"""T029 — School Excel reader for the minimal ingest layer.

Reads a school-issued ``.xlsx`` grade/attendance file and converts each
student row into ``CodexEntry`` instances (layer="minimal").  Fail-fast
on every boundary violation; no silent skip of malformed rows.

This is the canonical shared-reader convention: ``source_path`` is supplied
by the caller (a repo-relative string) so the resulting ``SourceRecord`` is
byte-identical across machines.  Subsequent source readers follow the same
signature.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from paideia_shared.schemas.metric_codex import CodexEntry, EntryKind, SourceRecord

from metric_codex.errors import LocatedInputError
from metric_codex.ingest.bronze_copies import SchoolExcelMap
from metric_codex.ingest.normalize import normalize_student_id
from metric_codex.ingest.result import SourceReadResult
from metric_codex.output.sha256 import compute_sha256

# Mapping from ColumnMap attribute name to its EntryKind equivalent.
_SCORE_FIELDS: list[tuple[str, EntryKind]] = [
    ("score_total", EntryKind.score_total),
    ("score_percent", EntryKind.score_percent),
    ("attendance", EntryKind.attendance),
]


def _select_sheet(
    wb: openpyxl.Workbook,
    sheet: str | int,
    filename: str,
) -> Worksheet | ReadOnlyWorksheet:
    """Select a worksheet by name or 0-based index.

    Args:
        wb: Loaded workbook.
        sheet: Sheet name (str) or 0-based index (int).
        filename: Source file name for error context.

    Returns:
        The selected worksheet.

    Raises:
        LocatedInputError: If a name is not found or an index is out of range.
    """
    if isinstance(sheet, str):
        if sheet not in wb.sheetnames:
            raise LocatedInputError(
                f"sheet '{sheet}' not found in workbook",
                file=filename,
            )
        return wb[sheet]

    # int path: 0-based index
    names = wb.sheetnames
    if sheet < 0 or sheet >= len(names):
        raise LocatedInputError(
            f"sheet index {sheet} is out of range (workbook has {len(names)} sheet(s))",
            file=filename,
        )
    return wb[names[sheet]]


def _find_column_indices(
    ws: Worksheet | ReadOnlyWorksheet,
    header_row: int,
    required_headers: list[str],
    filename: str,
) -> dict[str, int]:
    """Map each required header text to its 1-based column index.

    Args:
        ws: The active worksheet.
        header_row: 1-based row number of the header row.
        required_headers: Header texts that must all be present.
        filename: Source file name for error context.

    Returns:
        Dict mapping header text → 1-based column index.

    Raises:
        LocatedInputError: If the header row is beyond the populated range,
            a header text is duplicated, or any required header is absent.
    """
    # Fetch the header row.  ``next(gen, None)`` keeps an empty sheet (or a
    # header_row past the populated range) inside the LocatedInputError contract
    # instead of leaking a bare StopIteration.
    header_cells = next(
        ws.iter_rows(min_row=header_row, max_row=header_row),
        None,
    )
    if header_cells is None:
        raise LocatedInputError(
            f"header row {header_row} is beyond the sheet's populated range",
            file=filename,
            row=header_row,
        )

    # Build text → 1-based index, rejecting duplicate headers.
    found: dict[str, int] = {}
    for col_idx, cell in enumerate(header_cells, start=1):
        val = cell.value
        if val is None:
            continue
        text = str(val).strip()
        if text in found:
            raise LocatedInputError(
                f"duplicate column header '{text}' at columns {found[text]} and {col_idx}",
                file=filename,
                row=header_row,
            )
        found[text] = col_idx

    # Validate that every required header is present.
    for header in required_headers:
        if header not in found:
            raise LocatedInputError(
                f"required column header '{header}' not found in sheet",
                file=filename,
                row=header_row,
            )

    return {h: found[h] for h in required_headers}


def _coerce_score(
    raw: object,
    *,
    filename: str,
    row: int,
    column: str,
) -> float:
    """Coerce a cell value to float for a score/attendance column.

    Args:
        raw: The raw cell value from openpyxl.
        filename: Source file name for error location.
        row: 1-based row number for error location.
        column: Column header text for error location.

    Returns:
        The value as a float.

    Raises:
        LocatedInputError: If the value is a boolean or cannot be coerced to float.
    """
    # bool is an int subclass; float(True) == 1.0 would silently fabricate a
    # score.  Reject it explicitly before the float() coercion.
    if isinstance(raw, bool):
        raise LocatedInputError(
            "boolean is not a valid score/attendance value",
            file=filename,
            row=row,
            column=column,
            expected="numeric",
            actual=repr(raw),
        )
    try:
        return float(raw)  # type: ignore[arg-type]
    except (TypeError, ValueError) as exc:
        raise LocatedInputError(
            "non-numeric value in score/attendance column",
            file=filename,
            row=row,
            column=column,
            expected="numeric",
            actual=repr(raw),
        ) from exc


def read_school_excel(
    path: Path,
    excel_map: SchoolExcelMap,
    *,
    ingested_at: str,
    source_path: str,
) -> SourceReadResult:
    """Read a school Excel file into minimal-layer CodexEntry rows.

    Selects the configured sheet, finds columns by header text, and converts
    each data row into one or more ``CodexEntry`` instances (one per non-blank
    score/attendance column).  Empty score cells are silently skipped; every
    other boundary error raises ``LocatedInputError`` immediately.

    Args:
        path: Real filesystem path to the ``.xlsx`` file (used for opening and
            the SHA-256 digest only).
        excel_map: Parsed ``SchoolExcelMap`` config for this file.
        ingested_at: ISO-8601 UTC timestamp string to embed in the
            ``SourceRecord``.
        source_path: Repo-relative path string recorded verbatim in the
            ``SourceRecord.source_path`` field.  Caller-supplied so manifests
            are byte-identical across machines (independent of cwd).

    Returns:
        A ``SourceReadResult`` with ``entries`` sorted by
        ``(student_id, entry_kind, key)`` for determinism, the
        ``SourceRecord`` for this file, and an ``identities`` map of
        ``student_id → name_kr``.

    Raises:
        LocatedInputError: On any boundary violation — missing/out-of-range
            sheet, empty sheet, header row past the data, missing/duplicate
            header, invalid student_id, boolean/non-numeric score cell, or an
            invalid cohort_year.
    """
    filename = path.name
    cols = excel_map.columns

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = _select_sheet(wb, excel_map.sheet, filename)

        # Build the list of headers we need to locate.
        required_headers: list[str] = [cols.student_id]
        if cols.name_kr is not None:
            required_headers.append(cols.name_kr)
        for field, _ in _SCORE_FIELDS:
            header = getattr(cols, field)
            if header is not None:
                required_headers.append(header)
        if excel_map.cohort_year_column is not None:
            required_headers.append(excel_map.cohort_year_column)

        col_indices = _find_column_indices(
            ws, excel_map.header_row, required_headers, filename
        )

        source_id = f"school_excel:{filename}"
        entries: list[CodexEntry] = []
        identities: dict[str, str | None] = {}
        # Track which student_ids have been seen to detect duplicates (T030).
        seen_student_ids: set[str] = set()

        student_id_idx = col_indices[cols.student_id]

        # Iterate data rows: from header_row+1 onward.  ``row_num`` is computed
        # from the offset rather than ``cell.row`` because read-only EmptyCell
        # objects (e.g. a leading None cell) carry no ``.row`` attribute.
        first_data_row = excel_map.header_row + 1
        for offset, row_tuple in enumerate(ws.iter_rows(min_row=first_data_row)):
            # Skip entirely blank rows (e.g. trailing empty rows in the sheet).
            if all(cell.value is None for cell in row_tuple):
                continue

            row_num = first_data_row + offset  # 1-based

            # Normalize student_id — fail-fast at boundary (None included).
            raw_id = row_tuple[student_id_idx - 1].value
            student_id = normalize_student_id(
                raw_id,  # type: ignore[arg-type]
                source=filename,
                row=row_num,
            )

            # Fail-fast on duplicate student_id (last-write-wins would silently
            # discard conflicting scores — an audit-integrity violation).
            if student_id in seen_student_ids:
                raise LocatedInputError(
                    f"duplicate student_id {student_id!r} — each student must appear "
                    "exactly once per source file",
                    file=filename,
                    row=row_num,
                    column=cols.student_id,
                    expected="unique student_id per row",
                    actual=student_id,
                )
            seen_student_ids.add(student_id)

            # Capture identity.
            name_kr_val: str | None = None
            if cols.name_kr is not None:
                name_raw = row_tuple[col_indices[cols.name_kr] - 1].value
                name_kr_val = str(name_raw).strip() if name_raw is not None else None
            identities[student_id] = name_kr_val

            # Derive or read cohort_year.
            if excel_map.cohort_year_column is not None:
                raw_cy = row_tuple[col_indices[excel_map.cohort_year_column] - 1].value
                try:
                    cohort_year = int(float(raw_cy))  # type: ignore[arg-type]
                except (TypeError, ValueError) as exc:
                    raise LocatedInputError(
                        "non-numeric value in cohort_year column",
                        file=filename,
                        row=row_num,
                        column=excel_map.cohort_year_column,
                        expected="integer year (e.g. 2024)",
                        actual=repr(raw_cy),
                    ) from exc
                if not (2000 <= cohort_year <= 2100):
                    raise LocatedInputError(
                        f"cohort_year {cohort_year} out of range [2000, 2100]",
                        file=filename,
                        row=row_num,
                        column=excel_map.cohort_year_column,
                        expected="2000 <= cohort_year <= 2100",
                        actual=str(cohort_year),
                    )
            else:
                cohort_year = int(student_id[:4])
                if not (2000 <= cohort_year <= 2100):
                    raise LocatedInputError(
                        f"cohort_year {cohort_year} derived from id prefix is "
                        "out of range [2000, 2100]",
                        file=filename,
                        row=row_num,
                        column=cols.student_id,
                        expected="student id starting with a year in [2000, 2100]",
                        actual=student_id,
                    )

            # Emit one CodexEntry per non-blank score column.
            for field, entry_kind in _SCORE_FIELDS:
                header = getattr(cols, field)
                if header is None:
                    continue  # not configured for this source
                raw_score = row_tuple[col_indices[header] - 1].value
                if raw_score is None:
                    continue  # blank cell — skip this kind (not an error)

                value_num = _coerce_score(
                    raw_score,
                    filename=filename,
                    row=row_num,
                    column=header,
                )
                entries.append(
                    CodexEntry(
                        student_id=student_id,
                        semester=excel_map.semester,
                        cohort_year=cohort_year,
                        layer="minimal",
                        entry_kind=entry_kind,
                        key=entry_kind.value,
                        value_num=value_num,
                        value_text=None,
                        domain=None,
                        item_ref=None,
                        source_id=source_id,
                        observed_at=None,
                    )
                )

    finally:
        wb.close()

    # Deterministic sort: (student_id, entry_kind, key).
    entries.sort(key=lambda e: (e.student_id, e.entry_kind, e.key))

    source_record = SourceRecord(
        source_id=source_id,
        origin_module="school",
        origin_layer="bronze",
        source_path=source_path,
        sha256=compute_sha256(path),
        ingested_at=ingested_at,
    )

    return SourceReadResult(
        entries=entries,
        source_record=source_record,
        identities=identities,
    )


__all__ = ["read_school_excel"]
