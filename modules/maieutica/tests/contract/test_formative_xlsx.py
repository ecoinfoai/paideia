"""T042 — Contract test: LMS formative `.xlsx` (SC-003 / SC-009 roundtrip).

Writes a `.xlsx` from sample :class:`FormativeItemCandidate` objects via
:func:`maieutica.output.formative_xlsx.write_formative_xlsx`, reads it back with
openpyxl, and asserts the immutable bhu_text_mining contract
(``contracts/formative_xlsx.md``):

- single ``'Formative Test'`` sheet with 14 headers in exact order.
- M data rows mirroring the candidates.
- ``No.`` / ``Chapter`` cells are ``int``; every other cell is ``str``.
- ``Keywords`` is the keyword list joined by ``", "``.
- byte-identical across two writes of identical inputs (SC-009).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from maieutica.output.formative_xlsx import (
    FORMATIVE_HEADERS,
    formative_xlsx_filename,
    write_formative_xlsx,
)
from paideia_shared.schemas import FormativeItemCandidate

_EXPECTED_HEADERS = [
    "No.",
    "Chapter",
    "Topic",
    "Question",
    "Limit",
    "Model Answer",
    "Purpose",
    "Keywords",
    "Rubric(High)",
    "Rubric(Mid)",
    "Rubric(Low)",
    "Support(High)",
    "Support(Mid)",
    "Support(Low)",
]


def _candidate(no: int) -> FormativeItemCandidate:
    return FormativeItemCandidate(
        semester="2026-1",
        course_slug="anatomy",
        no=no,
        chapter_no=8,
        topic="개념이해",
        question=f"{no}번 문항: 허파꽈리의 가스교환 원리를 서술하시오.",
        limit="200자 내외",
        model_answer="허파꽈리는 제1형·제2형 허파세포로 덮여 있고 분압 차이로 확산한다.",
        purpose="허파꽈리 가스교환 원리 이해 확인.",
        keywords=["제1형 허파세포", "제2형 허파세포", "분압", "확산"],
        rubric_high="핵심 개념 전부 + 정확한 용어 사용.",
        rubric_mid="핵심 개념 일부 포함, 용어 부정확.",
        rubric_low="부분적 언급 또는 핵심 누락.",
        support_high="신생아 호흡곤란증후군과 연결하여 표면활성제 임상 중요성으로 도약.",
        support_mid="허파꽈리 구조 그림을 참고하여 복습하도록 지도.",
        support_low="허파꽈리 기본 구조부터 1:1로 재학습 안내.",
    )


def test_formative_xlsx_filename_contract() -> None:
    """The filename follows ``Ch{NN}_{chapter}_FormativeTest.xlsx``."""
    assert formative_xlsx_filename(8, "호흡계통") == "Ch08_호흡계통_FormativeTest.xlsx"


def test_formative_headers_match_real_sample() -> None:
    """The 14 headers match the real bronze sample verbatim and in order."""
    assert list(FORMATIVE_HEADERS) == _EXPECTED_HEADERS
    assert len(FORMATIVE_HEADERS) == 14


def test_formative_xlsx_roundtrip_structure_and_cell_types(tmp_path: Path) -> None:
    """SC-003: write → openpyxl read-back → assert headers, cell types, join."""
    candidates = [_candidate(1), _candidate(2), _candidate(3)]
    out = tmp_path / "Ch08_호흡계통_FormativeTest.xlsx"
    write_formative_xlsx(out, candidates)

    workbook = openpyxl.load_workbook(out)
    assert workbook.sheetnames == ["Formative Test"]
    sheet = workbook.active

    # Header row: 14 headers in exact order.
    header = [c.value for c in sheet[1]]
    assert header == list(FORMATIVE_HEADERS)
    assert len(header) == 14

    # M data rows.
    assert sheet.max_row == 1 + len(candidates)

    col = {h: i + 1 for i, h in enumerate(FORMATIVE_HEADERS)}

    for i, cand in enumerate(candidates):
        row = i + 2
        # No. / Chapter → int cells.
        no_cell = sheet.cell(row=row, column=col["No."])
        chapter_cell = sheet.cell(row=row, column=col["Chapter"])
        assert isinstance(no_cell.value, int)
        assert no_cell.value == cand.no
        assert isinstance(chapter_cell.value, int)
        assert chapter_cell.value == cand.chapter_no

        # Keywords → joined by ", ".
        kw_cell = sheet.cell(row=row, column=col["Keywords"])
        assert isinstance(kw_cell.value, str)
        assert kw_cell.value == ", ".join(cand.keywords)

        # All non-numeric columns → str.
        for name in (
            "Topic",
            "Question",
            "Limit",
            "Model Answer",
            "Purpose",
            "Keywords",
            "Rubric(High)",
            "Rubric(Mid)",
            "Rubric(Low)",
            "Support(High)",
            "Support(Mid)",
            "Support(Low)",
        ):
            assert isinstance(sheet.cell(row=row, column=col[name]).value, str)


def test_formative_xlsx_byte_identical_on_two_writes(tmp_path: Path) -> None:
    """SC-009: two writes of identical inputs produce byte-identical files."""
    candidates = [_candidate(1), _candidate(2)]
    out1 = tmp_path / "first.xlsx"
    out2 = tmp_path / "second.xlsx"
    write_formative_xlsx(out1, candidates)
    write_formative_xlsx(out2, candidates)
    assert out1.read_bytes() == out2.read_bytes()
