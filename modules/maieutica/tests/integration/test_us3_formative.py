"""T043 — Integration test: formative build path (US3).

Exercises ``maieutica.pipeline.build`` end-to-end and asserts the formative
``.xlsx`` output alongside the quiz ``.xls`` (US1 fixture pattern reused):

1. Build a temp Bronze tree with a chapter ``.txt`` whose body contains the
   canned formative item's ``topic`` term (so groundedness resolves to 확인).
2. Pre-fill ``responses/{slot_id}.json`` for every quiz AND formative slot.
3. Call ``build(...)`` with a ``SubscriptionBackend``.

Assertions (SC-003 / SC-009 / FR-014):
- ``Ch{NN}_{chapter}_FormativeTest.xlsx`` exists with ``formative_count`` rows.
- rubric High/Mid/Low + support High/Mid/Low + keywords + model_answer filled.
- ``support_high`` is the leap/도약 axis.
- every formative item is anchored (확인) or 미확인; the fixture arranges 확인.
- the quiz ``.xls`` is STILL emitted (both files coexist).
- everything under ``tmp_path`` — no real ``data/`` touched.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from maieutica.ingest.spec_load import load_curriculum_map, load_generation_spec
from maieutica.plan.slots import plan_slots
from paideia_shared.schemas import MaieuticaGenerationSpec

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_WEEK = 9
_CHAPTER_NO = 8
_CHAPTER = "8장 호흡계통"
_QUIZ_COUNT = 2
_FORMATIVE_COUNT = 2

# Quiz key_concepts (must appear in the chapter body for 확인).
_KEY_CONCEPTS = ["폐포", "기관지"]
# Formative topics (must appear in the chapter body for 확인).
_FORMATIVE_TOPICS = ["가스교환", "가로막"]

_CHAPTER_TXT = "\n".join(
    [
        "8장 호흡계통",
        "",
        "1. 호흡계통의 구조",
        "코는 후각과 공기 가습을 담당한다.",
        "폐포는 가스 교환이 일어나는 포상 구조이다.",
        "기관지는 공기를 폐로 전달하는 통로이다.",
        "허파꽈리에서 가스교환이 분압 차이로 일어난다.",
        "가로막은 수축하여 흉강 부피를 늘린다.",
        "",
    ]
)


def _quiz_item_json(item_no: int, key_concept: str, answer_no: int) -> dict:
    options = [
        f"{marker} {key_concept} 관련 보기 {item_no}-{i} 충분한 길이를 가진 보기입니다 abcde"
        for i, marker in enumerate("①②③④⑤", start=1)
    ]
    return {
        "question_type": "지식축적",
        "stem_polarity": "부정형",
        "text": f"{item_no}번 문제: {key_concept}에 대해 옳지 않은 것은?",
        "options": options,
        "answer_no": answer_no,
        "option_evidence": [f"{key_concept} 근거 {i}" for i in range(1, 6)],
        "wrong_explanation": f"{key_concept} 관련 오답 설명입니다.",
        "leap_explanation": f"{key_concept} 다음 개념으로의 도약 설명입니다.",
        "key_concept": key_concept,
        "section": "1. 호흡계통의 구조",
    }


def _formative_item_json(no: int, topic: str) -> dict:
    return {
        "no": no,
        "chapter_no": _CHAPTER_NO,
        "topic": topic,
        "question": f"{no}번 형성문항: {topic} 원리를 200자 내외로 서술하시오.",
        "limit": "200자 내외",
        "model_answer": f"{topic}는 교재에 따르면 핵심 과정이다.",
        "purpose": f"{topic} 이해 여부 확인.",
        "keywords": [topic, "분압", "확산"],
        "rubric_high": f"{topic} 핵심 개념 전부 + 정확한 용어.",
        "rubric_mid": f"{topic} 일부 포함, 용어 부정확.",
        "rubric_low": f"{topic} 핵심 누락.",
        "support_high": f"{topic}를 다음 심화 개념으로 잇는 도약 활동 안내.",
        "support_mid": f"{topic} 관련 교재 그림으로 복습 지도.",
        "support_low": f"{topic} 기본 개념부터 재학습 경로 안내.",
    }


def _spec() -> MaieuticaGenerationSpec:
    return MaieuticaGenerationSpec.model_validate(
        {
            "semester": _SEMESTER,
            "course_slug": _COURSE,
            "week": _WEEK,
            "chapter_no": _CHAPTER_NO,
            "chapter": _CHAPTER,
            "quiz_count": _QUIZ_COUNT,
            "formative_count": _FORMATIVE_COUNT,
        }
    )


def _build_bronze(tmp_path: Path) -> tuple[Path, Path]:
    data_root = tmp_path / "data"
    bronze = data_root / "bronze" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    bronze.mkdir(parents=True, exist_ok=True)

    (bronze / f"{_CHAPTER} 호흡.txt").write_text(_CHAPTER_TXT, encoding="utf-8")

    spec = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "week": _WEEK,
        "chapter_no": _CHAPTER_NO,
        "chapter": _CHAPTER,
        "quiz_count": _QUIZ_COUNT,
        "formative_count": _FORMATIVE_COUNT,
    }
    (bronze / "generation_spec.yaml").write_text(
        json.dumps(spec, ensure_ascii=False), encoding="utf-8"
    )

    curriculum = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "entries": [
            {
                "week": _WEEK,
                "chapter": _CHAPTER,
                "chapter_no": _CHAPTER_NO,
                "sections": ["1. 호흡계통의 구조"],
            }
        ],
    }
    (bronze / "curriculum_map.yaml").write_text(
        json.dumps(curriculum, ensure_ascii=False), encoding="utf-8"
    )
    return bronze, data_root


def _write_canned_responses(responses_dir: Path) -> None:
    responses_dir.mkdir(parents=True, exist_ok=True)
    slots = plan_slots(_spec())

    quiz_slots = [s for s in slots if s.kind == "quiz"]
    for idx, slot in enumerate(quiz_slots):
        key_concept = _KEY_CONCEPTS[idx % len(_KEY_CONCEPTS)]
        item_json = _quiz_item_json(slot.ordinal, key_concept, (idx % 5) + 1)
        _write_envelope(responses_dir, slot.slot_id, item_json)

    formative_slots = [s for s in slots if s.kind == "formative"]
    for idx, slot in enumerate(formative_slots):
        topic = _FORMATIVE_TOPICS[idx % len(_FORMATIVE_TOPICS)]
        item_json = _formative_item_json(slot.ordinal, topic)
        _write_envelope(responses_dir, slot.slot_id, item_json)


def _write_envelope(responses_dir: Path, slot_id: str, item_json: dict) -> None:
    envelope = {
        "slot_id": slot_id,
        "raw_text": json.dumps(item_json, ensure_ascii=False),
        "model": "canned-subscription",
    }
    (responses_dir / f"{slot_id}.json").write_text(
        json.dumps(envelope, ensure_ascii=False), encoding="utf-8"
    )


def test_us3_formative_build_end_to_end(tmp_path: Path) -> None:
    """Build → formative .xlsx with formative_count anchored rows + quiz .xls."""
    from maieutica.generate.backend import SubscriptionBackend
    from maieutica.output.formative_xlsx import FORMATIVE_HEADERS
    from maieutica.pipeline import build

    bronze, data_root = _build_bronze(tmp_path)
    spec = load_generation_spec(bronze / "generation_spec.yaml")
    curriculum_map = load_curriculum_map(bronze / "curriculum_map.yaml")

    silver = data_root / "silver" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    staging_dir = silver / "staging"
    responses_dir = silver / "responses"
    _write_canned_responses(responses_dir)

    backend = SubscriptionBackend(staging_dir=staging_dir, responses_dir=responses_dir)

    _items, run_dir = build(
        spec=spec,
        curriculum_map=curriculum_map,
        bronze_dir=bronze,
        data_root=data_root,
        backend=backend,
        generation_spec_path=bronze / "generation_spec.yaml",
        curriculum_map_path=bronze / "curriculum_map.yaml",
    )

    # Both files coexist: quiz .xls AND formative .xlsx.
    xls_path = run_dir / f"QuestionUploadExcel_{_WEEK}주차.xls"
    assert xls_path.is_file(), f"expected quiz xls at {xls_path}"

    formative_path = run_dir / f"Ch{_CHAPTER_NO:02d}_{_CHAPTER}_FormativeTest.xlsx"
    assert formative_path.is_file(), f"expected formative xlsx at {formative_path}"

    workbook = openpyxl.load_workbook(formative_path)
    sheet = workbook.active
    header = [c.value for c in sheet[1]]
    assert header == list(FORMATIVE_HEADERS)
    # formative_count data rows.
    assert sheet.max_row == 1 + _FORMATIVE_COUNT

    col = {h: i + 1 for i, h in enumerate(FORMATIVE_HEADERS)}
    for row in range(2, 2 + _FORMATIVE_COUNT):
        # rubric High/Mid/Low + support High/Mid/Low + keywords + model_answer filled.
        for name in (
            "Model Answer",
            "Keywords",
            "Rubric(High)",
            "Rubric(Mid)",
            "Rubric(Low)",
            "Support(High)",
            "Support(Mid)",
            "Support(Low)",
        ):
            value = sheet.cell(row=row, column=col[name]).value
            assert isinstance(value, str) and value, f"{name} empty at row {row}"
        # support_high is the leap/도약 axis.
        assert "도약" in sheet.cell(row=row, column=col["Support(High)"]).value

    # Manifest formative_count reflects the formative items.
    manifest = json.loads(
        (run_dir / "manifest_maieutica.json").read_text(encoding="utf-8")
    )
    assert manifest["formative_count"] == _FORMATIVE_COUNT

    # No real data/ dir touched.
    assert str(run_dir).startswith(str(tmp_path))


def test_us3_formative_items_anchored(tmp_path: Path) -> None:
    """Every formative item's textbook_evidence status is set (확인 here)."""
    from maieutica.generate.backend import SubscriptionBackend
    from maieutica.generate.formative_gen import generate_formative_item
    from maieutica.silver.chunk import chunk_chapter
    from maieutica.silver.evidence_index import EvidenceIndex
    from maieutica.verify.groundedness import ground_formative

    bronze, _data_root = _build_bronze(tmp_path)
    spec = load_generation_spec(bronze / "generation_spec.yaml")
    responses_dir = tmp_path / "responses"
    staging_dir = tmp_path / "staging"
    _write_canned_responses(responses_dir)

    from maieutica.generate.backend import InputHashCache

    backend = SubscriptionBackend(staging_dir=staging_dir, responses_dir=responses_dir)
    cache = InputHashCache(backend=backend, cache_dir=tmp_path / "cache")

    raw_lines = _CHAPTER_TXT.split("\n")
    chunks = chunk_chapter(
        raw_lines,
        chapter_no=_CHAPTER_NO,
        chapter=_CHAPTER,
        semester=_SEMESTER,
        course_slug=_COURSE,
        source_file="ch.txt",
    )
    index = EvidenceIndex.from_chapter(raw_lines, chunks=chunks, source_file="ch.txt")

    formative_slots = [s for s in plan_slots(spec) if s.kind == "formative"]
    for slot in formative_slots:
        item = generate_formative_item(slot, spec, chunks, cache)
        item = ground_formative(item, index)
        assert item.textbook_evidence is not None
        assert item.textbook_evidence.status in ("확인", "미확인")
        # The fixture topics appear in the chapter body → 확인.
        assert item.textbook_evidence.status == "확인"
        # support_high is the leap axis (FR-014).
        assert "도약" in item.support_high
