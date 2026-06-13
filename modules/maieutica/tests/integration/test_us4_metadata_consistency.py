"""T047 — Integration test: US4 full nested yaml + quality report + consistency.

Exercises the extended ``build`` pipeline end-to-end and asserts:
(a) ``출제후보_완전판.yaml`` has both ``quiz`` and ``formative`` top-level keys
    with full metadata fields populated (all FR-015 fields present).
(b) The flat ``.xls``/``.xlsx`` files and the nested yaml hold the same
    candidate set — same counts, same item numbers, answers match.
(c) ``출제품질리포트.md`` exists with the expected sections.
(d) ``adoption_status`` is present (default ``생성``) in both yaml sections.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import xlrd
import yaml
from maieutica.ingest.spec_load import load_curriculum_map, load_generation_spec
from maieutica.plan.slots import plan_slots
from paideia_shared.schemas import MaieuticaGenerationSpec

# ---------------------------------------------------------------------------
# Fixture constants (reuse US1/US3 pattern)
# ---------------------------------------------------------------------------

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_WEEK = 9
_CHAPTER_NO = 8
_CHAPTER = "8장 호흡계통"
_QUIZ_COUNT = 3
_FORMATIVE_COUNT = 2

_KEY_CONCEPTS = ["폐포", "기관지", "가로막"]
_FORMATIVE_TOPICS = ["가스교환", "가로막"]

# Verbatim subsection line per key_concept — the correct option's evidence must
# be a verbatim line of its assigned subsection for the answer-anchored
# groundedness check (US1) to confirm it (else it is 미확인-excluded in US3).
_CONCEPT_EVIDENCE = {
    "폐포": "폐포는 가스 교환이 일어나는 포상 구조이다.",
    "기관지": "기관지는 공기를 폐로 전달하는 통로이다.",
    "가로막": "가로막은 수축하여 흉강 부피를 늘린다.",
}

_CHAPTER_TXT = "\n".join(
    [
        "8장 호흡계통",
        "",
        "1. 호흡계통의 구조",
        "코는 후각과 공기 가습을 담당한다.",
        "폐포는 가스 교환이 일어나는 포상 구조이다.",
        "기관지는 공기를 폐로 전달하는 통로이다.",
        "가로막은 수축하여 흉강 부피를 늘린다.",
        "허파꽈리에서 가스교환이 분압 차이로 일어난다.",
        "",
    ]
)

# Quality report section headings that must be present.
_EXPECTED_SECTIONS = [
    "# 출제품질리포트",
    "## 총 문항 수",
    "## 정답 번호 분포",
    "## 줄기 극성 분포",
    "## 난이도 분포",
    "## 교재 근거 확인",
    "## 형성평가 요약",
    "## 주의 항목",
]


def _quiz_item_json(item_no: int, key_concept: str, answer_no: int) -> dict:
    options = [
        f"{marker} {key_concept} 관련 보기 {item_no}-{i} 충분한 길이를 가진 보기입니다 abcde"
        for i, marker in enumerate("①②③④⑤", start=1)
    ]
    return {
        "question_type": "지식축적",
        "stem_polarity": "부정형",
        "text": f"{item_no}번 문제: {key_concept}에 대해 옳지 않은 것은?",
        "options": options,
        "answer_no": answer_no,
        # Correct option's evidence = verbatim subsection line → status="확인".
        "option_evidence": [
            _CONCEPT_EVIDENCE[key_concept] if i == answer_no else f"{key_concept} 근거 {i}"
            for i in range(1, 6)
        ],
        "wrong_explanation": f"{key_concept} 관련 오답 설명입니다.",
        "leap_explanation": f"{key_concept} 다음 개념으로의 도약 설명입니다.",
        "key_concept": key_concept,
        "section": "1. 호흡계통의 구조",
    }


def _formative_item_json(no: int, topic: str) -> dict:
    return {
        "no": no,
        "chapter_no": _CHAPTER_NO,
        "topic": topic,
        "question": f"{no}번 형성문항: {topic} 원리를 200자 내외로 서술하시오.",
        "limit": "200자 내외",
        "model_answer": f"{topic}는 교재에 따르면 핵심 과정이다.",
        "purpose": f"{topic} 이해 여부 확인.",
        "keywords": [topic, "분압", "확산"],
        "rubric_high": f"{topic} 핵심 개념 전부 + 정확한 용어.",
        "rubric_mid": f"{topic} 일부 포함, 용어 부정확.",
        "rubric_low": f"{topic} 핵심 누락.",
        "support_high": f"{topic}를 다음 심화 개념으로 잇는 도약 활동 안내.",
        "support_mid": f"{topic} 관련 교재 그림으로 복습 지도.",
        "support_low": f"{topic} 기본 개념부터 재학습 경로 안내.",
    }


def _spec() -> MaieuticaGenerationSpec:
    return MaieuticaGenerationSpec.model_validate(
        {
            "semester": _SEMESTER,
            "course_slug": _COURSE,
            "week": _WEEK,
            "chapter_no": _CHAPTER_NO,
            "chapter": _CHAPTER,
            "quiz_count": _QUIZ_COUNT,
            "formative_count": _FORMATIVE_COUNT,
        }
    )


def _build_bronze(tmp_path: Path) -> tuple[Path, Path]:
    data_root = tmp_path / "data"
    bronze = data_root / "bronze" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    bronze.mkdir(parents=True, exist_ok=True)

    (bronze / f"{_CHAPTER} 호흡.txt").write_text(_CHAPTER_TXT, encoding="utf-8")

    spec = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "week": _WEEK,
        "chapter_no": _CHAPTER_NO,
        "chapter": _CHAPTER,
        "quiz_count": _QUIZ_COUNT,
        "formative_count": _FORMATIVE_COUNT,
    }
    (bronze / "generation_spec.yaml").write_text(
        json.dumps(spec, ensure_ascii=False), encoding="utf-8"
    )

    curriculum = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "entries": [
            {
                "week": _WEEK,
                "chapter": _CHAPTER,
                "chapter_no": _CHAPTER_NO,
                "sections": ["1. 호흡계통의 구조"],
            }
        ],
    }
    (bronze / "curriculum_map.yaml").write_text(
        json.dumps(curriculum, ensure_ascii=False), encoding="utf-8"
    )
    return bronze, data_root


def _write_canned_responses(responses_dir: Path) -> None:
    responses_dir.mkdir(parents=True, exist_ok=True)
    slots = plan_slots(_spec())

    quiz_slots = [s for s in slots if s.kind == "quiz"]
    for idx, slot in enumerate(quiz_slots):
        key_concept = _KEY_CONCEPTS[idx % len(_KEY_CONCEPTS)]
        item_json = _quiz_item_json(slot.ordinal, key_concept, (idx % 5) + 1)
        _write_envelope(responses_dir, slot.slot_id, item_json)

    formative_slots = [s for s in slots if s.kind == "formative"]
    for idx, slot in enumerate(formative_slots):
        topic = _FORMATIVE_TOPICS[idx % len(_FORMATIVE_TOPICS)]
        item_json = _formative_item_json(slot.ordinal, topic)
        _write_envelope(responses_dir, slot.slot_id, item_json)


def _write_envelope(responses_dir: Path, slot_id: str, item_json: dict) -> None:
    envelope = {
        "slot_id": slot_id,
        "raw_text": json.dumps(item_json, ensure_ascii=False),
        "model": "canned-subscription",
    }
    (responses_dir / f"{slot_id}.json").write_text(
        json.dumps(envelope, ensure_ascii=False), encoding="utf-8"
    )


def test_us4_full_nested_yaml_and_quality_report(tmp_path: Path) -> None:
    """Build → assert yaml has quiz+formative keys, quality report exists, consistency holds."""
    from maieutica.generate.backend import SubscriptionBackend
    from maieutica.pipeline import build

    bronze, data_root = _build_bronze(tmp_path)
    spec = load_generation_spec(bronze / "generation_spec.yaml")
    curriculum_map = load_curriculum_map(bronze / "curriculum_map.yaml")

    silver = data_root / "silver" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    staging_dir = silver / "staging"
    responses_dir = silver / "responses"
    _write_canned_responses(responses_dir)

    backend = SubscriptionBackend(staging_dir=staging_dir, responses_dir=responses_dir)

    _items, run_dir = build(
        spec=spec,
        curriculum_map=curriculum_map,
        bronze_dir=bronze,
        data_root=data_root,
        backend=backend,
        generation_spec_path=bronze / "generation_spec.yaml",
        curriculum_map_path=bronze / "curriculum_map.yaml",
    )

    # --- (a) yaml has both keys with full metadata fields ---
    yaml_path = run_dir / "출제후보_완전판.yaml"
    assert yaml_path.is_file(), f"expected yaml at {yaml_path}"

    doc = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
    assert isinstance(doc, dict), "top-level must be a mapping"
    assert "quiz" in doc, "yaml missing 'quiz' key"
    assert "formative" in doc, "yaml missing 'formative' key"

    # quiz list: full metadata present
    assert len(doc["quiz"]) == _QUIZ_COUNT
    for q in doc["quiz"]:
        for field in (
            "item_no",
            "text",
            "options",
            "answer_no",
            "option_evidence",
            "textbook_evidence",
            "leap",
            "key_concept",
            "question_type",
            "difficulty",
            "review_note",
            "adoption_status",
        ):
            assert field in q, f"quiz item missing field: {field}"
        # leap has nested textbook_evidence key
        assert "textbook_evidence" in q["leap"]

    # formative list: full metadata present
    assert len(doc["formative"]) == _FORMATIVE_COUNT
    for f in doc["formative"]:
        for field in (
            "no",
            "topic",
            "textbook_evidence",
            "review_note",
            "adoption_status",
            "rubric_high",
            "rubric_mid",
            "rubric_low",
            "support_high",
            "support_mid",
            "support_low",
            "keywords",
        ):
            assert field in f, f"formative item missing field: {field}"

    # --- (b) flat files and yaml hold the same candidate set ---
    xls_path = run_dir / f"QuestionUploadExcel_{_WEEK}주차.xls"
    assert xls_path.is_file()

    # quiz: item_no + answer_no must match between flat xls and nested yaml
    book = xlrd.open_workbook(str(xls_path))
    sheet1 = book.sheet_by_index(1)
    assert sheet1.nrows == 1 + _QUIZ_COUNT, "xls row count mismatch"

    # xls column layout: col0=문제번호 (numeric), col8=답안 (text str of answer_no)
    xls_item_nos = sorted(int(sheet1.cell(r, 0).value) for r in range(1, sheet1.nrows))
    xls_answers = {
        int(sheet1.cell(r, 0).value): sheet1.cell(r, 8).value
        for r in range(1, sheet1.nrows)
    }
    yaml_quiz_item_nos = sorted(q["item_no"] for q in doc["quiz"])
    assert xls_item_nos == yaml_quiz_item_nos, (
        f"xls item_nos {xls_item_nos} != yaml quiz item_nos {yaml_quiz_item_nos}"
    )
    for q in doc["quiz"]:
        ino = q["item_no"]
        expected_ans = str(q["answer_no"])
        assert xls_answers[ino] == expected_ans, (
            f"answer mismatch at item_no={ino}: xls={xls_answers[ino]!r}, "
            f"yaml={expected_ans!r}"
        )

    # formative: no must match between flat xlsx and nested yaml
    formative_path = run_dir / f"Ch{_CHAPTER_NO:02d}_{_CHAPTER}_FormativeTest.xlsx"
    assert formative_path.is_file()
    wb = openpyxl.load_workbook(formative_path)
    ws = wb.active
    xlsx_nos = sorted(int(ws.cell(r, 1).value) for r in range(2, ws.max_row + 1))
    yaml_formative_nos = sorted(f["no"] for f in doc["formative"])
    assert xlsx_nos == yaml_formative_nos, (
        f"formative no mismatch: xlsx={xlsx_nos}, yaml={yaml_formative_nos}"
    )

    # --- (c) quality report exists with expected sections ---
    report_path = run_dir / "출제품질리포트.md"
    assert report_path.is_file(), f"expected quality report at {report_path}"
    report_text = report_path.read_text(encoding="utf-8")
    for section in _EXPECTED_SECTIONS:
        assert section in report_text, f"quality report missing section: {section!r}"

    # --- (d) adoption_status present (default 생성) in both sections ---
    for q in doc["quiz"]:
        assert q["adoption_status"] == "생성", (
            f"quiz item {q['item_no']} adoption_status != '생성'"
        )
    for f in doc["formative"]:
        assert f["adoption_status"] == "생성", (
            f"formative item {f['no']} adoption_status != '생성'"
        )

    # No real data/ dir touched.
    assert str(run_dir).startswith(str(tmp_path))


def test_us4_flat_files_byte_identical_after_yaml_addition(tmp_path: Path) -> None:
    """Yaml/report additions must NOT perturb the .xls and .xlsx outputs.

    Run the build twice with identical inputs and verify the .xls and
    .xlsx files are byte-identical across both runs (yaml/report are
    separate files, they do not re-enter the xls/xlsx writers).
    """
    from maieutica.generate.backend import SubscriptionBackend
    from maieutica.pipeline import build

    bronze, data_root = _build_bronze(tmp_path)
    spec = load_generation_spec(bronze / "generation_spec.yaml")
    curriculum_map = load_curriculum_map(bronze / "curriculum_map.yaml")

    silver = data_root / "silver" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    staging_dir = silver / "staging"
    responses_dir = silver / "responses"
    _write_canned_responses(responses_dir)

    backend = SubscriptionBackend(staging_dir=staging_dir, responses_dir=responses_dir)

    kwargs = {
        "spec": spec,
        "curriculum_map": curriculum_map,
        "bronze_dir": bronze,
        "data_root": data_root,
        "backend": backend,
        "generation_spec_path": bronze / "generation_spec.yaml",
        "curriculum_map_path": bronze / "curriculum_map.yaml",
    }

    _, run_dir1 = build(**kwargs)  # type: ignore[arg-type]
    xls1 = (run_dir1 / f"QuestionUploadExcel_{_WEEK}주차.xls").read_bytes()
    xlsx1 = (
        run_dir1 / f"Ch{_CHAPTER_NO:02d}_{_CHAPTER}_FormativeTest.xlsx"
    ).read_bytes()
    yaml1 = (run_dir1 / "출제후보_완전판.yaml").read_bytes()
    report1 = (run_dir1 / "출제품질리포트.md").read_bytes()

    _, run_dir2 = build(**kwargs)  # type: ignore[arg-type]
    # run_id is deterministic: same run_dir
    assert run_dir1 == run_dir2

    xls2 = (run_dir2 / f"QuestionUploadExcel_{_WEEK}주차.xls").read_bytes()
    xlsx2 = (
        run_dir2 / f"Ch{_CHAPTER_NO:02d}_{_CHAPTER}_FormativeTest.xlsx"
    ).read_bytes()
    yaml2 = (run_dir2 / "출제후보_완전판.yaml").read_bytes()
    report2 = (run_dir2 / "출제품질리포트.md").read_bytes()

    assert xls1 == xls2, "quiz .xls is not byte-identical across runs"
    assert xlsx1 == xlsx2, "formative .xlsx is not byte-identical across runs"
    assert yaml1 == yaml2, "candidate yaml is not byte-identical across runs"
    assert report1 == report2, "quality report is not byte-identical across runs"
