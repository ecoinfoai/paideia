"""T031 — Determinism property/regression for the v0.1.1 pipeline (R10/SC-004/FR-013).

Exercises ``maieutica.pipeline.build`` twice on a synthetic multi-subsection
chapter with canned ``SubscriptionBackend`` responses (every item resolves 확인
with a DISTINCT verbatim subsection line, mirroring ``test_us1_diversity``), and
asserts the v0.1.1 determinism contract:

- **SC-004 byte-identical** — two builds with identical inputs into SEPARATE
  data roots produce byte-identical ``.xls``, ``.xlsx`` and ``출제후보_완전판.yaml``.
  The manifest is excluded (its ``generated_at`` is the lone non-deterministic
  field).
- **Cache hit → 0 LLM re-calls under serial avoid_list (R10)** — a SECOND build
  against the SAME data root (persisted ``cache/`` dir) serves every slot from
  the InputHashCache, so the second run's manifest ``cache_hit_rate == 1.0``.
  This proves the serially-accumulated avoid_list yields a stable cache key.
- **Stable sort under subsection shuffle** — the adopted items' subsection
  grouping (anchor ``chunk_id`` → count) is identical across the two separate
  builds.
- **xls/xlsx roundtrip cell types + zero-padding (FR-013)** — the produced
  ``.xls`` honours the immutable LMS cell-type contract (답안 TEXT "n",
  문제번호 NUMBER, 예상주차 zero-padded TEXT "00W", 문항유형 TEXT "002") and the
  ``.xlsx`` honours the bhu_text_mining contract (No./Chapter int, rest str).

All I/O is under ``tmp_path`` — no real ``data/`` directory is touched.
"""

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import openpyxl
import xlrd
from maieutica.generate.backend import SubscriptionBackend
from maieutica.ingest.spec_load import load_curriculum_map, load_generation_spec
from maieutica.output.formative_xlsx import FORMATIVE_HEADERS, formative_xlsx_filename
from maieutica.output.quiz_xls import QUIZ_HEADERS, guide_sheet_name
from maieutica.pipeline import build
from maieutica.plan.slots import assign_subsections, plan_slots
from maieutica.silver.chunk import chunk_chapter
from paideia_shared.schemas import MaieuticaGenerationSpec

# ---------------------------------------------------------------------------
# Fixture content — same multi-subsection chapter shape as test_us1_diversity.
# ---------------------------------------------------------------------------

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_WEEK = 9
_CHAPTER_NO = 8
_CHAPTER = "8장 호흡계통"
_QUIZ_COUNT = 12
_FORMATIVE_COUNT = 1

_CHAPTER_TXT = "\n".join(
    [
        "8장 호흡계통",
        "",
        "1. 코의 구조",
        "코는 들이마신 공기를 데우고 가습하는 첫 관문이다.",
        "코털과 점막은 먼지와 이물질을 거른다.",
        "코안은 비중격으로 좌우가 나뉜다.",
        "",
        "2. 인두와 후두",
        "인두는 공기와 음식이 함께 지나가는 통로이다.",
        "후두는 발성을 담당하며 기도를 보호한다.",
        "성대는 공기 흐름으로 진동하여 소리를 만든다.",
        "",
        "3. 기관과 기관지",
        "기관은 후두에서 갈라져 좌우 폐로 이어진다.",
        "기관지는 공기를 좌우 폐로 나누어 전달하는 통로이다.",
        "기관지는 점점 가늘어져 세기관지가 된다.",
        "",
        "4. 폐포와 가스 교환",
        "폐포는 모세혈관과 맞닿아 가스를 교환한다.",
        "폐포에서는 산소가 혈액으로 들어가고 이산화탄소가 나온다.",
        "이 교환은 분압 차이에 따른 단순 확산으로 일어난다.",
        "",
        "5. 호흡의 조절",
        "가로막은 수축하여 흉강의 부피를 늘리고 압력을 낮춘다.",
        "숨뇌의 호흡중추가 기본 리듬을 만든다.",
        "화학수용체가 이산화탄소 농도를 감지해 호흡수를 바꾼다.",
        "",
    ]
)


def _spec() -> MaieuticaGenerationSpec:
    """Build the validated generation spec for this fixture chapter."""
    return MaieuticaGenerationSpec.model_validate(
        {
            "semester": _SEMESTER,
            "course_slug": _COURSE,
            "week": _WEEK,
            "chapter_no": _CHAPTER_NO,
            "chapter": _CHAPTER,
            "quiz_count": _QUIZ_COUNT,
            "formative_count": _FORMATIVE_COUNT,
        }
    )


def _quiz_item_json(item_no: int, evidence_line: str, answer_no: int) -> dict:
    """Schema-valid quiz item whose CORRECT option's evidence is a verbatim line."""
    options = [
        f"{marker} 보기 {item_no}-{i} 충분한 길이를 가진 보기 문장입니다 abcde"
        for i, marker in enumerate("①②③④⑤", start=1)
    ]
    option_evidence = [f"보기 {item_no}-{i} 근거" for i in range(1, 6)]
    option_evidence[answer_no - 1] = evidence_line
    return {
        "question_type": "지식축적",
        "stem_polarity": "부정형",
        "text": f"{item_no}번 문제: 다음 중 옳지 않은 것은?",
        "options": options,
        "answer_no": answer_no,
        "option_evidence": option_evidence,
        "wrong_explanation": f"{item_no}번 문항의 오답 설명입니다.",
        "leap_explanation": f"{item_no}번 다음 개념으로의 도약 설명입니다.",
        "key_concept": f"개념-{item_no}",
        "section": f"섹션-{item_no}",
    }


def _formative_item_json(no: int) -> dict:
    """Minimal schema-valid formative item JSON for the canned response."""
    topic = "폐포"
    return {
        "no": no,
        "chapter_no": _CHAPTER_NO,
        "topic": topic,
        "question": f"{no}번 형성문항: {topic} 원리를 200자 내외로 서술하시오.",
        "limit": "200자 내외",
        "model_answer": f"{topic}는 교재에 따르면 핵심 과정이다.",
        "purpose": f"{topic} 이해 여부 확인.",
        "keywords": [topic, "분압", "확산"],
        "rubric_high": f"{topic} 핵심 개념 전부.",
        "rubric_mid": f"{topic} 일부 포함.",
        "rubric_low": f"{topic} 핵심 누락.",
        "support_high": f"{topic}를 다음 심화 개념으로 잇는 도약 활동.",
        "support_mid": f"{topic} 복습 지도.",
        "support_low": f"{topic} 기본 재학습 안내.",
    }


def _subsection_body_lines(chunk_text: str) -> list[str]:
    """Return the non-heading body lines of a chunk's text (verbatim anchors)."""
    lines = [ln for ln in chunk_text.splitlines() if ln.strip()]
    return lines[1:] if len(lines) > 1 else lines


def _build_bronze(root: Path) -> Path:
    """Lay out the Bronze tree under ``root`` and return the bronze dir."""
    bronze = root / "bronze" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    bronze.mkdir(parents=True, exist_ok=True)

    (bronze / f"{_CHAPTER} 호흡.txt").write_text(_CHAPTER_TXT, encoding="utf-8")

    spec = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "week": _WEEK,
        "chapter_no": _CHAPTER_NO,
        "chapter": _CHAPTER,
        "quiz_count": _QUIZ_COUNT,
        "formative_count": _FORMATIVE_COUNT,
    }
    (bronze / "generation_spec.yaml").write_text(
        json.dumps(spec, ensure_ascii=False), encoding="utf-8"
    )

    curriculum = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "entries": [
            {
                "week": _WEEK,
                "chapter": _CHAPTER,
                "chapter_no": _CHAPTER_NO,
                "sections": [
                    "1. 코의 구조",
                    "2. 인두와 후두",
                    "3. 기관과 기관지",
                    "4. 폐포와 가스 교환",
                    "5. 호흡의 조절",
                ],
            }
        ],
    }
    (bronze / "curriculum_map.yaml").write_text(
        json.dumps(curriculum, ensure_ascii=False), encoding="utf-8"
    )
    return bronze


def _write_envelope(responses_dir: Path, slot_id: str, item_json: dict) -> None:
    """Write a ``responses/{slot_id}.json`` envelope for the SubscriptionBackend."""
    envelope = {
        "slot_id": slot_id,
        "raw_text": json.dumps(item_json, ensure_ascii=False),
        "model": "canned-subscription",
    }
    (responses_dir / f"{slot_id}.json").write_text(
        json.dumps(envelope, ensure_ascii=False), encoding="utf-8"
    )


def _write_canned_responses(responses_dir: Path) -> None:
    """Write canned per-slot responses anchored in each slot's ASSIGNED subsection.

    Mirrors the pipeline assignment so each correct-option evidence is a distinct
    verbatim subsection line → distinct anchors, all 확인 (so dedup/exclusion drop
    nothing and the adopted set is the full N).
    """
    responses_dir.mkdir(parents=True, exist_ok=True)
    spec = _spec()

    raw_lines = _CHAPTER_TXT.splitlines()
    chunks = chunk_chapter(
        raw_lines,
        chapter_no=_CHAPTER_NO,
        chapter=_CHAPTER,
        semester=_SEMESTER,
        course_slug=_COURSE,
        source_file=f"{_CHAPTER} 호흡.txt",
    )
    chunk_by_id = {c.chunk_id: c for c in chunks}

    slots = plan_slots(spec)
    quiz_slots = [s for s in slots if s.kind == "quiz"]
    assigned = assign_subsections(quiz_slots, chunks)

    body_cursor: dict[str, int] = {}
    for slot in assigned:
        chunk = chunk_by_id[slot.subsection_chunk_id]
        body = _subsection_body_lines(chunk.text)
        idx = body_cursor.get(slot.subsection_chunk_id, 0)
        evidence_line = body[idx % len(body)]
        body_cursor[slot.subsection_chunk_id] = idx + 1
        answer_no = ((slot.ordinal - 1) % 5) + 1
        item_json = _quiz_item_json(slot.ordinal, evidence_line, answer_no)
        _write_envelope(responses_dir, slot.slot_id, item_json)

    for slot in (s for s in slots if s.kind == "formative"):
        _write_envelope(responses_dir, slot.slot_id, _formative_item_json(slot.ordinal))


def _run_build(data_root: Path) -> tuple[list, Path]:
    """Lay out a full Bronze+responses tree under ``data_root`` and run build once."""
    bronze = _build_bronze(data_root)
    silver = data_root / "silver" / "maieutica" / f"{_SEMESTER}-{_COURSE}"
    responses_dir = silver / "responses"
    _write_canned_responses(responses_dir)

    spec = load_generation_spec(bronze / "generation_spec.yaml")
    curriculum_map = load_curriculum_map(bronze / "curriculum_map.yaml")
    backend = SubscriptionBackend(staging_dir=silver / "staging", responses_dir=responses_dir)
    return build(
        spec=spec,
        curriculum_map=curriculum_map,
        bronze_dir=bronze,
        data_root=data_root,
        backend=backend,
        generation_spec_path=bronze / "generation_spec.yaml",
        curriculum_map_path=bronze / "curriculum_map.yaml",
    )


def _anchor_subsection_counts(items: list) -> dict[str, int]:
    """Return ``chunk_id → adopted-item count`` over confirmed anchors."""
    return dict(
        Counter(
            i.textbook_evidence.chunk_id
            for i in items
            if i.textbook_evidence is not None and i.textbook_evidence.status == "확인"
        )
    )


def test_build_byte_identical_xls_xlsx_yaml_across_two_runs(tmp_path: Path) -> None:
    """SC-004: .xls / .xlsx / 완전판 yaml are byte-identical across two data roots.

    Also asserts stable subsection grouping (R10 stable sort) across the runs;
    the manifest is excluded (its generated_at is the lone non-deterministic field).
    """
    root_a = tmp_path / "a" / "data"
    root_b = tmp_path / "b" / "data"
    items_a, run_a = _run_build(root_a)
    items_b, run_b = _run_build(root_b)

    xls_name = f"QuestionUploadExcel_{_WEEK}주차.xls"
    xlsx_name = formative_xlsx_filename(_CHAPTER_NO, _CHAPTER)
    yaml_name = "출제후보_완전판.yaml"

    for name in (xls_name, xlsx_name, yaml_name):
        bytes_a = (run_a / name).read_bytes()
        bytes_b = (run_b / name).read_bytes()
        assert bytes_a == bytes_b, f"{name} differs across two identical-input builds"

    # Stable sort under subsection shuffle: identical subsection grouping.
    assert _anchor_subsection_counts(items_a) == _anchor_subsection_counts(items_b)

    # The run dirs share the same deterministic run_id (same inputs).
    assert run_a.name == run_b.name


def test_cached_rebuild_makes_zero_llm_recalls(tmp_path: Path) -> None:
    """R10/SC-004: a second build against the same data root is fully cache-served.

    The first build populates ``cache/``; the second (same inputs, same root)
    serves every slot from the InputHashCache under the serial avoid_list, so the
    second manifest's ``cache_hit_rate == 1.0``.
    """
    data_root = tmp_path / "data"
    _items1, run1 = _run_build(data_root)
    _items2, run2 = _run_build(data_root)
    assert run1 == run2, "deterministic run_id must be stable across rebuilds"

    manifest = json.loads((run2 / "manifest_maieutica.json").read_text(encoding="utf-8"))
    assert manifest["cache_hit_rate"] == 1.0, (
        "second build must be fully cache-served (0 LLM re-calls) under the "
        f"serial avoid_list — got {manifest['cache_hit_rate']!r}"
    )


def test_build_xls_roundtrip_cell_types_and_zero_padding(tmp_path: Path) -> None:
    """FR-013: the built .xls honours the immutable LMS cell-type / 0-padding contract."""
    data_root = tmp_path / "data"
    _items, run_dir = _run_build(data_root)
    xls_path = run_dir / f"QuestionUploadExcel_{_WEEK}주차.xls"

    book = xlrd.open_workbook(str(xls_path))
    assert book.nsheets == 2
    assert book.sheet_by_index(0).name == guide_sheet_name()
    sheet = book.sheet_by_index(1)
    assert sheet.name == "Sheet1"

    header_values = [sheet.cell(0, c).value for c in range(sheet.ncols)]
    assert header_values == list(QUIZ_HEADERS)
    col = {h: i for i, h in enumerate(QUIZ_HEADERS)}

    # Every data row: 문제번호 NUMBER, 답안 TEXT "n", 예상주차 TEXT "009", 문항유형 TEXT "002".
    assert sheet.nrows == 1 + _QUIZ_COUNT
    for r in range(1, sheet.nrows):
        cell_num = sheet.cell(r, col["문제번호"])
        assert cell_num.ctype == xlrd.XL_CELL_NUMBER
        assert cell_num.value == r  # renumbered 1..N

        cell_ans = sheet.cell(r, col["답안"])
        assert cell_ans.ctype == xlrd.XL_CELL_TEXT
        assert cell_ans.value.isdigit()  # text digit, not number

        cell_week = sheet.cell(r, col["예상주차"])
        assert cell_week.ctype == xlrd.XL_CELL_TEXT
        assert cell_week.value == "009"  # zero-padded preserved

        cell_type = sheet.cell(r, col["문항유형"])
        assert cell_type.ctype == xlrd.XL_CELL_TEXT
        assert cell_type.value == "002"  # zero-padded constant


def test_build_xlsx_roundtrip_cell_types(tmp_path: Path) -> None:
    """FR-013: the built formative .xlsx honours the bhu_text_mining cell-type contract."""
    data_root = tmp_path / "data"
    _items, run_dir = _run_build(data_root)
    xlsx_path = run_dir / formative_xlsx_filename(_CHAPTER_NO, _CHAPTER)

    workbook = openpyxl.load_workbook(xlsx_path)
    assert workbook.sheetnames == ["Formative Test"]
    sheet = workbook.active

    header = [c.value for c in sheet[1]]
    assert header == list(FORMATIVE_HEADERS)
    col = {h: i + 1 for i, h in enumerate(FORMATIVE_HEADERS)}

    assert sheet.max_row == 1 + _FORMATIVE_COUNT
    for row in range(2, sheet.max_row + 1):
        assert isinstance(sheet.cell(row=row, column=col["No."]).value, int)
        assert isinstance(sheet.cell(row=row, column=col["Chapter"]).value, int)
        for name in ("Topic", "Question", "Keywords", "Model Answer"):
            assert isinstance(sheet.cell(row=row, column=col[name]).value, str)
