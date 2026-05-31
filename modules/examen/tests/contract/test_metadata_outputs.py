"""T044 — Contract tests for US4: metadata completeness, dedup, length verify,
ExamItem projection, and output-path separation.

Contracts asserted:
- C1  28 columns in exact order (already covered by T018; re-asserted here for
      completeness with US4 items that carry duplicate_flag=True and flagged
      review_note).
- C2  excel↔yaml hold the same items (same item_nos, same duplicate_flag values).
- C3  wrong_explanation / leap_explanation outside 270–330 codepoints are
      flagged in review_note; in-range items are NOT flagged.
- C4  intent outside 40–60 codepoints is flagged in review_note.
- C5  ExamItem projection round-trips: project_to_exam_item returns a valid
      immersio ExamItem that can be constructed without error.
- C6  write_exam_item_projection produces a loadable YAML sidecar; each
      projected item satisfies ExamItem(**item_dict) successfully.
- C7  detect_duplicates is deterministic across two calls with the same input.
- C8  detect_duplicates flags items sharing the same key_concept (keeps first,
      flags rest).
- C9  items written to run_dir (run-isolated Gold path), not the base gold_dir.
- C10 difficulty mapping "1_쉬움"→"easy", "2_보통"→"medium", "3_어려움"→"hard".
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import yaml
from paideia_shared.schemas import ExamItem, ExamItemDraft, TextbookEvidence

# ---------------------------------------------------------------------------
# Helpers — canonical column list (mirrors T018, re-asserted here)
# ---------------------------------------------------------------------------

_EXPECTED_COLUMNS = [
    "번호",
    "출처",
    "원본출처식별자",
    "챕터",
    "절",
    "주차",
    "핵심개념",
    "강조여부",
    "문제유형",
    "난이도",
    "문두방향",
    "문제",
    "보기1",
    "보기2",
    "보기3",
    "보기4",
    "보기5",
    "정답",
    "보기별오답근거",
    "오답설명",
    "도약설명",
    "교재근거위치",
    "출제의도",
    "보기글자수검증",
    "중복플래그",
    "문제검증",
    "채택상태",
    "비고",
]
assert len(_EXPECTED_COLUMNS) == 28

_PINNED_WHEN = datetime.datetime(2026, 1, 1, 0, 0, 0, tzinfo=datetime.UTC)

# 옵션 — 번호 포함 30자 (OK 범위)
_OPTIONS_OK = [
    "① " + "가" * 28,
    "② " + "나" * 28,
    "③ " + "다" * 28,
    "④ " + "라" * 28,
    "⑤ " + "마" * 28,
]

_DISTRACTOR = [f"근거{i}" for i in range(1, 6)]


def _make_item(
    item_no: int = 1,
    key_concept: str | None = "폐포",
    wrong_explanation: str | None = None,
    leap_explanation: str | None = None,
    intent: str | None = None,
    duplicate_flag: bool = False,
    difficulty: str = "1_쉬움",
    source: str = "textbook",
) -> ExamItemDraft:
    """Build a minimal valid ExamItemDraft for US4 contract tests."""
    # 기본값: 길이 범위 내 (270~330자)
    wrong = wrong_explanation or ("오답 설명 테스트입니다." * 27)[:300]
    leap = leap_explanation or ("도약 설명 테스트입니다." * 27)[:300]
    intent_str = intent or ("출제의도 확인 테스트입니다. 폐포 기능.")  # ~22자 → 범위 내 조정

    # intent 기본값을 정확히 40~60자로 맞추기
    if intent is None:
        base = "폐포의 기능과 가스교환 기전을 정확히 파악하는지 확인한다"
        # len(base) 확인 후 조정
        while len(base) < 40:
            base += "."
        intent_str = base[:55]

    ev = TextbookEvidence(
        source_file="8장 호흡계통.txt",
        line=42,
        found_text="호흡계통 주요 내용",
        status="확인",
        search_term="폐포",
    )
    return ExamItemDraft(
        semester="2026-1",
        course_slug="anatomy",
        item_no=item_no,
        source=source,  # type: ignore[arg-type]
        source_ref=None,
        chapter="8장 호흡계통",
        chapter_no=8,
        section="1. 기도",
        week=None,
        key_concept=key_concept,
        is_emphasized=True,
        emphasis_class_count=3,
        question_type="지식축적",
        bloom=None,
        difficulty=difficulty,  # type: ignore[arg-type]
        stem_polarity="부정형",
        text="다음 중 폐포에 대한 설명으로 가장 옳지 않은 것은?",
        options=_OPTIONS_OK,
        answer_no=3,
        distractor_rationale=_DISTRACTOR,
        wrong_explanation=wrong,
        leap_explanation=leap,
        textbook_evidence=ev,
        intent=intent_str,
        option_length_ok=True,
        duplicate_flag=duplicate_flag,
        review_note="",
        adoption_status="생성",
        note=None,
    )


# ---------------------------------------------------------------------------
# C1 — 28 columns (with US4 fields set)
# ---------------------------------------------------------------------------


class TestColumnContract:
    """28-column contract with US4 duplicate_flag and review_note populated."""

    def test_28_columns_with_flagged_items(self, tmp_path: Path) -> None:
        """xlsx still has exactly 28 columns when duplicate_flag=True."""
        from examen.output.determinism import finalize_xlsx
        from examen.output.xlsx import write_xlsx

        items = [
            _make_item(item_no=1, duplicate_flag=False),
            _make_item(item_no=2, duplicate_flag=True),
        ]
        dest = tmp_path / "draft.xlsx"
        write_xlsx(items, dest)
        finalize_xlsx(dest, _PINNED_WHEN)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == _EXPECTED_COLUMNS, (
            f"Column order mismatch.\nExpected: {_EXPECTED_COLUMNS}\nGot: {headers}"
        )


# ---------------------------------------------------------------------------
# C2 — excel↔yaml consistency with duplicate_flag
# ---------------------------------------------------------------------------


class TestExcelYamlConsistency:
    """xlsx and yaml hold the same duplicate_flag and item_no values."""

    def test_duplicate_flag_consistent(self, tmp_path: Path) -> None:
        """duplicate_flag values in xlsx col-25 match yaml duplicate_flag fields."""
        from examen.output.determinism import finalize_xlsx
        from examen.output.xlsx import write_xlsx
        from examen.output.yaml_out import write_yaml

        items = [
            _make_item(item_no=1, duplicate_flag=False),
            _make_item(item_no=2, duplicate_flag=True),
            _make_item(item_no=3, duplicate_flag=False),
        ]
        xlsx_path = tmp_path / "draft.xlsx"
        yaml_path = tmp_path / "draft.yaml"
        write_xlsx(items, xlsx_path)
        finalize_xlsx(xlsx_path, _PINNED_WHEN)
        write_yaml(items, yaml_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        xlsx_flags = [ws.cell(row=r, column=25).value for r in range(2, 5)]

        yaml_data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
        yaml_flags = [d["duplicate_flag"] for d in yaml_data]

        # xlsx stores Python bool; yaml stores bool — both should be equivalent
        assert [bool(f) for f in xlsx_flags] == yaml_flags, (
            f"duplicate_flag mismatch: xlsx={xlsx_flags} yaml={yaml_flags}"
        )

    def test_review_note_consistent(self, tmp_path: Path) -> None:
        """review_note in xlsx col-26 matches yaml review_note."""
        from examen.output.determinism import finalize_xlsx
        from examen.output.xlsx import write_xlsx
        from examen.output.yaml_out import write_yaml

        item = _make_item(item_no=1)
        # Simulate a length-flagged item by adding a review_note
        item_with_note = item.model_copy(update={"review_note": "[length_check] wrong_explanation 길이 위반"})
        items = [item_with_note]

        xlsx_path = tmp_path / "note.xlsx"
        yaml_path = tmp_path / "note.yaml"
        write_xlsx(items, xlsx_path)
        finalize_xlsx(xlsx_path, _PINNED_WHEN)
        write_yaml(items, yaml_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        xlsx_note = ws.cell(row=2, column=26).value

        yaml_data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
        yaml_note = yaml_data[0]["review_note"]

        assert xlsx_note == yaml_note, (
            f"review_note mismatch: xlsx={xlsx_note!r} yaml={yaml_note!r}"
        )


# ---------------------------------------------------------------------------
# C3/C4 — length verification flags
# ---------------------------------------------------------------------------


class TestLengthVerification:
    """check_explanation_lengths flags out-of-range explanations/intent."""

    def test_in_range_wrong_explanation_not_flagged(self) -> None:
        """wrong_explanation 270–330 codepoints: no length flag added."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(wrong_explanation="가" * 300, leap_explanation="나" * 300)
        result = check_explanation_lengths(item)
        assert "[length_check]" not in result.review_note

    def test_too_short_wrong_explanation_flagged(self) -> None:
        """wrong_explanation < 270 codepoints: flagged in review_note."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(wrong_explanation="가" * 100, leap_explanation="나" * 300)
        result = check_explanation_lengths(item)
        assert "[length_check]" in result.review_note
        assert "wrong_explanation" in result.review_note

    def test_too_long_wrong_explanation_flagged(self) -> None:
        """wrong_explanation > 330 codepoints: flagged in review_note."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(wrong_explanation="가" * 400, leap_explanation="나" * 300)
        result = check_explanation_lengths(item)
        assert "[length_check]" in result.review_note
        assert "wrong_explanation" in result.review_note

    def test_too_short_leap_explanation_flagged(self) -> None:
        """leap_explanation < 270 codepoints: flagged in review_note."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(wrong_explanation="가" * 300, leap_explanation="나" * 50)
        result = check_explanation_lengths(item)
        assert "[length_check]" in result.review_note
        assert "leap_explanation" in result.review_note

    def test_in_range_intent_not_flagged(self) -> None:
        """intent 40–60 codepoints: no length flag added."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(
            wrong_explanation="가" * 300,
            leap_explanation="나" * 300,
            intent="가" * 50,
        )
        result = check_explanation_lengths(item)
        # No length_check flag for intent
        assert "intent" not in result.review_note or "[length_check]" not in result.review_note

    def test_too_short_intent_flagged(self) -> None:
        """intent < 40 codepoints: flagged in review_note."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(
            wrong_explanation="가" * 300,
            leap_explanation="나" * 300,
            intent="짧음",
        )
        result = check_explanation_lengths(item)
        assert "[length_check]" in result.review_note
        assert "intent" in result.review_note

    def test_too_long_intent_flagged(self) -> None:
        """intent > 60 codepoints: flagged in review_note."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(
            wrong_explanation="가" * 300,
            leap_explanation="나" * 300,
            intent="가" * 70,
        )
        result = check_explanation_lengths(item)
        assert "[length_check]" in result.review_note
        assert "intent" in result.review_note

    def test_does_not_raise_on_violation(self) -> None:
        """check_explanation_lengths never raises — only flags."""
        from examen.verify.format_checks import check_explanation_lengths

        # All three fields out of range — must not raise
        item = _make_item(
            wrong_explanation="짧",
            leap_explanation="짧",
            intent="짧",
        )
        result = check_explanation_lengths(item)  # must not raise
        assert isinstance(result, ExamItemDraft)

    def test_existing_review_note_preserved(self) -> None:
        """Pre-existing review_note is preserved when new flag is added."""
        from examen.verify.format_checks import check_explanation_lengths

        item = _make_item(wrong_explanation="가" * 50, leap_explanation="나" * 300)
        item_with_note = item.model_copy(update={"review_note": "기존 메모"})
        result = check_explanation_lengths(item_with_note)
        assert "기존 메모" in result.review_note
        assert "[length_check]" in result.review_note


# ---------------------------------------------------------------------------
# C5/C6 — ExamItem projection
# ---------------------------------------------------------------------------


class TestExamItemProjection:
    """project_to_exam_item maps ExamItemDraft → valid ExamItem."""

    def test_project_returns_exam_item(self) -> None:
        """project_to_exam_item returns an ExamItem instance."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item(item_no=1, difficulty="1_쉬움")
        result = project_to_exam_item(
            item,
            semester="2026-1",
            course_slug="anatomy",
        )
        assert isinstance(result, ExamItem)

    def test_difficulty_easy_maps(self) -> None:
        """1_쉬움 → expected_difficulty='easy'."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item(difficulty="1_쉬움")
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert result.expected_difficulty == "easy"

    def test_difficulty_medium_maps(self) -> None:
        """2_보통 → expected_difficulty='medium'."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item(difficulty="2_보통")
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert result.expected_difficulty == "medium"

    def test_difficulty_hard_maps(self) -> None:
        """3_어려움 → expected_difficulty='hard'."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item(difficulty="3_어려움")
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert result.expected_difficulty == "hard"

    def test_answer_key_is_string(self) -> None:
        """answer_key is a string (str(answer_no))."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item()
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert isinstance(result.answer_key, str)
        assert result.answer_key == str(item.answer_no)

    def test_source_propagated(self) -> None:
        """source literal propagates unchanged."""
        from examen.output.exam_item_projection import project_to_exam_item

        for src in ("textbook", "formative", "quiz"):
            item = _make_item(source=src)
            result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
            assert result.source == src

    def test_distractors_are_options(self) -> None:
        """ExamItem.distractors == list(draft.options)."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item()
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert result.distractors == list(item.options)

    def test_text_propagated(self) -> None:
        """ExamItem.text == draft.text."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item()
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert result.text == item.text

    def test_chapter_propagated(self) -> None:
        """ExamItem.chapter == draft.chapter."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item()
        result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        assert result.chapter == item.chapter

    def test_exam_item_is_immersio_loadable(self) -> None:
        """Projected ExamItem can be re-constructed via ExamItem(**data)."""
        from examen.output.exam_item_projection import project_to_exam_item

        item = _make_item()
        projected = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
        # Round-trip via model_dump → construct
        data = projected.model_dump(mode="python")
        reconstructed = ExamItem(**data)
        assert reconstructed == projected

    def test_write_exam_item_projection_creates_yaml(self, tmp_path: Path) -> None:
        """write_exam_item_projection writes a loadable yaml sidecar."""
        from examen.output.exam_item_projection import write_exam_item_projection

        items = [_make_item(item_no=i, difficulty=d) for i, d in enumerate(
            ["1_쉬움", "2_보통", "3_어려움"], start=1
        )]
        dest = tmp_path / "exam_items.yaml"
        write_exam_item_projection(items, dest, semester="2026-1", course_slug="anatomy")
        assert dest.exists()

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        assert isinstance(data, list)
        assert len(data) == 3

    def test_projection_yaml_each_item_loadable(self, tmp_path: Path) -> None:
        """Each dict in the projection yaml satisfies ExamItem(**dict)."""
        from examen.output.exam_item_projection import write_exam_item_projection

        items = [_make_item(item_no=i) for i in range(1, 4)]
        dest = tmp_path / "exam_items.yaml"
        write_exam_item_projection(items, dest, semester="2026-1", course_slug="anatomy")

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        for d in data:
            ExamItem(**d)  # must not raise

    def test_projection_deterministic(self, tmp_path: Path) -> None:
        """Two writes with the same items produce byte-identical yaml."""
        from examen.output.exam_item_projection import write_exam_item_projection

        items = [_make_item(item_no=i) for i in range(1, 4)]
        p1 = tmp_path / "a.yaml"
        p2 = tmp_path / "b.yaml"
        write_exam_item_projection(items, p1, semester="2026-1", course_slug="anatomy")
        write_exam_item_projection(items, p2, semester="2026-1", course_slug="anatomy")
        assert p1.read_bytes() == p2.read_bytes(), "exam_items.yaml is not byte-identical"


# ---------------------------------------------------------------------------
# C7/C8 — detect_duplicates
# ---------------------------------------------------------------------------


class TestDetectDuplicates:
    """detect_duplicates flags items sharing the same key_concept."""

    def test_no_duplicates_returns_same(self) -> None:
        """Items with distinct key_concepts: duplicate_flag stays False."""
        from examen.verify.format_checks import detect_duplicates

        items = [
            _make_item(item_no=1, key_concept="폐포"),
            _make_item(item_no=2, key_concept="가로막"),
            _make_item(item_no=3, key_concept="산소포화도"),
        ]
        result = detect_duplicates(items)
        assert all(not i.duplicate_flag for i in result)

    def test_shared_key_concept_flags_second(self) -> None:
        """Items with the same key_concept: first kept, rest flagged."""
        from examen.verify.format_checks import detect_duplicates

        items = [
            _make_item(item_no=1, key_concept="폐포"),
            _make_item(item_no=2, key_concept="폐포"),
            _make_item(item_no=3, key_concept="가로막"),
        ]
        result = detect_duplicates(items)
        assert result[0].duplicate_flag is False, "first item should NOT be flagged"
        assert result[1].duplicate_flag is True, "second item with same key_concept should be flagged"
        assert result[2].duplicate_flag is False, "different key_concept should NOT be flagged"

    def test_three_duplicates_flags_second_and_third(self) -> None:
        """Three items with same key_concept: first kept, second+third flagged."""
        from examen.verify.format_checks import detect_duplicates

        items = [
            _make_item(item_no=1, key_concept="폐포"),
            _make_item(item_no=2, key_concept="폐포"),
            _make_item(item_no=3, key_concept="폐포"),
        ]
        result = detect_duplicates(items)
        assert result[0].duplicate_flag is False
        assert result[1].duplicate_flag is True
        assert result[2].duplicate_flag is True

    def test_none_key_concept_not_flagged_as_duplicate(self) -> None:
        """Items with key_concept=None are not grouped together as duplicates."""
        from examen.verify.format_checks import detect_duplicates

        items = [
            _make_item(item_no=1, key_concept=None),
            _make_item(item_no=2, key_concept=None),
        ]
        result = detect_duplicates(items)
        # key_concept=None items should not be flagged as duplicates of each other
        assert result[0].duplicate_flag is False
        assert result[1].duplicate_flag is False

    def test_already_flagged_item_stays_flagged(self) -> None:
        """If an item was already duplicate_flag=True, it remains True after dedup."""
        from examen.verify.format_checks import detect_duplicates

        items = [
            _make_item(item_no=1, key_concept="폐포"),
            _make_item(item_no=2, key_concept="다른개념", duplicate_flag=True),
        ]
        result = detect_duplicates(items)
        assert result[1].duplicate_flag is True  # pre-set flag preserved

    def test_deterministic_two_calls(self) -> None:
        """detect_duplicates with the same input returns identical results."""
        from examen.verify.format_checks import detect_duplicates

        items = [
            _make_item(item_no=1, key_concept="폐포"),
            _make_item(item_no=2, key_concept="폐포"),
            _make_item(item_no=3, key_concept="가로막"),
        ]
        r1 = detect_duplicates(items)
        r2 = detect_duplicates(items)
        assert [i.duplicate_flag for i in r1] == [i.duplicate_flag for i in r2]

    def test_empty_list_returns_empty(self) -> None:
        """detect_duplicates([]) returns []."""
        from examen.verify.format_checks import detect_duplicates

        assert detect_duplicates([]) == []


# ---------------------------------------------------------------------------
# C9 — Output-path separation (run dir vs. base gold dir)
# ---------------------------------------------------------------------------


class TestOutputPathSeparation:
    """Items are written to run_gold_dir, not the base gold_dir."""

    def test_run_dir_is_under_runs_subdir(self, tmp_path: Path) -> None:
        """run_gold_dir returns a path under .../runs/{run_id}/."""
        from examen.output.paths import gold_dir, run_gold_dir

        run_dir = run_gold_dir("2026-1", "anatomy", run_id="abc123", data_root=tmp_path)
        base = gold_dir("2026-1", "anatomy", data_root=tmp_path)

        assert str(run_dir).startswith(str(base))
        assert "runs" in str(run_dir)
        assert "abc123" in str(run_dir)

    def test_run_dir_does_not_equal_base_gold_dir(self, tmp_path: Path) -> None:
        """run_gold_dir != gold_dir (run-isolated, not base)."""
        from examen.output.paths import gold_dir, run_gold_dir

        run_dir = run_gold_dir("2026-1", "anatomy", run_id="abc123", data_root=tmp_path)
        base = gold_dir("2026-1", "anatomy", data_root=tmp_path)
        assert run_dir != base


# ---------------------------------------------------------------------------
# C10 — Difficulty mapping (covered in C5 projection tests above; re-assert)
# ---------------------------------------------------------------------------


class TestDifficultyMapping:
    """Explicit difficulty→expected_difficulty mapping table."""

    def test_all_three_difficulties(self) -> None:
        """All three difficulty levels map correctly."""
        from examen.output.exam_item_projection import project_to_exam_item

        mapping = {
            "1_쉬움": "easy",
            "2_보통": "medium",
            "3_어려움": "hard",
        }
        for diff, expected in mapping.items():
            item = _make_item(difficulty=diff)
            result = project_to_exam_item(item, semester="2026-1", course_slug="anatomy")
            assert result.expected_difficulty == expected, (
                f"difficulty={diff!r} should map to expected_difficulty={expected!r}, "
                f"got {result.expected_difficulty!r}"
            )
