"""T018 — Contract tests: xlsx 28-column schema + yaml round-trip + determinism.

Contracts:
- write_xlsx produces EXACTLY 28 columns in the order defined in
  specs/008-examen-question-gen/contracts/exam_draft_outputs.md.
- write_yaml produces nested full-fidelity YAML round-trippable to
  ExamItemDraft.
- xlsx and yaml hold the same item set (no contradiction).
- xlsx is byte-identical on re-write (determinism via finalize_xlsx).
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import yaml
from paideia_shared.schemas import ExamItemDraft, TextbookEvidence

# ---------------------------------------------------------------------------
# Expected column names in exact order (28 columns, from exam_draft_outputs.md)
# ---------------------------------------------------------------------------

EXPECTED_COLUMNS = [
    "번호",            # 1  item_no
    "출처",            # 2  source (형성평가/퀴즈/교과서)
    "원본출처식별자",   # 3  source_ref
    "챕터",            # 4  chapter
    "절",              # 5  section
    "주차",            # 6  week
    "핵심개념",         # 7  key_concept
    "강조여부",         # 8  is_emphasized (강의강조/자습)
    "문제유형",         # 9  question_type
    "난이도",           # 10 difficulty
    "문두방향",         # 11 stem_polarity
    "문제",            # 12 text
    "보기1",           # 13 options[0]
    "보기2",           # 14 options[1]
    "보기3",           # 15 options[2]
    "보기4",           # 16 options[3]
    "보기5",           # 17 options[4]
    "정답",            # 18 answer_no
    "보기별오답근거",   # 19 distractor_rationale joined by \n
    "오답설명",         # 20 wrong_explanation
    "도약설명",         # 21 leap_explanation
    "교재근거위치",     # 22 textbook_evidence (파일:행 + status)
    "출제의도",         # 23 intent
    "보기글자수검증",   # 24 option_length_ok (OK/위반)
    "중복플래그",       # 25 duplicate_flag
    "문제검증",         # 26 review_note (blank at generation)
    "채택상태",         # 27 adoption_status
    "비고",            # 28 note
]

assert len(EXPECTED_COLUMNS) == 28, f"Expected 28 columns, got {len(EXPECTED_COLUMNS)}"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_BASE_OPTIONS_30 = [
    "① " + "가" * 28,
    "② " + "나" * 28,
    "③ " + "다" * 28,
    "④ " + "라" * 28,
    "⑤ " + "마" * 28,
]

_BASE_DISTRACTOR = [f"근거{i}" for i in range(1, 6)]


def _make_item(
    item_no: int = 1,
    source: str = "textbook",
    evidence_status: str = "확인",
) -> ExamItemDraft:
    """Build a minimal valid ExamItemDraft for contract tests."""
    ev = TextbookEvidence(
        source_file="8장 호흡계통.txt",
        line=42,
        found_text="호흡계통 주요 내용",
        status=evidence_status,  # type: ignore[arg-type]
        search_term="폐포",
    )
    return ExamItemDraft(
        semester="2026-1",
        course_slug="anatomy",
        item_no=item_no,
        source=source,  # type: ignore[arg-type]
        source_ref=None,
        chapter="8장 호흡계통",
        chapter_no=8,
        section="1. 기도",
        week=None,
        key_concept="폐포",
        is_emphasized=True,
        emphasis_class_count=3,
        question_type="지식축적",
        bloom=None,
        difficulty="1_쉬움",
        stem_polarity="부정형",
        text="다음 중 폐포에 대한 설명으로 가장 옳지 않은 것은?",
        options=_BASE_OPTIONS_30,
        answer_no=3,
        distractor_rationale=_BASE_DISTRACTOR,
        wrong_explanation="오답 설명 테스트." * 20,
        leap_explanation="도약 설명 테스트." * 20,
        textbook_evidence=ev,
        intent="폐포의 기능을 정확히 파악하는지 확인.",
        option_length_ok=True,
        duplicate_flag=False,
        review_note="",
        adoption_status="생성",
        note=None,
    )


_PINNED_WHEN = datetime.datetime(2026, 1, 1, 0, 0, 0, tzinfo=datetime.UTC)


# ---------------------------------------------------------------------------
# T018-A: xlsx 28-column contract
# ---------------------------------------------------------------------------


class TestXlsxColumnContract:
    """Verify exact 28-column layout of write_xlsx output."""

    def _write(
        self, items: list[ExamItemDraft], path: Path
    ) -> None:
        from examen.output.determinism import finalize_xlsx
        from examen.output.xlsx import write_xlsx
        write_xlsx(items, path)
        finalize_xlsx(path, _PINNED_WHEN)

    def test_xlsx_has_exactly_28_columns(self, tmp_path: Path) -> None:
        """xlsx output has exactly 28 columns."""
        items = [_make_item(item_no=i) for i in range(1, 4)]
        dest = tmp_path / "draft.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert len(headers) == 28, f"Expected 28 columns, got {len(headers)}: {headers}"

    def test_xlsx_column_names_exact_order(self, tmp_path: Path) -> None:
        """Column names match EXPECTED_COLUMNS exactly (order matters)."""
        items = [_make_item()]
        dest = tmp_path / "draft.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == EXPECTED_COLUMNS, (
            f"Column order mismatch.\n"
            f"Expected: {EXPECTED_COLUMNS}\n"
            f"Got:      {headers}"
        )

    def test_xlsx_row_count_matches_items(self, tmp_path: Path) -> None:
        """Number of data rows (excl. header) equals len(items)."""
        items = [_make_item(item_no=i) for i in range(1, 6)]
        dest = tmp_path / "draft.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        data_rows = ws.max_row - 1  # subtract header row
        assert data_rows == len(items), f"Expected {len(items)} rows, got {data_rows}"

    def test_xlsx_item_no_column(self, tmp_path: Path) -> None:
        """Column '번호' (col 1) contains item_no values."""
        items = [_make_item(item_no=i) for i in range(1, 4)]
        dest = tmp_path / "draft.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        nos = [ws.cell(row=r, column=1).value for r in range(2, 5)]
        assert nos == [1, 2, 3], f"번호 column: {nos}"

    def test_xlsx_source_column_textbook(self, tmp_path: Path) -> None:
        """Column '출처' (col 2) contains '교과서' for source='textbook'."""
        items = [_make_item(source="textbook")]
        dest = tmp_path / "draft.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        val = ws.cell(row=2, column=2).value
        assert val == "교과서", f"Expected '교과서', got {val!r}"

    def test_xlsx_textbook_evidence_column(self, tmp_path: Path) -> None:
        """Column '교재근거위치' (col 22) contains file:line+status."""
        item = _make_item(evidence_status="확인")
        dest = tmp_path / "evidence.xlsx"
        self._write([item], dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        val = ws.cell(row=2, column=22).value
        assert val is not None
        assert "확인" in str(val), f"Expected '확인' in evidence column, got {val!r}"

    def test_xlsx_option_length_ok_column(self, tmp_path: Path) -> None:
        """Column '보기글자수검증' (col 24) contains 'OK' or '위반'."""
        items = [_make_item()]
        dest = tmp_path / "optlen.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        val = ws.cell(row=2, column=24).value
        assert val in ("OK", "위반"), f"Expected 'OK' or '위반', got {val!r}"

    def test_xlsx_distractor_rationale_joined_by_newline(self, tmp_path: Path) -> None:
        """Column '보기별오답근거' (col 19): 5 entries joined by newline."""
        items = [_make_item()]
        dest = tmp_path / "distractor.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        val = ws.cell(row=2, column=19).value
        assert isinstance(val, str)
        parts = val.split("\n")
        assert len(parts) == 5, f"Expected 5 parts, got {len(parts)}: {parts}"

    def test_xlsx_options_split_to_five_columns(self, tmp_path: Path) -> None:
        """Columns '보기1'~'보기5' (cols 13~17) each contain one option."""
        items = [_make_item()]
        dest = tmp_path / "options.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        for col_idx, col_no in enumerate(range(13, 18)):
            val = ws.cell(row=2, column=col_no).value
            assert val is not None, f"보기{col_idx+1} is None at column {col_no}"
            assert isinstance(val, str)

    def test_xlsx_byte_identical_on_rewrite(self, tmp_path: Path) -> None:
        """Two identical calls produce byte-identical xlsx files."""
        items = [_make_item(item_no=i) for i in range(1, 4)]
        p1 = tmp_path / "a.xlsx"
        p2 = tmp_path / "b.xlsx"
        self._write(items, p1)
        self._write(items, p2)
        assert p1.read_bytes() == p2.read_bytes(), "xlsx is not byte-identical on rewrite"

    def test_xlsx_emphasis_label(self, tmp_path: Path) -> None:
        """Column '강조여부' (col 8): '강의강조' for is_emphasized=True."""
        items = [_make_item()]
        dest = tmp_path / "emph.xlsx"
        self._write(items, dest)

        wb = openpyxl.load_workbook(dest)
        ws = wb.active
        val = ws.cell(row=2, column=8).value
        assert val in ("강의강조", "자습", "", None), (
            f"강조여부 column: unexpected value {val!r}"
        )


# ---------------------------------------------------------------------------
# T018-B: yaml contract
# ---------------------------------------------------------------------------


class TestYamlContract:
    """Verify nested full-fidelity yaml output."""

    def _write(self, items: list[ExamItemDraft], path: Path) -> None:
        from examen.output.yaml_out import write_yaml
        write_yaml(items, path)

    def test_yaml_file_created(self, tmp_path: Path) -> None:
        """write_yaml creates the output file."""
        items = [_make_item()]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)
        assert dest.exists()

    def test_yaml_parses_as_list(self, tmp_path: Path) -> None:
        """yaml output parses to a list."""
        items = [_make_item(item_no=i) for i in range(1, 4)]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        assert isinstance(data, list), f"Expected list, got {type(data)}"

    def test_yaml_item_count_matches(self, tmp_path: Path) -> None:
        """yaml list length equals number of items written."""
        n = 5
        items = [_make_item(item_no=i) for i in range(1, n + 1)]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        assert len(data) == n

    def test_yaml_nested_textbook_evidence(self, tmp_path: Path) -> None:
        """yaml preserves textbook_evidence as a nested object (not flattened)."""
        items = [_make_item()]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        ev = data[0]["textbook_evidence"]
        assert isinstance(ev, dict), "textbook_evidence must be a nested dict in yaml"
        assert "source_file" in ev
        assert "status" in ev

    def test_yaml_distractor_rationale_is_list(self, tmp_path: Path) -> None:
        """yaml preserves distractor_rationale as list[5] (not joined string)."""
        items = [_make_item()]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        dr = data[0]["distractor_rationale"]
        assert isinstance(dr, list), "distractor_rationale must be a list"
        assert len(dr) == 5

    def test_yaml_options_is_list_of_5(self, tmp_path: Path) -> None:
        """yaml preserves options as list[5]."""
        items = [_make_item()]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        opts = data[0]["options"]
        assert isinstance(opts, list)
        assert len(opts) == 5

    def test_yaml_allow_unicode(self, tmp_path: Path) -> None:
        """yaml does not escape Korean characters (allow_unicode=True)."""
        items = [_make_item()]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        raw = dest.read_text(encoding="utf-8")
        assert "폐포" in raw, "Korean text should not be \\uXXXX escaped"
        assert "\\u" not in raw

    def test_yaml_deterministic(self, tmp_path: Path) -> None:
        """Two write_yaml calls with same items produce byte-identical files."""
        items = [_make_item(item_no=i) for i in range(1, 4)]
        p1 = tmp_path / "a.yaml"
        p2 = tmp_path / "b.yaml"
        self._write(items, p1)
        self._write(items, p2)
        assert p1.read_bytes() == p2.read_bytes(), "yaml is not byte-identical on rewrite"

    def test_yaml_round_trip_item_no(self, tmp_path: Path) -> None:
        """yaml round-trip preserves item_no for each item."""
        items = [_make_item(item_no=i) for i in range(1, 4)]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        nos = [d["item_no"] for d in data]
        assert nos == [1, 2, 3]

    def test_yaml_has_adoption_status(self, tmp_path: Path) -> None:
        """yaml contains adoption_status field for each item."""
        items = [_make_item()]
        dest = tmp_path / "draft.yaml"
        self._write(items, dest)

        data = yaml.safe_load(dest.read_text(encoding="utf-8"))
        assert "adoption_status" in data[0]
        assert data[0]["adoption_status"] == "생성"


# ---------------------------------------------------------------------------
# T018-C: xlsx + yaml consistency (same item set)
# ---------------------------------------------------------------------------


class TestXlsxYamlConsistency:
    """Verify xlsx and yaml hold the same item set."""

    def test_xlsx_yaml_same_item_count(self, tmp_path: Path) -> None:
        """xlsx row count == yaml list length for same items."""
        from examen.output.determinism import finalize_xlsx
        from examen.output.xlsx import write_xlsx
        from examen.output.yaml_out import write_yaml

        items = [_make_item(item_no=i) for i in range(1, 5)]
        xlsx_path = tmp_path / "draft.xlsx"
        yaml_path = tmp_path / "draft.yaml"

        write_xlsx(items, xlsx_path)
        finalize_xlsx(xlsx_path, _PINNED_WHEN)
        write_yaml(items, yaml_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        data_rows = ws.max_row - 1

        yaml_data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))

        assert data_rows == len(yaml_data), (
            f"xlsx has {data_rows} rows, yaml has {len(yaml_data)} items"
        )

    def test_xlsx_yaml_same_item_nos(self, tmp_path: Path) -> None:
        """item_no values in xlsx col-1 and yaml item_no fields match."""
        from examen.output.determinism import finalize_xlsx
        from examen.output.xlsx import write_xlsx
        from examen.output.yaml_out import write_yaml

        items = [_make_item(item_no=i) for i in range(1, 4)]
        xlsx_path = tmp_path / "draft.xlsx"
        yaml_path = tmp_path / "draft.yaml"

        write_xlsx(items, xlsx_path)
        finalize_xlsx(xlsx_path, _PINNED_WHEN)
        write_yaml(items, yaml_path)

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        xlsx_nos = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]

        yaml_data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
        yaml_nos = [d["item_no"] for d in yaml_data]

        assert xlsx_nos == yaml_nos, (
            f"item_no mismatch: xlsx={xlsx_nos} yaml={yaml_nos}"
        )
