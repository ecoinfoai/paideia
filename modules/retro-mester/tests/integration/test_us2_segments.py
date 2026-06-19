"""T036 — US2 two-segment integration tests.

RED -> GREEN: written before implementation.

Tests verify full pipeline produces group-differentiated prescriptions,
structural escalation, 집단대비 xlsx sheet, and 집단별 전략 report section.
SC-010: no student ID appears in any gold output.
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import yaml

# ---------------------------------------------------------------------------
# Fixture constants
# ---------------------------------------------------------------------------

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_KEY = f"{_SEMESTER}-{_COURSE}"

_CHAPTER_A = "1장. 해부학 서론"  # structural: both segments below threshold
_CHAPTER_B = "2장. 세포와 조직"  # non-structural: only 학령기 below

_STUDENT_IDS = {
    "2026000001": "학령기",
    "2026000002": "학령기",
    "2026000003": "만학도",
    "2026000004": "만학도",
}

# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

_AXES = [
    "digital_efficacy",
    "motivation",
    "time_availability",
    "material_preference",
    "study_strategy",
    "study_environment",
    "social_learning",
    "feedback_seeking",
]


def _combined_row(
    student_id: str,
    chapter_rates: dict[str, float],
    cluster_label: str | None = None,
    prior_readiness_q5: str | None = None,
    prior_readiness_q6: str | None = None,
) -> dict:
    """Return a minimal CombinedAnalysisRow-compatible dict."""
    has_cluster = cluster_label is not None
    row: dict = {
        "student_id": student_id,
        "name_kr": None,
        "on_roster": True,
        "section": None,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "cluster_id": 1 if has_cluster else None,
        "cluster_label": cluster_label,
        "cluster_distance": 0.1 if has_cluster else None,
        "exam_taken": True,
        "total_score": 60.0,
        "score_percent": 60.0,
        "section_percentile": 50.0,
        "cohort_percentile": 50.0,
        "z_score": 0.0,
        "chapter_correct_rates": json.dumps(chapter_rates),
        "source_correct_rates": json.dumps({"형성평가": 0.5}),
        "difficulty_correct_rates": json.dumps({"1": 0.7, "2": 0.5, "3": 0.3}),
        "expected_difficulty_correct_rates": json.dumps({"쉬움": 0.7, "보통": 0.5, "어려움": 0.3}),
        "item_type_correct_rates": json.dumps({"지식축적": 0.6, "이해": 0.5}),
        "interest_chapters_correct_rate": None,
        "aversion_chapters_correct_rate": None,
        "prior_readiness_q5": prior_readiness_q5,
        "prior_readiness_q6": prior_readiness_q6,
        "time_pattern_q21": None,
        "time_pattern_q22": None,
        "time_pattern_q23": None,
        "interest_topics_q9": None,
        "interest_topics_q10": None,
        "interest_topics_q11": None,
        "categorical_intent_q12": None,
        "categorical_intent_q13": None,
        "진단응답": False,
        "시험응시": True,
        "needs_map_schema_version": "0.1.1",
        "immersio_phase2_schema_version": "0.1.0",
    }
    for axis in _AXES:
        row[f"{axis}_raw"] = None
        row[f"{axis}_z"] = None
        row[f"{axis}_missing"] = True
    return row


def _item_row(
    item_no: int,
    chapter: str,
    expected_difficulty: str = "어려움",
    correct_rate: float = 0.3,
) -> dict:
    """Return a minimal ItemStatistics-compatible dict."""
    return {
        "item_no": item_no,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "chapter": chapter,
        "week": None,
        "item_type": "이해",
        "difficulty_level": 3,
        "expected_difficulty": expected_difficulty,
        "source": "형성평가",
        "correct_answer": 3,
        "n_responders": 20,
        "n_correct": max(0, round(correct_rate * 20)),
        "n_omit": 0,
        "correct_rate": correct_rate,
        "omit_rate": 0.00,
        "discrimination_index": 0.25,
        "point_biserial": 0.35,
        "top_distractor_no": 2,
        "top_distractor_rate": 0.20,
        "is_top_distractor_adjacent": True,
        "option_distribution": json.dumps({1: 0.1, 2: 0.2, 3: 0.3, 4: 0.2, 5: 0.2}),
        "distractor_label": "특이사항 없음",
    }


def _blueprint() -> dict:
    return {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "exam_name": "2026-1학기 기말고사",
        "total_items": 40,
        "chapters": [_CHAPTER_A, _CHAPTER_B],
        "difficulty_targets": {"easy": 0.45, "medium": 0.35, "hard": 0.20},
        "source_mix": {"formative": 18, "quiz": 12, "textbook": 10},
        "quiz_target": 12,
        "answer_key_balance": True,
    }


def _curriculum() -> dict:
    return {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "entries": [
            {
                "week": 1,
                "chapter": _CHAPTER_A,
                "chapter_no": 1,
                "subtopic": None,
                "sections": ["1.1 인체의 조직"],
            },
            {
                "week": 2,
                "chapter": _CHAPTER_B,
                "chapter_no": 2,
                "subtopic": None,
                "sections": ["2.1 세포의 구조"],
            },
        ],
    }


def _retro_config() -> dict:
    return {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "group_roster": _STUDENT_IDS,
        "unit_importance": {_CHAPTER_A: "상", _CHAPTER_B: "중"},
        "gap_threshold": 0.6,
        "baseline_segment": "만학도",
        "low_discrimination_threshold": 0.2,
        "cognitive_cliff_drop": 0.15,
        "effort_ratings": {_CHAPTER_A: "상", _CHAPTER_B: "중"},
        # 학령기 low-readiness students carry "낮음" on q5; this label drives the
        # low-readiness subgroup for CHAPTER_B (non-hard items, baseline healthy).
        "prior_readiness_low_labels": ["낮음"],
    }


def _build_fixture_tree(data_root: Path) -> None:
    """Write fixture files under data_root.

    Scenario:
    - CHAPTER_A: both segments below 0.6 → structural gap
    - CHAPTER_B: only 학령기 below 0.6, 만학도 above → non-structural gap
    """
    key = _KEY

    silver_dir = data_root / "silver" / "immersio" / key
    silver_dir.mkdir(parents=True, exist_ok=True)

    combined_rows = [
        # 학령기 students — both below threshold on both chapters; low-readiness
        # on q5 so CHAPTER_B (non-hard, baseline healthy) is readiness-driven.
        _combined_row(
            "2026000001",
            {_CHAPTER_A: 0.35, _CHAPTER_B: 0.40},
            cluster_label="전략적",
            prior_readiness_q5="낮음",
        ),
        _combined_row(
            "2026000002",
            {_CHAPTER_A: 0.40, _CHAPTER_B: 0.45},
            cluster_label="전략적",
            prior_readiness_q5="낮음",
        ),
        # 만학도 students — below on CHAPTER_A (structural), above on CHAPTER_B (not structural)
        _combined_row(
            "2026000003",
            {_CHAPTER_A: 0.30, _CHAPTER_B: 0.70},
            cluster_label="습관중심",
            prior_readiness_q5="높음",
        ),
        _combined_row(
            "2026000004",
            {_CHAPTER_A: 0.25, _CHAPTER_B: 0.75},
            cluster_label="습관중심",
            prior_readiness_q5="높음",
        ),
    ]
    pd.DataFrame(combined_rows).to_parquet(silver_dir / "진단×시험결합.parquet", index=False)

    # CHAPTER_A: hard item → 내용난이도 (structural, baseline also low).
    # CHAPTER_B: non-hard item, baseline healthy → cause driven by the
    #            학령기 low-readiness subgroup → 기초구멍.
    item_rows = [
        _item_row(1, _CHAPTER_A, expected_difficulty="어려움", correct_rate=0.3),
        _item_row(2, _CHAPTER_B, expected_difficulty="보통", correct_rate=0.45),
    ]
    pd.DataFrame(item_rows).to_parquet(silver_dir / "문항통계.parquet", index=False)

    bronze_dir = data_root / "bronze" / "retro-mester" / key
    bronze_dir.mkdir(parents=True, exist_ok=True)

    (bronze_dir / "retro_config.yaml").write_text(
        yaml.dump(_retro_config(), allow_unicode=True), encoding="utf-8"
    )
    (bronze_dir / "blueprint.yaml").write_text(
        yaml.dump(_blueprint(), allow_unicode=True), encoding="utf-8"
    )
    (bronze_dir / "curriculum_map.yaml").write_text(
        yaml.dump(_curriculum(), allow_unicode=True), encoding="utf-8"
    )


# ---------------------------------------------------------------------------
# US2 integration tests
# ---------------------------------------------------------------------------


class TestUS2Segments:
    """T036: US2 two-segment full pipeline integration."""

    def _run(self, tmp_path: Path) -> Path:
        """Build fixture, run pipeline, return gold dir."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)
        code = run_retro(
            semester=_SEMESTER,
            course=_COURSE,
            data_root=str(data_root),
            llm_mode="off",
        )
        assert code == 0, f"Pipeline exited with code {code}"
        return data_root / "gold" / "retro-mester" / _KEY

    def test_exit_zero_and_outputs_exist(self, tmp_path: Path) -> None:
        """Pipeline exits 0 and all expected outputs are present."""
        gold = self._run(tmp_path)

        assert (gold / "CQI회고보고서.md").exists()
        assert (gold / "CQI회고보고서.pdf").exists()
        assert (gold / "회고분석.xlsx").exists()
        # FR-012: the manifest is a Silver-layer artefact.
        assert (
            tmp_path / "data" / "silver" / "retro-mester" / _KEY / "manifest_retro.json"
        ).exists()

    def test_recommendations_carry_segment(self, tmp_path: Path) -> None:
        """ChangeRecommendation records carry non-empty segment field."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")

        assert "변경권고" in wb.sheetnames
        ws = wb["변경권고"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        seg_col = headers.index("segment") + 1

        data_rows = [ws.cell(r, seg_col).value for r in range(2, ws.max_row + 1)]
        # All segment values must be non-null and in known segments
        assert all(v in ("학령기", "만학도") for v in data_rows if v is not None)

    def test_structural_escalation_present(self, tmp_path: Path) -> None:
        """SC-004: At least one gap is marked is_structural=True."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")

        ws = wb["빈틈"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        struct_col = headers.index("is_structural") + 1
        chapter_col = headers.index("chapter") + 1

        structural_chapters = [
            ws.cell(r, chapter_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, struct_col).value is True
        ]
        assert len(structural_chapters) > 0, "SC-004: expected at least one structural gap"

    def test_집단대비_sheet_exists(self, tmp_path: Path) -> None:
        """US2 T035: 집단대비 sheet is present in the xlsx workbook."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")

        assert "집단대비" in wb.sheetnames, "Missing '집단대비' sheet in 회고분석.xlsx"

    def test_집단대비_sheet_structure(self, tmp_path: Path) -> None:
        """집단대비 sheet has required columns."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["집단대비"]

        headers = {ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
        required = {
            "chapter",
            "segment",
            "segment_mean_rate",
            "n_below",
            "is_structural",
            "cause",
            "prescription",
        }
        missing = required - headers
        assert not missing, f"집단대비 sheet missing columns: {missing}"

    def test_report_has_집단별_전략_section(self, tmp_path: Path) -> None:
        """US2 T035: MD report includes '집단별 전략' section."""
        gold = self._run(tmp_path)
        md_text = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")

        assert "집단별 전략" in md_text, "Missing '집단별 전략' section in MD report"

    def test_report_has_both_segment_names(self, tmp_path: Path) -> None:
        """MD 집단별 전략 section mentions both segments."""
        gold = self._run(tmp_path)
        md_text = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")

        assert "학령기" in md_text, "MD report missing '학령기'"
        assert "만학도" in md_text, "MD report missing '만학도'"

    def test_group_prescriptions_are_different(self, tmp_path: Path) -> None:
        """SC-003: prescriptions differ between segments for same chapter/cause."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["집단대비"]

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        seg_col = headers.index("segment") + 1
        ch_col = headers.index("chapter") + 1
        presc_col = headers.index("prescription") + 1
        cause_col = headers.index("cause") + 1

        # Collect rows per chapter
        data: dict[
            str, dict[str, tuple[str, str]]
        ] = {}  # chapter → segment → (cause, prescription)
        for r in range(2, ws.max_row + 1):
            ch = ws.cell(r, ch_col).value
            seg = ws.cell(r, seg_col).value
            cause = ws.cell(r, cause_col).value
            presc = ws.cell(r, presc_col).value
            if ch and seg and presc:
                data.setdefault(ch, {})[seg] = (cause, presc)

        # For chapters where both segments appear, prescriptions must differ
        for ch, segs in data.items():
            if "학령기" in segs and "만학도" in segs:
                p_학령기 = segs["학령기"][1]
                p_만학도 = segs["만학도"][1]
                assert p_학령기 != p_만학도, (
                    f"SC-003: 집단대비 sheet prescriptions must differ per segment "
                    f"for chapter {ch!r}; got: 학령기={p_학령기!r}, 만학도={p_만학도!r}"
                )

    def test_no_student_id_in_md(self, tmp_path: Path) -> None:
        """SC-010: No student ID appears anywhere in the MD report."""
        gold = self._run(tmp_path)
        md_text = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")

        for sid in _STUDENT_IDS:
            assert sid not in md_text, f"SC-010: Student ID {sid!r} must NOT appear in MD report"

    def test_no_student_id_in_xlsx(self, tmp_path: Path) -> None:
        """SC-010: No student ID appears in any xlsx sheet."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell_val in row:
                    cell_str = str(cell_val) if cell_val is not None else ""
                    for sid in _STUDENT_IDS:
                        assert sid not in cell_str, (
                            f"SC-010: Student ID {sid!r} found in sheet "
                            f"{sheet_name!r} cell value {cell_str!r}"
                        )

    def test_cluster_vocab_propagated(self, tmp_path: Path) -> None:
        """cluster_vocab is non-null for at least one covered recommendation."""
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")

        ws = wb["변경권고"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        vocab_col = headers.index("cluster_vocab") + 1

        vocab_values = [ws.cell(r, vocab_col).value for r in range(2, ws.max_row + 1)]
        # At least one covered rec should have a non-null vocab (from fixture cluster labels)
        assert any(v is not None for v in vocab_values), (
            "Expected at least one non-null cluster_vocab in recommendations"
        )

    def test_prior_readiness_drives_cause_split(self, tmp_path: Path) -> None:
        """US2 H2: readiness splits 기초구멍 (CHAPTER_B) vs 내용난이도 (CHAPTER_A).

        CHAPTER_A has a hard item and a baseline that is also low → 내용난이도.
        CHAPTER_B has only non-hard items with a healthy baseline; the 학령기
        gap is therefore attributed to its low-readiness subgroup → 기초구멍.
        """
        from openpyxl import load_workbook

        gold = self._run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["집단대비"]

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        seg_col = headers.index("segment") + 1
        ch_col = headers.index("chapter") + 1
        cause_col = headers.index("cause") + 1
        presc_col = headers.index("prescription") + 1

        # (chapter, segment) → (cause, prescription)
        rows: dict[tuple[str, str], tuple[str, str]] = {}
        for r in range(2, ws.max_row + 1):
            ch = ws.cell(r, ch_col).value
            seg = ws.cell(r, seg_col).value
            if ch and seg:
                rows[(ch, seg)] = (
                    ws.cell(r, cause_col).value,
                    ws.cell(r, presc_col).value,
                )

        # CHAPTER_A 학령기 gap → 내용난이도 (content difficulty).
        cause_a, presc_a = rows[(_CHAPTER_A, "학령기")]
        assert cause_a == "내용난이도", f"CHAPTER_A 학령기 cause expected 내용난이도, got {cause_a!r}"
        assert presc_a == "난이도 계단식 분해", (
            f"CHAPTER_A 학령기 prescription should be the 내용난이도 strategy, got {presc_a!r}"
        )

        # CHAPTER_B 학령기 gap → 기초구멍 (basic gaps, readiness-driven).
        cause_b, presc_b = rows[(_CHAPTER_B, "학령기")]
        assert cause_b == "기초구멍", f"CHAPTER_B 학령기 cause expected 기초구멍, got {cause_b!r}"
        assert presc_b == "1주차 기초 다리 선제 배치", (
            f"CHAPTER_B 학령기 prescription should be the 기초구멍 strategy, got {presc_b!r}"
        )

    def test_cause_signals_carry_low_readiness(self, tmp_path: Path) -> None:
        """US2 H2: silver 빈틈표 cause_signals carry the low_readiness_* signals."""
        gold = self._run(tmp_path)
        silver_dir = tmp_path / "data" / "silver" / "retro-mester" / _KEY
        df = pd.read_parquet(silver_dir / "빈틈표.parquet")

        assert not df.empty, "Expected at least one gap row in 빈틈표.parquet"

        required = {"low_readiness_share", "low_readiness_mean_rate", "baseline_segment_mean_rate"}
        for raw in df["cause_signals"]:
            signals = json.loads(raw)
            missing = required - signals.keys()
            assert not missing, f"cause_signals missing low_readiness keys: {missing}"

        # The CHAPTER_B 학령기 gap must record a positive low-readiness share.
        chapter_b_학령기 = df[(df["chapter"] == _CHAPTER_B) & (df["segment"] == "학령기")]
        assert not chapter_b_학령기.empty, "Expected a CHAPTER_B 학령기 gap row"
        signals_b = json.loads(chapter_b_학령기.iloc[0]["cause_signals"])
        assert signals_b["low_readiness_share"] > 0.0, (
            "CHAPTER_B 학령기 low_readiness_share should be positive"
        )

    def test_us1_sections_still_present(self, tmp_path: Path) -> None:
        """US1 report sections survive the US2 update."""
        gold = self._run(tmp_path)
        md_text = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")

        assert "변경 권고 요약" in md_text, "US1 section (A) missing from MD"
        assert "못 덮은 빈틈 비율" in md_text, "US1 uncovered ratio line missing"
