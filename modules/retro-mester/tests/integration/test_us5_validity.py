"""T052 — US5 integration test: assessment-validity gate (SC-006).

RED phase: written before implementation.

Scenario:
  - CHAPTER_BAD: 3 items, all with discrimination_index=0.05 (below threshold 0.2)
    → majority low-disc → validity="문항수선"
  - CHAPTER_OK: 3 items, all with discrimination_index=0.35 (above threshold)
    → validity="건전"

SC-006 assertions:
1. UnitGap.validity for CHAPTER_BAD gaps == "문항수선".
2. ChangeRecommendation.validity for CHAPTER_BAD recs == "문항수선".
3. ChangeRecommendation.prescription_key for CHAPTER_BAD recs ==
   REPAIR_PRESCRIPTION (not a re-teach string).
4. ChangeRecommendation.prescription_key for CHAPTER_OK recs is a
   normal re-teaching prescription (not the repair string).
5. '타당도' sheet is present in 회고분석.xlsx.
6. '타당도' sheet contains rows for CHAPTER_BAD and CHAPTER_OK with
   their respective verdicts.
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
import yaml

# Repair prescription constant — must match implementation.
REPAIR_PRESCRIPTION = "문항 재검토·교체 — 학습 처방 보류"

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_KEY = f"{_SEMESTER}-{_COURSE}"

_CHAPTER_BAD = "1장. 해부학 서론"   # all items low-discrimination → 문항수선
_CHAPTER_OK = "2장. 세포와 조직"   # all items ok → 건전

_STUDENT_IDS = {
    "2026000001": "학령기",
    "2026000002": "학령기",
    "2026000003": "만학도",
    "2026000004": "만학도",
}

_AXES = [
    "digital_efficacy",
    "motivation",
    "time_availability",
    "material_preference",
    "study_strategy",
    "study_environment",
    "social_learning",
    "feedback_seeking",
]


def _combined_row(student_id: str, chapter_rates: dict[str, float]) -> dict:
    """Build a CombinedAnalysisRow-like dict for parquet."""
    row: dict = {
        "student_id": student_id,
        "name_kr": None,
        "on_roster": True,
        "section": None,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "cluster_id": 1,
        "cluster_label": "전략적",
        "cluster_distance": 0.1,
        "exam_taken": True,
        "total_score": 50.0,
        "score_percent": 50.0,
        "section_percentile": 40.0,
        "cohort_percentile": 40.0,
        "z_score": -0.5,
        "chapter_correct_rates": json.dumps(chapter_rates),
        "source_correct_rates": json.dumps({"형성평가": 0.5}),
        "difficulty_correct_rates": json.dumps({"1": 0.6, "2": 0.4, "3": 0.3}),
        "expected_difficulty_correct_rates": json.dumps({"쉬움": 0.6, "보통": 0.4, "어려움": 0.3}),
        "item_type_correct_rates": json.dumps({"지식축적": 0.5, "이해": 0.4}),
        "interest_chapters_correct_rate": None,
        "aversion_chapters_correct_rate": None,
        "prior_readiness_q5": None,
        "prior_readiness_q6": None,
        "time_pattern_q21": None,
        "time_pattern_q22": None,
        "time_pattern_q23": None,
        "interest_topics_q9": None,
        "interest_topics_q10": None,
        "interest_topics_q11": None,
        "categorical_intent_q12": None,
        "categorical_intent_q13": None,
        "진단응답": False,
        "시험응시": True,
        "needs_map_schema_version": "0.1.1",
        "immersio_phase2_schema_version": "0.1.0",
    }
    for axis in _AXES:
        row[f"{axis}_raw"] = None
        row[f"{axis}_z"] = None
        row[f"{axis}_missing"] = True
    return row


def _item_row(
    item_no: int,
    chapter: str,
    discrimination_index: float = 0.30,
    distractor_label: str = "특이사항 없음",
) -> dict:
    """Build an ItemStatistics-like dict for parquet."""
    n_responders = 20
    cr = 0.50
    remainder = round(1.0 - cr, 4)
    each = round(remainder / 4, 4)
    dist = {1: cr, 2: each, 3: each, 4: each, 5: round(remainder - 3 * each, 4)}
    return {
        "item_no": item_no,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "chapter": chapter,
        "week": None,
        "item_type": "이해",
        "difficulty_level": 3,
        "expected_difficulty": "보통",
        "source": "형성평가",
        "correct_answer": 1,
        "n_responders": n_responders,
        "n_correct": 10,
        "n_omit": 0,
        "correct_rate": cr,
        "omit_rate": 0.0,
        "discrimination_index": discrimination_index,
        "point_biserial": 0.35,
        "top_distractor_no": 2,
        "top_distractor_rate": 0.20,
        "is_top_distractor_adjacent": True,
        "option_distribution": json.dumps(dist),
        "distractor_label": distractor_label,
    }


def _build_fixture_tree(data_root: Path) -> None:
    """Build fixture with two chapters: one bad-validity, one healthy."""
    key = _KEY

    silver_dir = data_root / "silver" / "immersio" / key
    silver_dir.mkdir(parents=True, exist_ok=True)

    # All students below gap_threshold (0.6) on both chapters → gaps exist
    combined_rows = [
        _combined_row("2026000001", {_CHAPTER_BAD: 0.30, _CHAPTER_OK: 0.35}),
        _combined_row("2026000002", {_CHAPTER_BAD: 0.25, _CHAPTER_OK: 0.40}),
        _combined_row("2026000003", {_CHAPTER_BAD: 0.35, _CHAPTER_OK: 0.45}),
        _combined_row("2026000004", {_CHAPTER_BAD: 0.20, _CHAPTER_OK: 0.50}),
    ]
    pd.DataFrame(combined_rows).to_parquet(
        silver_dir / "진단×시험결합.parquet", index=False
    )

    # CHAPTER_BAD: 3 items, all low-discrimination (0.05 < threshold 0.2)
    # CHAPTER_OK:  3 items, all good-discrimination (0.35 > threshold 0.2)
    item_rows = [
        _item_row(1, _CHAPTER_BAD, discrimination_index=0.05),
        _item_row(2, _CHAPTER_BAD, discrimination_index=0.08),
        _item_row(3, _CHAPTER_BAD, discrimination_index=0.10),
        _item_row(4, _CHAPTER_OK, discrimination_index=0.35),
        _item_row(5, _CHAPTER_OK, discrimination_index=0.40),
        _item_row(6, _CHAPTER_OK, discrimination_index=0.30),
    ]
    pd.DataFrame(item_rows).to_parquet(silver_dir / "문항통계.parquet", index=False)

    bronze = data_root / "bronze" / "retro-mester" / key
    bronze.mkdir(parents=True, exist_ok=True)

    retro_cfg = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "group_roster": _STUDENT_IDS,
        "unit_importance": {_CHAPTER_BAD: "상", _CHAPTER_OK: "중"},
        "gap_threshold": 0.6,
        "baseline_segment": "만학도",
        "low_discrimination_threshold": 0.2,
        "cognitive_cliff_drop": 0.15,
        "effort_ratings": {_CHAPTER_BAD: "상", _CHAPTER_OK: "중"},
    }
    blueprint = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "exam_name": "2026-1학기 기말고사",
        "total_items": 40,
        "chapters": [_CHAPTER_BAD, _CHAPTER_OK],
        "difficulty_targets": {"easy": 0.45, "medium": 0.35, "hard": 0.20},
        "source_mix": {"formative": 18, "quiz": 12, "textbook": 10},
        "quiz_target": 12,
        "answer_key_balance": True,
    }
    curriculum = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "entries": [
            {"week": 1, "chapter": _CHAPTER_BAD, "chapter_no": 1, "subtopic": None, "sections": ["1.1"]},
            {"week": 2, "chapter": _CHAPTER_OK, "chapter_no": 2, "subtopic": None, "sections": ["2.1"]},
        ],
    }

    (bronze / "retro_config.yaml").write_text(
        yaml.dump(retro_cfg, allow_unicode=True), encoding="utf-8"
    )
    (bronze / "blueprint.yaml").write_text(
        yaml.dump(blueprint, allow_unicode=True), encoding="utf-8"
    )
    (bronze / "curriculum_map.yaml").write_text(
        yaml.dump(curriculum, allow_unicode=True), encoding="utf-8"
    )


def _run(tmp_path: Path) -> tuple[Path, Path]:
    """Build fixture, run pipeline, return (gold_dir, data_root)."""
    from retro_mester.pipeline import run_retro

    data_root = tmp_path / "data"
    _build_fixture_tree(data_root)

    code = run_retro(
        semester=_SEMESTER,
        course=_COURSE,
        data_root=str(data_root),
        llm_mode="off",
    )
    assert code == 0, f"Pipeline exited with code {code}"
    gold = data_root / "gold" / "retro-mester" / _KEY
    return gold, data_root


# -------------------------------------------------------------------------
# SC-006: UnitGap.validity set correctly
# -------------------------------------------------------------------------


class TestSC006UnitGapValidity:
    """SC-006a — UnitGap.validity is enriched from chapter_validity."""

    def test_bad_chapter_gap_validity_is_문항수선(self, tmp_path: Path) -> None:
        """CHAPTER_BAD gaps must have validity='문항수선'."""
        import pyarrow.parquet as pq

        gold, data_root = _run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        gaps_df = pq.read_table(silver / "빈틈표.parquet").to_pandas()

        bad_gaps = gaps_df[gaps_df["chapter"] == _CHAPTER_BAD]
        assert len(bad_gaps) > 0, f"No gaps found for {_CHAPTER_BAD}"
        for _, row in bad_gaps.iterrows():
            assert row["validity"] == "문항수선", (
                f"Expected '문항수선' for {_CHAPTER_BAD} gap, got {row['validity']!r}"
            )

    def test_ok_chapter_gap_validity_is_건전(self, tmp_path: Path) -> None:
        """CHAPTER_OK gaps must have validity='건전'."""
        import pyarrow.parquet as pq

        gold, data_root = _run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        gaps_df = pq.read_table(silver / "빈틈표.parquet").to_pandas()

        ok_gaps = gaps_df[gaps_df["chapter"] == _CHAPTER_OK]
        assert len(ok_gaps) > 0, f"No gaps found for {_CHAPTER_OK}"
        for _, row in ok_gaps.iterrows():
            assert row["validity"] == "건전", (
                f"Expected '건전' for {_CHAPTER_OK} gap, got {row['validity']!r}"
            )


# -------------------------------------------------------------------------
# SC-006: ChangeRecommendation — repair prescription for 문항수선
# -------------------------------------------------------------------------


class TestSC006RecPrescription:
    """SC-006b — 문항수선 recs get repair prescription, not re-teach."""

    def test_bad_chapter_rec_validity_is_문항수선(self, tmp_path: Path) -> None:
        """ChangeRecommendation.validity=='문항수선' for CHAPTER_BAD."""
        import pyarrow.parquet as pq

        gold, data_root = _run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        recs_df = pq.read_table(silver / "변경권고.parquet").to_pandas()

        bad_recs = recs_df[recs_df["chapter"] == _CHAPTER_BAD]
        assert len(bad_recs) > 0, f"No recs found for {_CHAPTER_BAD}"
        for _, row in bad_recs.iterrows():
            assert row["validity"] == "문항수선", (
                f"Expected '문항수선' validity for {_CHAPTER_BAD} rec, "
                f"got {row['validity']!r}"
            )

    def test_bad_chapter_rec_gets_repair_prescription(self, tmp_path: Path) -> None:
        """CHAPTER_BAD recs must carry the repair prescription string (SC-006)."""
        import pyarrow.parquet as pq

        gold, data_root = _run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        recs_df = pq.read_table(silver / "변경권고.parquet").to_pandas()

        bad_recs = recs_df[recs_df["chapter"] == _CHAPTER_BAD]
        assert len(bad_recs) > 0, f"No recs found for {_CHAPTER_BAD}"
        for _, row in bad_recs.iterrows():
            assert row["prescription_key"] == REPAIR_PRESCRIPTION, (
                f"Expected repair prescription for {_CHAPTER_BAD}, "
                f"got {row['prescription_key']!r}"
            )

    def test_ok_chapter_rec_does_not_get_repair_prescription(
        self, tmp_path: Path
    ) -> None:
        """CHAPTER_OK recs must NOT carry the repair prescription (normal re-teach)."""
        import pyarrow.parquet as pq

        gold, data_root = _run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        recs_df = pq.read_table(silver / "변경권고.parquet").to_pandas()

        ok_recs = recs_df[recs_df["chapter"] == _CHAPTER_OK]
        assert len(ok_recs) > 0, f"No recs found for {_CHAPTER_OK}"
        for _, row in ok_recs.iterrows():
            assert row["prescription_key"] != REPAIR_PRESCRIPTION, (
                f"CHAPTER_OK rec should NOT have repair prescription, "
                f"got {row['prescription_key']!r}"
            )

    def test_bad_chapter_rec_prescription_not_a_reteach_string(
        self, tmp_path: Path
    ) -> None:
        """CHAPTER_BAD prescriptions must not be re-teaching catalogue entries."""
        import pyarrow.parquet as pq

        gold, data_root = _run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        recs_df = pq.read_table(silver / "변경권고.parquet").to_pandas()

        # These are the re-teaching prescriptions from the catalogue.
        reteach_prescriptions = {
            "1주차 기초 다리 선제 배치",
            "난이도 계단식 분해",
            "수업 관찰 후 재진단",
            "단편지식 재구조화 스캐폴딩",
            "핵심 개념 반복·속도 배려",
            "개별 면담 후 재진단",
            "담당 교수 협의 후 재진단",
        }
        bad_recs = recs_df[recs_df["chapter"] == _CHAPTER_BAD]
        for _, row in bad_recs.iterrows():
            assert row["prescription_key"] not in reteach_prescriptions, (
                f"CHAPTER_BAD rec must not have a re-teach prescription, "
                f"got {row['prescription_key']!r}"
            )


# -------------------------------------------------------------------------
# T051: 타당도 sheet
# -------------------------------------------------------------------------


class TestSC006ValiditySheet:
    """T051 — 타당도 sheet present in xlsx and contains correct data."""

    def test_타당도_sheet_exists(self, tmp_path: Path) -> None:
        """회고분석.xlsx must contain a '타당도' sheet."""
        from openpyxl import load_workbook

        gold, _ = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        assert "타당도" in wb.sheetnames, (
            f"'타당도' sheet missing; sheets: {wb.sheetnames}"
        )

    def test_타당도_sheet_header_columns(self, tmp_path: Path) -> None:
        """타당도 sheet header must include expected columns."""
        from openpyxl import load_workbook

        gold, _ = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["타당도"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for col in ("chapter", "verdict", "mean_discrimination", "low_disc_share", "bad_distractor_share"):
            assert col in headers, f"Missing column '{col}' in 타당도 header: {headers}"

    def test_타당도_sheet_has_bad_chapter_row(self, tmp_path: Path) -> None:
        """타당도 sheet must have a row for CHAPTER_BAD with verdict '문항수선'."""
        from openpyxl import load_workbook

        gold, _ = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["타당도"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        chapter_col = headers.index("chapter") + 1
        verdict_col = headers.index("verdict") + 1

        rows_found: dict[str, str] = {}
        for r in range(2, ws.max_row + 1):
            ch = ws.cell(r, chapter_col).value
            verd = ws.cell(r, verdict_col).value
            if ch:
                rows_found[ch] = verd

        assert _CHAPTER_BAD in rows_found, (
            f"'{_CHAPTER_BAD}' not found in 타당도 sheet; rows: {rows_found}"
        )
        assert rows_found[_CHAPTER_BAD] == "문항수선", (
            f"Expected '문항수선' for {_CHAPTER_BAD}, got {rows_found[_CHAPTER_BAD]!r}"
        )

    def test_타당도_sheet_has_ok_chapter_row(self, tmp_path: Path) -> None:
        """타당도 sheet must have a row for CHAPTER_OK with verdict '건전'."""
        from openpyxl import load_workbook

        gold, _ = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["타당도"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        chapter_col = headers.index("chapter") + 1
        verdict_col = headers.index("verdict") + 1

        rows_found: dict[str, str] = {}
        for r in range(2, ws.max_row + 1):
            ch = ws.cell(r, chapter_col).value
            verd = ws.cell(r, verdict_col).value
            if ch:
                rows_found[ch] = verd

        assert _CHAPTER_OK in rows_found, (
            f"'{_CHAPTER_OK}' not found in 타당도 sheet; rows: {rows_found}"
        )
        assert rows_found[_CHAPTER_OK] == "건전", (
            f"Expected '건전' for {_CHAPTER_OK}, got {rows_found[_CHAPTER_OK]!r}"
        )
