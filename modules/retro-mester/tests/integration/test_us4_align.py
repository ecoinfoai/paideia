"""T048 — US4 alignment + cognitive-cliff integration tests.

RED phase: written before implementation.

Verifies:
1. Pipeline produces a '정렬' xlsx sheet.
2. Markdown report includes '인지수준·정렬' section.
3. figs/*.png files exist in the gold dir.
4. UnitGap.cohort_failing_item_types is populated (non-empty) for chapters
   with a cognitive cliff.
5. At least one ChangeRecommendation.target_cognitive_level != '미상' when
   a cliff exists.
6. PNG determinism: two identical runs produce byte-identical PNGs.
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import yaml

# ---------------------------------------------------------------------------
# Fixture constants
# ---------------------------------------------------------------------------

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_KEY = f"{_SEMESTER}-{_COURSE}"

_CHAPTER_A = "1장. 해부학 서론"    # will have cognitive cliff
_CHAPTER_B = "2장. 세포와 조직"    # balanced, no cliff

_STUDENT_IDS = {
    "2026000001": "학령기",
    "2026000002": "학령기",
    "2026000003": "만학도",
    "2026000004": "만학도",
}

_AXES = [
    "digital_efficacy",
    "motivation",
    "time_availability",
    "material_preference",
    "study_strategy",
    "study_environment",
    "social_learning",
    "feedback_seeking",
]


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _combined_row(
    student_id: str,
    chapter_rates: dict[str, float],
    cluster_label: str | None = None,
    interest_rate: float | None = None,
    aversion_rate: float | None = None,
) -> dict:
    has_cluster = cluster_label is not None
    row: dict = {
        "student_id": student_id,
        "name_kr": None,
        "on_roster": True,
        "section": None,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "cluster_id": 1 if has_cluster else None,
        "cluster_label": cluster_label,
        "cluster_distance": 0.1 if has_cluster else None,
        "exam_taken": True,
        "total_score": 60.0,
        "score_percent": 60.0,
        "section_percentile": 50.0,
        "cohort_percentile": 50.0,
        "z_score": 0.0,
        "chapter_correct_rates": json.dumps(chapter_rates),
        "source_correct_rates": json.dumps({"형성평가": 0.5}),
        "difficulty_correct_rates": json.dumps({"1": 0.7, "2": 0.5, "3": 0.3}),
        "expected_difficulty_correct_rates": json.dumps({"쉬움": 0.7, "보통": 0.5, "어려움": 0.3}),
        # item_type_correct_rates: 지식축적 high, 이해 low → cliff signal per student
        "item_type_correct_rates": json.dumps({"지식축적": 0.82, "이해": 0.45}),
        "interest_chapters_correct_rate": interest_rate,
        "aversion_chapters_correct_rate": aversion_rate,
        "prior_readiness_q5": None,
        "prior_readiness_q6": None,
        "time_pattern_q21": None,
        "time_pattern_q22": None,
        "time_pattern_q23": None,
        "interest_topics_q9": None,
        "interest_topics_q10": None,
        "interest_topics_q11": None,
        "categorical_intent_q12": None,
        "categorical_intent_q13": None,
        "진단응답": False,
        "시험응시": True,
        "needs_map_schema_version": "0.1.1",
        "immersio_phase2_schema_version": "0.1.0",
    }
    for axis in _AXES:
        row[f"{axis}_raw"] = None
        row[f"{axis}_z"] = None
        row[f"{axis}_missing"] = True
    return row


def _item_row(
    item_no: int,
    chapter: str,
    item_type: str = "이해",
    correct_rate: float = 0.3,
) -> dict:
    cr = correct_rate
    remainder = round(1.0 - cr, 4)
    each = round(remainder / 4, 4)
    dist = {1: cr, 2: each, 3: each, 4: each, 5: round(remainder - 3 * each, 4)}
    return {
        "item_no": item_no,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "chapter": chapter,
        "week": None,
        "item_type": item_type,
        "difficulty_level": 3,
        "expected_difficulty": "어려움",
        "source": "형성평가",
        "correct_answer": 3,
        "n_responders": 20,
        "n_correct": max(0, round(cr * 20)),
        "n_omit": 0,
        "correct_rate": cr,
        "omit_rate": 0.00,
        "discrimination_index": 0.25,
        "point_biserial": 0.35,
        "top_distractor_no": 2,
        "top_distractor_rate": 0.20,
        "is_top_distractor_adjacent": True,
        "option_distribution": json.dumps(dist),
        "distractor_label": "특이사항 없음",
    }


def _build_fixture_tree(data_root: Path) -> None:
    """Build fixture with cliff scenario for CHAPTER_A (1장).

    CHAPTER_A item types:
      - 지식축적: correct_rate=0.82 (high)
      - 이해: correct_rate=0.45 (low; 0.82 - 0.45 = 0.37 > 0.15 → cliff)

    CHAPTER_B item types:
      - 지식축적: correct_rate=0.70 (balanced)
      - 이해: correct_rate=0.65 (balanced; 0.70 - 0.65 = 0.05 < 0.15 → no cliff)
    """
    key = _KEY

    silver_dir = data_root / "silver" / "immersio" / key
    silver_dir.mkdir(parents=True, exist_ok=True)

    combined_rows = [
        # All students below gap threshold to guarantee gaps exist
        _combined_row("2026000001", {_CHAPTER_A: 0.35, _CHAPTER_B: 0.40},
                      cluster_label="전략적", interest_rate=0.70, aversion_rate=0.45),
        _combined_row("2026000002", {_CHAPTER_A: 0.40, _CHAPTER_B: 0.45},
                      cluster_label="전략적", interest_rate=0.65, aversion_rate=0.40),
        _combined_row("2026000003", {_CHAPTER_A: 0.30, _CHAPTER_B: 0.70},
                      cluster_label="습관중심"),
        _combined_row("2026000004", {_CHAPTER_A: 0.25, _CHAPTER_B: 0.75},
                      cluster_label="습관중심"),
    ]
    pd.DataFrame(combined_rows).to_parquet(
        silver_dir / "진단×시험결합.parquet", index=False
    )

    # Items: CHAPTER_A has cliff (지식축적 high, 이해 low)
    #        CHAPTER_B is balanced
    item_rows = [
        _item_row(1, _CHAPTER_A, item_type="지식축적", correct_rate=0.82),
        _item_row(2, _CHAPTER_A, item_type="이해", correct_rate=0.45),
        _item_row(3, _CHAPTER_A, item_type="이해", correct_rate=0.42),
        _item_row(4, _CHAPTER_B, item_type="지식축적", correct_rate=0.70),
        _item_row(5, _CHAPTER_B, item_type="이해", correct_rate=0.65),
    ]
    pd.DataFrame(item_rows).to_parquet(silver_dir / "문항통계.parquet", index=False)

    bronze = data_root / "bronze" / "retro-mester" / key
    bronze.mkdir(parents=True, exist_ok=True)

    retro_cfg = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "group_roster": _STUDENT_IDS,
        "unit_importance": {_CHAPTER_A: "상", _CHAPTER_B: "중"},
        "gap_threshold": 0.6,
        "baseline_segment": "만학도",
        "low_discrimination_threshold": 0.2,
        "cognitive_cliff_drop": 0.15,
        "effort_ratings": {_CHAPTER_A: "상", _CHAPTER_B: "중"},
    }
    blueprint = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "exam_name": "2026-1학기 기말고사",
        "total_items": 40,
        "chapters": [_CHAPTER_A, _CHAPTER_B],
        "difficulty_targets": {"easy": 0.45, "medium": 0.35, "hard": 0.20},
        "source_mix": {"formative": 18, "quiz": 12, "textbook": 10},
        "quiz_target": 12,
        "answer_key_balance": True,
    }
    curriculum = {
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "entries": [
            {"week": 1, "chapter": _CHAPTER_A, "chapter_no": 1, "subtopic": None, "sections": ["1.1"]},
            {"week": 2, "chapter": _CHAPTER_B, "chapter_no": 2, "subtopic": None, "sections": ["2.1"]},
        ],
    }

    (bronze / "retro_config.yaml").write_text(yaml.dump(retro_cfg, allow_unicode=True), encoding="utf-8")
    (bronze / "blueprint.yaml").write_text(yaml.dump(blueprint, allow_unicode=True), encoding="utf-8")
    (bronze / "curriculum_map.yaml").write_text(yaml.dump(curriculum, allow_unicode=True), encoding="utf-8")


def _run(tmp_path: Path) -> Path:
    """Build fixture, run pipeline, return gold dir."""
    from retro_mester.pipeline import run_retro

    data_root = tmp_path / "data"
    if not (data_root / "silver").exists():
        _build_fixture_tree(data_root)

    code = run_retro(
        semester=_SEMESTER,
        course=_COURSE,
        data_root=str(data_root),
        llm_mode="off",
    )
    assert code == 0, f"Pipeline exited with code {code}"
    return data_root / "gold" / "retro-mester" / _KEY


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestUS4XlsxAlignSheet:
    """T047 — 정렬 sheet exists in xlsx."""

    def test_정렬_sheet_exists(self, tmp_path: Path) -> None:
        """회고분석.xlsx contains a '정렬' sheet."""
        from openpyxl import load_workbook

        gold = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        assert "정렬" in wb.sheetnames

    def test_정렬_sheet_has_header_row(self, tmp_path: Path) -> None:
        """정렬 sheet has a header row with 'chapter' column."""
        from openpyxl import load_workbook

        gold = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["정렬"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "chapter" in headers

    def test_정렬_sheet_has_data_rows(self, tmp_path: Path) -> None:
        """정렬 sheet has at least one data row."""
        from openpyxl import load_workbook

        gold = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["정렬"]
        assert ws.max_row >= 2, "No data rows in 정렬 sheet"

    def test_정렬_sheet_has_flag_column(self, tmp_path: Path) -> None:
        """정렬 sheet has a 'flag' column."""
        from openpyxl import load_workbook

        gold = _run(tmp_path)
        wb = load_workbook(gold / "회고분석.xlsx")
        ws = wb["정렬"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "flag" in headers


class TestUS4ReportSection:
    """T047 — 인지수준·정렬 section in Markdown report."""

    def test_report_has_인지수준_정렬_section(self, tmp_path: Path) -> None:
        """CQI회고보고서.md contains '인지수준·정렬' section."""
        gold = _run(tmp_path)
        md = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")
        assert "인지수준" in md and "정렬" in md

    def test_report_mentions_chapter_with_cliff(self, tmp_path: Path) -> None:
        """Alignment section mentions the cliff chapter (1장)."""
        gold = _run(tmp_path)
        md = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")
        assert "1장" in md or "해부학 서론" in md


class TestUS4FiguresExist:
    """T047 — PNG figures exist in gold figs/ directory."""

    def test_figs_directory_exists(self, tmp_path: Path) -> None:
        """gold/figs/ directory is created."""
        gold = _run(tmp_path)
        assert (gold / "figs").is_dir()

    def test_at_least_one_png_exists(self, tmp_path: Path) -> None:
        """At least one .png file is written to figs/."""
        gold = _run(tmp_path)
        pngs = list((gold / "figs").glob("*.png"))
        assert len(pngs) >= 1, "No PNG files found in figs/"

    def test_png_has_nonzero_size(self, tmp_path: Path) -> None:
        """Each PNG file is non-empty."""
        gold = _run(tmp_path)
        for png in (gold / "figs").glob("*.png"):
            assert png.stat().st_size > 0, f"Empty PNG: {png}"


class TestUS4CognitiveLevelPopulated:
    """T044/T045/Enrich — cohort_failing_item_types and target_cognitive_level filled."""

    def test_unit_gap_cohort_failing_item_types_populated(self, tmp_path: Path) -> None:
        """UnitGap.cohort_failing_item_types is non-empty for CHAPTER_A (cliff exists)."""
        import pyarrow.parquet as pq

        _run(tmp_path)
        # Read silver parquet to check gaps
        data_root = tmp_path / "data"
        silver = data_root / "silver" / "retro-mester" / _KEY
        gaps_df = pq.read_table(silver / "빈틈표.parquet").to_pandas()

        # Filter for CHAPTER_A gaps
        cliff_gaps = gaps_df[gaps_df["chapter"] == _CHAPTER_A]
        assert len(cliff_gaps) > 0, f"No gaps for {_CHAPTER_A}"

        # At least one gap should have cohort_failing_item_types populated
        any_populated = any(
            row != "[]" and row != "" and row is not None
            for row in cliff_gaps["cohort_failing_item_types"].astype(str)
        )
        assert any_populated, (
            f"cohort_failing_item_types still empty for {_CHAPTER_A} gaps; "
            f"values: {cliff_gaps['cohort_failing_item_types'].tolist()}"
        )

    def test_recommendation_target_cognitive_level_not_미상(self, tmp_path: Path) -> None:
        """At least one ChangeRecommendation.target_cognitive_level != '미상'."""
        import pyarrow.parquet as pq

        _run(tmp_path)
        data_root = tmp_path / "data"
        silver = data_root / "silver" / "retro-mester" / _KEY
        recs_df = pq.read_table(silver / "변경권고.parquet").to_pandas()

        non_미상 = (recs_df["target_cognitive_level"] != "미상").any()
        assert non_미상, (
            "All target_cognitive_level are still '미상'; "
            f"values: {recs_df['target_cognitive_level'].tolist()}"
        )


class TestUS4PNGDeterminism:
    """T047 — PNG files are deterministic across runs."""

    def test_png_byte_identical_across_runs(self, tmp_path: Path) -> None:
        """Two identical pipeline runs produce byte-identical PNG files."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)

        # Run 1
        code1 = run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))
        assert code1 == 0

        gold = data_root / "gold" / "retro-mester" / _KEY
        figs_dir = gold / "figs"
        pngs = list(figs_dir.glob("*.png"))
        assert len(pngs) >= 1, "No PNGs after run 1"

        # Collect bytes from run 1
        bytes_run1: dict[str, bytes] = {p.name: p.read_bytes() for p in pngs}

        # Run 2 — archives existing gold, rewrites
        code2 = run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))
        assert code2 == 0

        # figs are recreated in the fresh gold dir
        pngs2 = list(figs_dir.glob("*.png"))
        bytes_run2: dict[str, bytes] = {p.name: p.read_bytes() for p in pngs2}

        assert set(bytes_run1) == set(bytes_run2), "Different PNG files between runs"
        for name in bytes_run1:
            assert bytes_run1[name] == bytes_run2[name], (
                f"PNG {name} is not byte-identical between runs"
            )
