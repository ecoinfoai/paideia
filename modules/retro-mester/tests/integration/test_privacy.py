"""T059a — Privacy integration tests for retro-mester (FR-026 + 헌장 V).

Verifies:
- All retro-mester outputs are written only under data/
- Repo-root .gitignore includes data/
- No student IDs appear in gold CQI회고보고서.md, 회고분석.xlsx, 차년도방향.yaml
- No secret-like tokens appear in gold outputs
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
import pandas as pd
import yaml

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_KEY = f"{_SEMESTER}-{_COURSE}"

_CHAPTER_A = "1장. 해부학 서론"
_CHAPTER_B = "2장. 세포와 조직"

_STUDENT_IDS = ["2026000001", "2026000002", "2026000003", "2026000004"]

_AXES = [
    "digital_efficacy", "motivation", "time_availability", "material_preference",
    "study_strategy", "study_environment", "social_learning", "feedback_seeking",
]

# Pattern that looks like a secret/token (long hex or base64-like string)
_SECRET_PATTERN = re.compile(r"[A-Za-z0-9+/]{32,}={0,2}")


# ---------------------------------------------------------------------------
# Fixture helpers (self-contained)
# ---------------------------------------------------------------------------


def _combined_row(
    student_id: str,
    chapter_rates: dict[str, float] | None = None,
) -> dict:
    if chapter_rates is None:
        chapter_rates = {_CHAPTER_A: 0.4, _CHAPTER_B: 0.3}
    row: dict = {
        "student_id": student_id,
        "name_kr": None,
        "on_roster": True,
        "section": None,
        "semester": _SEMESTER,
        "course_slug": _COURSE,
        "cluster_id": None,
        "cluster_label": None,
        "cluster_distance": None,
        "exam_taken": True,
        "total_score": 60.0,
        "score_percent": 60.0,
        "section_percentile": 50.0,
        "cohort_percentile": 50.0,
        "z_score": 0.0,
        "chapter_correct_rates": json.dumps(chapter_rates),
        "source_correct_rates": json.dumps({"형성평가": 0.5}),
        "difficulty_correct_rates": json.dumps({"1": 0.7, "2": 0.5, "3": 0.3}),
        "expected_difficulty_correct_rates": json.dumps(
            {"쉬움": 0.7, "보통": 0.5, "어려움": 0.3}
        ),
        "item_type_correct_rates": json.dumps({"지식축적": 0.6, "이해": 0.5}),
        "interest_chapters_correct_rate": None,
        "aversion_chapters_correct_rate": None,
        "prior_readiness_q5": None,
        "prior_readiness_q6": None,
        "time_pattern_q21": None,
        "time_pattern_q22": None,
        "time_pattern_q23": None,
        "interest_topics_q9": None,
        "interest_topics_q10": None,
        "interest_topics_q11": None,
        "categorical_intent_q12": None,
        "categorical_intent_q13": None,
        "진단응답": False,
        "시험응시": True,
        "needs_map_schema_version": "0.1.1",
        "immersio_phase2_schema_version": "0.1.0",
    }
    for axis in _AXES:
        row[f"{axis}_raw"] = None
        row[f"{axis}_z"] = None
        row[f"{axis}_missing"] = True
    return row


def _item_row(item_no: int, chapter: str = _CHAPTER_A) -> dict:
    return {
        "item_no": item_no, "semester": _SEMESTER, "course_slug": _COURSE,
        "chapter": chapter, "week": None, "item_type": "지식축적",
        "difficulty_level": 2, "expected_difficulty": "보통", "source": "형성평가",
        "correct_answer": 3, "n_responders": 20, "n_correct": 10, "n_omit": 0,
        "correct_rate": 0.50, "omit_rate": 0.00, "discrimination_index": 0.25,
        "point_biserial": 0.35, "top_distractor_no": 2, "top_distractor_rate": 0.20,
        "is_top_distractor_adjacent": True,
        "option_distribution": json.dumps({1: 0.1, 2: 0.2, 3: 0.5, 4: 0.1, 5: 0.1}),
        "distractor_label": "특이사항 없음",
    }


def _blueprint() -> dict:
    return {
        "semester": _SEMESTER, "course_slug": _COURSE, "exam_name": "기말고사",
        "total_items": 40, "chapters": [_CHAPTER_A, _CHAPTER_B],
        "difficulty_targets": {"easy": 0.45, "medium": 0.35, "hard": 0.20},
        "source_mix": {"formative": 18, "quiz": 12, "textbook": 10},
        "quiz_target": 12, "answer_key_balance": True,
    }


def _curriculum() -> dict:
    return {
        "semester": _SEMESTER, "course_slug": _COURSE,
        "entries": [
            {"week": 1, "chapter": _CHAPTER_A, "chapter_no": 1,
             "subtopic": None, "sections": ["1.1 인체의 조직"]},
            {"week": 2, "chapter": _CHAPTER_B, "chapter_no": 2,
             "subtopic": None, "sections": ["2.1 세포의 구조"]},
        ],
    }


def _retro_config() -> dict:
    return {
        "semester": _SEMESTER, "course_slug": _COURSE,
        "group_roster": {
            "2026000001": "학령기", "2026000002": "학령기",
            "2026000003": "만학도", "2026000004": "만학도",
        },
        "unit_importance": {_CHAPTER_A: "상", _CHAPTER_B: "중"},
        "gap_threshold": 0.6, "baseline_segment": "만학도",
        "low_discrimination_threshold": 0.2, "cognitive_cliff_drop": 0.15,
        "effort_ratings": {_CHAPTER_A: "상", _CHAPTER_B: "중"},
    }


def _build_fixture_and_run(data_root: Path) -> None:
    silver_im = data_root / "silver" / "immersio" / _KEY
    silver_im.mkdir(parents=True, exist_ok=True)

    combined = [
        _combined_row("2026000001", {_CHAPTER_A: 0.4, _CHAPTER_B: 0.3}),
        _combined_row("2026000002", {_CHAPTER_A: 0.5, _CHAPTER_B: 0.35}),
        _combined_row("2026000003", {_CHAPTER_A: 0.45, _CHAPTER_B: 0.25}),
        _combined_row("2026000004", {_CHAPTER_A: 0.4, _CHAPTER_B: 0.2}),
    ]
    pd.DataFrame(combined).to_parquet(silver_im / "진단×시험결합.parquet", index=False)

    items = [_item_row(1, _CHAPTER_A), _item_row(2, _CHAPTER_B)]
    pd.DataFrame(items).to_parquet(silver_im / "문항통계.parquet", index=False)

    bronze = data_root / "bronze" / "retro-mester" / _KEY
    bronze.mkdir(parents=True, exist_ok=True)
    (bronze / "retro_config.yaml").write_text(
        yaml.dump(_retro_config(), allow_unicode=True), encoding="utf-8"
    )
    (bronze / "blueprint.yaml").write_text(
        yaml.dump(_blueprint(), allow_unicode=True), encoding="utf-8"
    )
    (bronze / "curriculum_map.yaml").write_text(
        yaml.dump(_curriculum(), allow_unicode=True), encoding="utf-8"
    )

    from retro_mester.pipeline import run_retro
    code = run_retro(
        semester=_SEMESTER,
        course=_COURSE,
        data_root=str(data_root),
        llm_mode="off",
    )
    assert code == 0, f"Pipeline failed with exit code {code}"


# ---------------------------------------------------------------------------
# FR-026: all outputs under data/
# ---------------------------------------------------------------------------


class TestOutputUnderDataDir:
    """All retro-mester outputs must be written under the data/ root (FR-026)."""

    def test_gold_outputs_under_data(self, tmp_path: Path) -> None:
        """Gold artefacts reside strictly within data_root/gold/."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        gold = data_root / "gold" / "retro-mester" / _KEY
        for artefact in ["CQI회고보고서.md", "CQI회고보고서.pdf", "회고분석.xlsx",
                         "manifest_retro.json", "차년도방향.yaml"]:
            p = gold / artefact
            assert p.exists(), f"Missing gold artefact: {artefact}"
            assert str(p).startswith(str(data_root)), (
                f"{artefact} is outside data_root: {p}"
            )

    def test_silver_outputs_under_data(self, tmp_path: Path) -> None:
        """Silver parquet files reside strictly within data_root/silver/."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        silver = data_root / "silver" / "retro-mester" / _KEY
        for artefact in ["빈틈표.parquet", "변경권고.parquet"]:
            p = silver / artefact
            assert p.exists(), f"Missing silver artefact: {artefact}"
            assert str(p).startswith(str(data_root)), (
                f"{artefact} is outside data_root: {p}"
            )


# ---------------------------------------------------------------------------
# .gitignore: data/ must be excluded
# ---------------------------------------------------------------------------


class TestGitignore:
    """Repo-root .gitignore must include data/ (FR-026)."""

    def test_gitignore_excludes_data_dir(self) -> None:
        """The repo-root .gitignore contains a data/ rule."""
        repo_root = Path(__file__).parents[4]  # climb to paideia root
        gitignore = repo_root / ".gitignore"
        assert gitignore.exists(), f"No .gitignore found at {gitignore}"

        content = gitignore.read_text(encoding="utf-8")
        lines = [ln.strip() for ln in content.splitlines()]
        # Accept 'data/' or 'data' as valid exclusion patterns
        data_excluded = any(
            ln in ("data/", "data") or ln.startswith("data/")
            for ln in lines
            if not ln.startswith("#") and ln
        )
        assert data_excluded, (
            f"data/ is not excluded in {gitignore}. "
            "Add 'data/' to .gitignore to prevent student PII from being committed."
        )


# ---------------------------------------------------------------------------
# Privacy: no student IDs in gold outputs
# ---------------------------------------------------------------------------


class TestNoStudentIdsInGoldOutputs:
    """Gold outputs must not contain individual student IDs (FR-026, SC-003)."""

    def test_no_student_id_in_md_report(self, tmp_path: Path) -> None:
        """CQI회고보고서.md must not contain any student ID."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        md_text = (
            data_root / "gold" / "retro-mester" / _KEY / "CQI회고보고서.md"
        ).read_text(encoding="utf-8")

        for sid in _STUDENT_IDS:
            assert sid not in md_text, (
                f"Student ID {sid} must NOT appear in CQI회고보고서.md"
            )

    def test_no_student_id_in_xlsx(self, tmp_path: Path) -> None:
        """회고분석.xlsx must not contain any student ID in any cell."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        xlsx_path = (
            data_root / "gold" / "retro-mester" / _KEY / "회고분석.xlsx"
        )
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value is None:
                        continue
                    cell_str = str(cell_value)
                    for sid in _STUDENT_IDS:
                        assert sid not in cell_str, (
                            f"Student ID {sid} found in xlsx sheet '{sheet_name}'"
                        )
        wb.close()

    def test_no_student_id_in_forward_yaml(self, tmp_path: Path) -> None:
        """차년도방향.yaml must not contain any student ID."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        yaml_text = (
            data_root / "gold" / "retro-mester" / _KEY / "차년도방향.yaml"
        ).read_text(encoding="utf-8")

        for sid in _STUDENT_IDS:
            assert sid not in yaml_text, (
                f"Student ID {sid} must NOT appear in 차년도방향.yaml"
            )


# ---------------------------------------------------------------------------
# Privacy: no secret-like tokens in gold outputs
# ---------------------------------------------------------------------------


class TestNoSecretsInGoldOutputs:
    """Gold text outputs must not contain hardcoded secret-like tokens."""

    def test_no_secrets_in_md_report(self, tmp_path: Path) -> None:
        """CQI회고보고서.md contains no long hex/base64 token-like strings."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        md_text = (
            data_root / "gold" / "retro-mester" / _KEY / "CQI회고보고서.md"
        ).read_text(encoding="utf-8")

        # Look for long opaque tokens (exclude Korean content and common base words)
        suspicious = [
            m.group()
            for m in _SECRET_PATTERN.finditer(md_text)
            # Exclude SHA-256 digests (64 hex chars) from manifest — they're content hashes
            # and exclude short version strings like "0.1.0" → keep only truly long tokens
            if len(m.group()) > 40
        ]
        assert len(suspicious) == 0, (
            f"Secret-like tokens found in CQI회고보고서.md: {suspicious[:3]}"
        )

    def test_no_secrets_in_forward_yaml(self, tmp_path: Path) -> None:
        """차년도방향.yaml contains no long hex/base64 token-like strings."""
        data_root = tmp_path / "data"
        _build_fixture_and_run(data_root)

        yaml_text = (
            data_root / "gold" / "retro-mester" / _KEY / "차년도방향.yaml"
        ).read_text(encoding="utf-8")

        suspicious = [
            m.group()
            for m in _SECRET_PATTERN.finditer(yaml_text)
            if len(m.group()) > 40
        ]
        assert len(suspicious) == 0, (
            f"Secret-like tokens found in 차년도방향.yaml: {suspicious[:3]}"
        )
