"""T030 — US1 end-to-end integration tests for the retro-mester pipeline.

RED phase: written before pipeline.py exists.  All tests must FAIL until
``retro_mester.pipeline.run_retro`` is implemented.

Tests build a minimal but complete fixture tree under ``tmp_path/data/`` and
call ``run_retro(..., data_root=tmp_path / "data")``.

Fixture layout:
  data/silver/immersio/{key}/진단×시험결합.parquet
  data/silver/immersio/{key}/문항통계.parquet
  data/bronze/retro-mester/{key}/retro_config.yaml
  data/bronze/retro-mester/{key}/blueprint.yaml
  data/bronze/retro-mester/{key}/curriculum_map.yaml
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import yaml

# ---------------------------------------------------------------------------
# Fixture helpers (shared with test_us1_determinism.py via inline duplication
# — no conftest import so each test module is self-contained)
# ---------------------------------------------------------------------------

_SEMESTER = "2026-1"
_COURSE = "anatomy"
_KEY = f"{_SEMESTER}-{_COURSE}"

_CHAPTER_A = "1장. 해부학 서론"
_CHAPTER_B = "2장. 세포와 조직"


def _combined_row(
    student_id: str,
    chapter_rates: dict[str, float] | None = None,
    *,
    semester: str = _SEMESTER,
    course_slug: str = _COURSE,
) -> dict:
    """Return a minimal CombinedAnalysisRow-compatible dict."""
    if chapter_rates is None:
        chapter_rates = {_CHAPTER_A: 0.4, _CHAPTER_B: 0.3}

    axes = [
        "digital_efficacy",
        "motivation",
        "time_availability",
        "material_preference",
        "study_strategy",
        "study_environment",
        "social_learning",
        "feedback_seeking",
    ]
    row: dict = {
        "student_id": student_id,
        "name_kr": None,
        "on_roster": True,
        "section": None,
        "semester": semester,
        "course_slug": course_slug,
        "cluster_id": None,
        "cluster_label": None,
        "cluster_distance": None,
        "exam_taken": True,
        "total_score": 60.0,
        "score_percent": 60.0,
        "section_percentile": 50.0,
        "cohort_percentile": 50.0,
        "z_score": 0.0,
        "chapter_correct_rates": json.dumps(chapter_rates),
        "source_correct_rates": json.dumps({"형성평가": 0.5}),
        "difficulty_correct_rates": json.dumps({"1": 0.7, "2": 0.5, "3": 0.3}),
        "expected_difficulty_correct_rates": json.dumps({"쉬움": 0.7, "보통": 0.5, "어려움": 0.3}),
        "item_type_correct_rates": json.dumps({"지식축적": 0.6, "이해": 0.5}),
        "interest_chapters_correct_rate": None,
        "aversion_chapters_correct_rate": None,
        "prior_readiness_q5": None,
        "prior_readiness_q6": None,
        "time_pattern_q21": None,
        "time_pattern_q22": None,
        "time_pattern_q23": None,
        "interest_topics_q9": None,
        "interest_topics_q10": None,
        "interest_topics_q11": None,
        "categorical_intent_q12": None,
        "categorical_intent_q13": None,
        "진단응답": False,
        "시험응시": True,
        "needs_map_schema_version": "0.1.1",
        "immersio_phase2_schema_version": "0.1.0",
    }
    for axis in axes:
        row[f"{axis}_raw"] = None
        row[f"{axis}_z"] = None
        row[f"{axis}_missing"] = True
    return row


def _item_row(
    item_no: int,
    chapter: str = _CHAPTER_A,
    *,
    semester: str = _SEMESTER,
    course_slug: str = _COURSE,
) -> dict:
    """Return a minimal ItemStatistics-compatible dict."""
    return {
        "item_no": item_no,
        "semester": semester,
        "course_slug": course_slug,
        "chapter": chapter,
        "week": None,
        "item_type": "지식축적",
        "difficulty_level": 2,
        "expected_difficulty": "보통",
        "source": "형성평가",
        "correct_answer": 3,
        "n_responders": 20,
        "n_correct": 10,
        "n_omit": 0,
        "correct_rate": 0.50,
        "omit_rate": 0.00,
        "discrimination_index": 0.25,
        "point_biserial": 0.35,
        "top_distractor_no": 2,
        "top_distractor_rate": 0.20,
        "is_top_distractor_adjacent": True,
        "option_distribution": json.dumps({1: 0.1, 2: 0.2, 3: 0.5, 4: 0.1, 5: 0.1}),
        "distractor_label": "특이사항 없음",
    }


def _blueprint(semester: str = _SEMESTER, course_slug: str = _COURSE) -> dict:
    return {
        "semester": semester,
        "course_slug": course_slug,
        "exam_name": "2026-1학기 기말고사",
        "total_items": 40,
        "chapters": [_CHAPTER_A, _CHAPTER_B],
        "difficulty_targets": {"easy": 0.45, "medium": 0.35, "hard": 0.20},
        "source_mix": {"formative": 18, "quiz": 12, "textbook": 10},
        "quiz_target": 12,
        "answer_key_balance": True,
    }


def _curriculum(semester: str = _SEMESTER, course_slug: str = _COURSE) -> dict:
    return {
        "semester": semester,
        "course_slug": course_slug,
        "entries": [
            {
                "week": 1,
                "chapter": _CHAPTER_A,
                "chapter_no": 1,
                "subtopic": None,
                "sections": ["1.1 인체의 조직"],
            },
            {
                "week": 2,
                "chapter": _CHAPTER_B,
                "chapter_no": 2,
                "subtopic": None,
                "sections": ["2.1 세포의 구조"],
            },
        ],
    }


def _retro_config(semester: str = _SEMESTER, course_slug: str = _COURSE) -> dict:
    return {
        "semester": semester,
        "course_slug": course_slug,
        "group_roster": {
            "2026000001": "학령기",
            "2026000002": "학령기",
            "2026000003": "만학도",
            "2026000004": "만학도",
        },
        "unit_importance": {
            _CHAPTER_A: "상",
            _CHAPTER_B: "중",
        },
        "gap_threshold": 0.6,
        "baseline_segment": "만학도",
        "low_discrimination_threshold": 0.2,
        "cognitive_cliff_drop": 0.15,
        "effort_ratings": {
            _CHAPTER_A: "상",
            _CHAPTER_B: "중",
        },
    }


def _build_fixture_tree(data_root: Path) -> None:
    """Write the complete fixture file tree under ``data_root``."""
    key = _KEY

    # Silver immersio
    silver_dir = data_root / "silver" / "immersio" / key
    silver_dir.mkdir(parents=True, exist_ok=True)

    # Four students: all below gap_threshold (0.6) so we get gaps to rank
    combined_rows = [
        _combined_row("2026000001", {_CHAPTER_A: 0.4, _CHAPTER_B: 0.3}),
        _combined_row("2026000002", {_CHAPTER_A: 0.5, _CHAPTER_B: 0.35}),
        _combined_row("2026000003", {_CHAPTER_A: 0.45, _CHAPTER_B: 0.25}),
        _combined_row("2026000004", {_CHAPTER_A: 0.4, _CHAPTER_B: 0.2}),
    ]
    pd.DataFrame(combined_rows).to_parquet(silver_dir / "진단×시험결합.parquet", index=False)

    item_rows = [
        _item_row(1, _CHAPTER_A),
        _item_row(2, _CHAPTER_B),
    ]
    pd.DataFrame(item_rows).to_parquet(silver_dir / "문항통계.parquet", index=False)

    # Bronze retro-mester
    bronze_dir = data_root / "bronze" / "retro-mester" / key
    bronze_dir.mkdir(parents=True, exist_ok=True)

    (bronze_dir / "retro_config.yaml").write_text(
        yaml.dump(_retro_config(), allow_unicode=True), encoding="utf-8"
    )
    (bronze_dir / "blueprint.yaml").write_text(
        yaml.dump(_blueprint(), allow_unicode=True), encoding="utf-8"
    )
    (bronze_dir / "curriculum_map.yaml").write_text(
        yaml.dump(_curriculum(), allow_unicode=True), encoding="utf-8"
    )


# ---------------------------------------------------------------------------
# Happy-path E2E test
# ---------------------------------------------------------------------------


class TestUS1E2E:
    """T030: full pipeline wiring, US1 (no LLM)."""

    def test_exit_zero_and_all_outputs_present(self, tmp_path: Path) -> None:
        """run_retro exits 0 and produces all expected Gold+Silver artefacts."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)

        code = run_retro(
            semester=_SEMESTER,
            course=_COURSE,
            data_root=str(data_root),
            llm_mode="off",
        )

        assert code == 0, f"Expected exit 0, got {code}"

        # Gold outputs
        gold = data_root / "gold" / "retro-mester" / _KEY
        assert (gold / "CQI회고보고서.md").exists(), "Missing CQI회고보고서.md"
        assert (gold / "CQI회고보고서.pdf").exists(), "Missing CQI회고보고서.pdf"
        assert (gold / "회고분석.xlsx").exists(), "Missing 회고분석.xlsx"

        # Silver outputs (FR-012: manifest is a Silver-layer artefact)
        silver = data_root / "silver" / "retro-mester" / _KEY
        assert (silver / "빈틈표.parquet").exists(), "Missing 빈틈표.parquet"
        assert (silver / "변경권고.parquet").exists(), "Missing 변경권고.parquet"
        assert (silver / "근거부족단원.parquet").exists(), "Missing 근거부족단원.parquet"
        assert (silver / "manifest_retro.json").exists(), "Missing manifest_retro.json"
        assert not (gold / "manifest_retro.json").exists(), "manifest must not be in Gold"

    def test_md_contains_ranked_changes_and_uncovered_ratio(self, tmp_path: Path) -> None:
        """The generated MD report has a ranked changes table and uncovered ratio."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)
        run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))

        md_path = data_root / "gold" / "retro-mester" / _KEY / "CQI회고보고서.md"
        md_text = md_path.read_text(encoding="utf-8")

        # The report must contain the ranked-changes table header
        assert "변경 권고 요약" in md_text, "Missing section header"
        # And the uncovered-ratio summary line
        assert "못 덮은 빈틈 비율" in md_text, "Missing uncovered ratio line"

    def test_no_student_id_in_md(self, tmp_path: Path) -> None:
        """No student ID appears in the MD report (privacy guard)."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)
        run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))

        md_path = data_root / "gold" / "retro-mester" / _KEY / "CQI회고보고서.md"
        md_text = md_path.read_text(encoding="utf-8")

        # Student IDs embedded in fixture
        for sid in ["2026000001", "2026000002", "2026000003", "2026000004"]:
            assert sid not in md_text, f"Student ID {sid} must NOT appear in MD"

    def test_manifest_structure(self, tmp_path: Path) -> None:
        """manifest_retro.json has required top-level keys and valid JSON."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)
        run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))

        manifest_path = data_root / "silver" / "retro-mester" / _KEY / "manifest_retro.json"
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

        required_keys = {
            "module_version",
            "schema_version",
            "semester",
            "course_slug",
            "inputs",
            "thresholds",
            "counts",
            "degrade",
            "generated_at_utc",
        }
        for k in required_keys:
            assert k in manifest, f"Missing manifest key: {k}"

        # US1 — llm_used must be False
        assert manifest["degrade"]["llm_used"] is False

    def test_missing_combined_parquet_exits_2(self, tmp_path: Path) -> None:
        """Missing 진단×시험결합.parquet → InputError → exit code 2."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)

        # Remove the combined parquet to trigger InputError
        combined = data_root / "silver" / "immersio" / _KEY / "진단×시험결합.parquet"
        combined.unlink()

        code = run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))

        assert code == 2, f"Expected exit 2 for missing input, got {code}"

    def test_missing_config_exits_2(self, tmp_path: Path) -> None:
        """Missing retro_config.yaml → InputError → exit code 2."""
        from retro_mester.pipeline import run_retro

        data_root = tmp_path / "data"
        _build_fixture_tree(data_root)

        config_path = data_root / "bronze" / "retro-mester" / _KEY / "retro_config.yaml"
        config_path.unlink()

        code = run_retro(semester=_SEMESTER, course=_COURSE, data_root=str(data_root))

        assert code == 2, f"Expected exit 2 for missing config, got {code}"


# ---------------------------------------------------------------------------
# T013 — US1 no-silent-omission end-to-end (all six quickstart checks)
# ---------------------------------------------------------------------------

_CHAPTER_C = "9장. 신경계통"  # in blueprint/curriculum/items but ZERO cohort data
_STUDENT_IDS = ["2026000001", "2026000002", "2026000003", "2026000004"]


def _build_omission_fixture(data_root: Path, prior_yaml: Path) -> None:
    """Build a fixture exercising all six US1 no-silent-omission checks.

    - ``_CHAPTER_C`` is present in blueprint/curriculum/items but absent from
      every student's ``chapter_correct_rates`` → zero cohort evidence
      (insufficient unit) AND an items↔combined chapter-set mismatch.
    - A prior-year yaml is supplied for provenance.

    Args:
        data_root: Test data root.
        prior_yaml: Path to an existing prior 차년도방향.yaml.
    """
    key = _KEY
    silver_dir = data_root / "silver" / "immersio" / key
    silver_dir.mkdir(parents=True, exist_ok=True)

    # Students cover A and B (below threshold → gaps); none cover C.
    combined_rows = [
        _combined_row("2026000001", {_CHAPTER_A: 0.4, _CHAPTER_B: 0.3}),
        _combined_row("2026000002", {_CHAPTER_A: 0.5, _CHAPTER_B: 0.35}),
        _combined_row("2026000003", {_CHAPTER_A: 0.45, _CHAPTER_B: 0.25}),
        _combined_row("2026000004", {_CHAPTER_A: 0.4, _CHAPTER_B: 0.2}),
    ]
    pd.DataFrame(combined_rows).to_parquet(silver_dir / "진단×시험결합.parquet", index=False)

    # Items include C → items_not_in_combined mismatch + insufficient chapter.
    item_rows = [
        _item_row(1, _CHAPTER_A),
        _item_row(2, _CHAPTER_B),
        _item_row(3, _CHAPTER_C),
    ]
    pd.DataFrame(item_rows).to_parquet(silver_dir / "문항통계.parquet", index=False)

    bronze_dir = data_root / "bronze" / "retro-mester" / key
    bronze_dir.mkdir(parents=True, exist_ok=True)

    cfg = _retro_config()
    cfg["unit_importance"][_CHAPTER_C] = "중"
    cfg["effort_ratings"][_CHAPTER_C] = "중"
    bp = _blueprint()
    bp["chapters"] = [_CHAPTER_A, _CHAPTER_B, _CHAPTER_C]
    cur = _curriculum()
    cur["entries"].append(
        {
            "week": 3,
            "chapter": _CHAPTER_C,
            "chapter_no": 3,
            "subtopic": None,
            "sections": ["3.1 신경계 개요"],
        }
    )

    (bronze_dir / "retro_config.yaml").write_text(
        yaml.dump(cfg, allow_unicode=True), encoding="utf-8"
    )
    (bronze_dir / "blueprint.yaml").write_text(yaml.dump(bp, allow_unicode=True), encoding="utf-8")
    (bronze_dir / "curriculum_map.yaml").write_text(
        yaml.dump(cur, allow_unicode=True), encoding="utf-8"
    )

    # prior-year yaml already written by caller at `prior_yaml`.
    _ = prior_yaml


class TestUS1NoSilentOmissionE2E:
    """T013: every US1 quickstart check holds end-to-end (no silent omission)."""

    def _run(self, tmp_path: Path) -> tuple[Path, dict]:
        """Run the omission fixture; return (data_root, parsed manifest)."""
        from retro_mester.pipeline import run_retro

        from tests.fixtures.factories import write_prior_forward_yaml

        data_root = tmp_path / "data"
        prior_yaml = write_prior_forward_yaml(tmp_path / "prior")
        _build_omission_fixture(data_root, prior_yaml)

        code = run_retro(
            semester=_SEMESTER,
            course=_COURSE,
            data_root=str(data_root),
            llm_mode="off",
            prior_yaml_path=str(prior_yaml),
        )
        assert code == 0, f"Pipeline failed: exit {code}"

        manifest = json.loads(
            (data_root / "silver" / "retro-mester" / _KEY / "manifest_retro.json").read_text(
                encoding="utf-8"
            )
        )
        return data_root, manifest

    def test_check1_insufficient_parquet_has_unit(self, tmp_path: Path) -> None:
        """(1) 근거부족단원.parquet carries the unit with the right reason."""
        data_root, _ = self._run(tmp_path)
        silver = data_root / "silver" / "retro-mester" / _KEY
        df = pd.read_parquet(silver / "근거부족단원.parquet")
        assert len(df) >= 1
        c_rows = df[df["chapter"] == _CHAPTER_C]
        assert len(c_rows) >= 1, f"{_CHAPTER_C} missing from insufficient parquet"
        assert set(c_rows["reason"]) == {"근거부족-자료없음"}

    def test_check2_md_and_xlsx_show_insufficient(self, tmp_path: Path) -> None:
        """(2) report md and xlsx 빈틈 sheet show '근거 부족'."""
        from openpyxl import load_workbook

        data_root, _ = self._run(tmp_path)
        gold = data_root / "gold" / "retro-mester" / _KEY

        md_text = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")
        assert "근거 부족" in md_text
        assert _CHAPTER_C in md_text

        wb = load_workbook(gold / "회고분석.xlsx", read_only=True)
        ws = wb["빈틈"]
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert any("근거 부족" in v for v in values)
        assert any(_CHAPTER_C in v for v in values)

    def test_check3_uncovered_ratio_includes_insufficient(self, tmp_path: Path) -> None:
        """(3) uncovered_ratio denominator includes insufficient units."""
        data_root, manifest = self._run(tmp_path)
        counts = manifest["counts"]
        gaps_n = counts["gaps"]
        insuf_n = counts["insufficient_evidence_units"]
        covered_n = counts["covered"]
        ratio = counts["uncovered_ratio"]

        assert insuf_n >= 1, "expected at least one insufficient unit"
        # Honest denominator: total = gaps + insufficient.
        expected = ((gaps_n - covered_n) + insuf_n) / (gaps_n + insuf_n)
        assert abs(ratio - expected) < 1e-9
        # Coverage is strictly lower than if insufficient were excluded.
        ratio_excluding = (gaps_n - covered_n) / gaps_n if gaps_n else 0.0
        assert ratio > ratio_excluding

    def test_check4_warnings_has_chapter_mismatch(self, tmp_path: Path) -> None:
        """(4) manifest.warnings carries the chapter-mismatch warning."""
        _, manifest = self._run(tmp_path)
        warnings = manifest["warnings"]
        assert isinstance(warnings, list) and warnings
        assert any(_CHAPTER_C in w for w in warnings)

    def test_check5_prior_year_provenance(self, tmp_path: Path) -> None:
        """(5) manifest.inputs.prior_year has {path, sha256}."""
        _, manifest = self._run(tmp_path)
        assert "prior_year" in manifest["inputs"]
        prov = manifest["inputs"]["prior_year"]
        assert "path" in prov and prov["path"]
        assert isinstance(prov["sha256"], str) and len(prov["sha256"]) == 64

    def test_check6_manifest_in_silver(self, tmp_path: Path) -> None:
        """(6) manifest lives in the Silver dir, not Gold."""
        data_root, _ = self._run(tmp_path)
        silver_m = data_root / "silver" / "retro-mester" / _KEY / "manifest_retro.json"
        gold_m = data_root / "gold" / "retro-mester" / _KEY / "manifest_retro.json"
        assert silver_m.exists()
        assert not gold_m.exists()

    def test_no_individual_student_rows_in_outputs(self, tmp_path: Path) -> None:
        """Constitution IV: no individual-student row appears in any output."""
        from openpyxl import load_workbook

        data_root, manifest = self._run(tmp_path)
        gold = data_root / "gold" / "retro-mester" / _KEY
        silver = data_root / "silver" / "retro-mester" / _KEY

        md_text = (gold / "CQI회고보고서.md").read_text(encoding="utf-8")
        manifest_text = (silver / "manifest_retro.json").read_text(encoding="utf-8")

        for sid in _STUDENT_IDS:
            assert sid not in md_text, f"Student ID {sid} leaked into MD"
            assert sid not in manifest_text, f"Student ID {sid} leaked into manifest"

        # Parquet group-level tables must not carry a student_id column.
        for name in ("빈틈표.parquet", "변경권고.parquet", "근거부족단원.parquet"):
            df = pd.read_parquet(silver / name)
            assert "student_id" not in df.columns, f"{name} carries student_id"

        # xlsx cells must not contain any student ID.
        wb = load_workbook(gold / "회고분석.xlsx", read_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    val = str(cell.value)
                    for sid in _STUDENT_IDS:
                        assert sid not in val, f"Student ID {sid} in xlsx sheet {ws.title}"
