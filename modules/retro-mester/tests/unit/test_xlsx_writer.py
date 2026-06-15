"""T028 — XLSX writer tests (RED phase).

Tests for ``retro_mester.output.xlsx.write_xlsx``.
- Both sheets present.
- Expected column headers.
- Two writes with same ``when`` are byte-identical.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from paideia_shared.schemas.change_recommendation import ChangeRecommendation
from paideia_shared.schemas.unit_gap import UnitGap

from retro_mester.output.xlsx import write_xlsx

_WHEN = datetime.datetime(2026, 6, 15, 9, 0, 0)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_gaps() -> list[UnitGap]:
    common = dict(
        semester="2026-1",
        course_slug="anatomy",
        segment_mean_rate=0.60,
        n_below=5,
        pct_segment=0.25,
        pct_cohort=0.10,
        is_structural=True,
        cohort_failing_item_types=["지식"],
        cause="내용난이도",
        cause_signals={"diff": -0.1},
        validity="건전",
        unit_importance="상",
        weight=3.0,
        evidence_n=20,
        impact_score=15.0,
    )
    return [
        UnitGap(**common, chapter="1장 세포", segment="학령기"),
        UnitGap(**common, chapter="2장 조직", segment="만학도"),
    ]


def _make_recs() -> list[ChangeRecommendation]:
    base = dict(
        semester="2026-1",
        course_slug="anatomy",
        target_cognitive_level="이해",
        cause_hypothesis="내용난이도",
        covered_n=5,
        covered_pct_segment=0.25,
        covered_pct_cohort=0.10,
        unit_importance="상",
        weight=3.0,
        effort_level="중",
        priority_quadrant="빠른승리",
        prescription_key="scaffold_concepts",
        cluster_vocab=None,
        validity="건전",
        impact_score=15.0,
    )
    return [
        ChangeRecommendation(
            **base,
            rank=1,
            chapter="1장 세포",
            segment="학령기",
            is_covered=True,
        ),
        ChangeRecommendation(
            **base,
            rank=None,
            chapter="2장 조직",
            segment="만학도",
            is_covered=False,
        ),
    ]


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestWriteXlsx:
    def test_file_created(self, tmp_path: Path) -> None:
        """write_xlsx creates the xlsx file at the given path."""
        out = tmp_path / "retro.xlsx"
        write_xlsx(_make_gaps(), _make_recs(), out, _WHEN)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_both_sheets_present(self, tmp_path: Path) -> None:
        """Workbook contains sheets '빈틈' and '변경권고'."""
        out = tmp_path / "retro.xlsx"
        write_xlsx(_make_gaps(), _make_recs(), out, _WHEN)
        wb = load_workbook(out, read_only=True)
        sheet_names = wb.sheetnames
        assert "빈틈" in sheet_names
        assert "변경권고" in sheet_names

    def test_gap_sheet_has_chapter_column(self, tmp_path: Path) -> None:
        """'빈틈' sheet header row contains 'chapter'."""
        out = tmp_path / "retro.xlsx"
        write_xlsx(_make_gaps(), _make_recs(), out, _WHEN)
        wb = load_workbook(out, read_only=True)
        ws = wb["빈틈"]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        assert "chapter" in headers

    def test_rec_sheet_has_rank_column(self, tmp_path: Path) -> None:
        """'변경권고' sheet header row contains 'rank'."""
        out = tmp_path / "retro.xlsx"
        write_xlsx(_make_gaps(), _make_recs(), out, _WHEN)
        wb = load_workbook(out, read_only=True)
        ws = wb["변경권고"]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        assert "rank" in headers

    def test_byte_identical_on_two_writes(self, tmp_path: Path) -> None:
        """Two writes with the same ``when`` produce byte-identical xlsx files."""
        out1 = tmp_path / "retro1.xlsx"
        out2 = tmp_path / "retro2.xlsx"
        gaps = _make_gaps()
        recs = _make_recs()
        write_xlsx(gaps, recs, out1, _WHEN)
        write_xlsx(gaps, recs, out2, _WHEN)
        assert out1.read_bytes() == out2.read_bytes()

    def test_numbers_stored_as_numbers(self, tmp_path: Path) -> None:
        """Numeric columns (impact_score, weight) are stored as numbers not strings."""
        out = tmp_path / "retro.xlsx"
        write_xlsx(_make_gaps(), _make_recs(), out, _WHEN)
        wb = load_workbook(out, read_only=True)
        ws = wb["빈틈"]
        # Find impact_score column
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        if "impact_score" in headers:
            col_idx = headers.index("impact_score") + 1
            # Read first data row
            for row in ws.iter_rows(min_row=2, max_row=2):
                val = row[col_idx - 1].value
                assert isinstance(val, (int, float)), (
                    f"impact_score should be numeric, got {type(val)}"
                )
