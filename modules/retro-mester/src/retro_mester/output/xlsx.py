"""T028 / T035 / T047 / T051 — Gold-layer xlsx writer for retro-mester.

Entry point: ``write_xlsx(gaps, recs, xlsx_path, when)``.

Sheets:
- ``빈틈``       — one row per UnitGap.
- ``변경권고``   — one row per ChangeRecommendation.
- ``집단대비``   — per chapter × segment comparison (US2 T035).
- ``정렬``       — one row per AlignmentFinding (US4 T047).
- ``타당도``     — per-chapter validity verdict + psychometric signals (US5 T051).

Determinism:
- Row order: gaps sorted by (chapter, segment); recs by (rank nulls-last,
  chapter, segment); 집단대비 sorted by (chapter, segment); 정렬 sorted by
  (chapter); 타당도 sorted by (chapter).
- Workbook creator / lastModifiedBy pinned to ``_PRODUCER``.
- ``finalize_xlsx(xlsx_path, when)`` rewrites ``<dcterms:modified>`` and
  ``<dcterms:created>`` post-save for byte-identical output.

Note: ``변경권고`` rows where ``validity == "문항수선"`` carry the repair
prescription (not a re-teaching prescription) — annotated via the
``validity`` and ``prescription_key`` columns.
"""

from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from paideia_shared.schemas import InsufficientEvidenceUnit
from paideia_shared.schemas.alignment_finding import AlignmentFinding
from paideia_shared.schemas.change_recommendation import ChangeRecommendation
from paideia_shared.schemas.unit_gap import UnitGap

from retro_mester.output.determinism import finalize_xlsx

# Columns for the US2 집단대비 (group comparison) sheet.
_CONTRAST_COLUMNS: list[str] = [
    "chapter",
    "segment",
    "segment_mean_rate",
    "n_below",
    "is_structural",
    "cause",
    "prescription",
]

_ALIGN_COLUMNS: list[str] = [
    "semester",
    "course_slug",
    "chapter",
    "taught_weeks",
    "tested_items",
    "learned_rate",
    "flag",
    "interest_gap",
    "aversion_gap",
    "note",
]

_PRODUCER = "paideia/retro-mester/0.1.0"

# Column orders are derived from the model field order for predictability.
_GAP_COLUMNS: list[str] = [
    "semester",
    "course_slug",
    "chapter",
    "segment",
    "segment_mean_rate",
    "n_below",
    "pct_segment",
    "pct_cohort",
    "is_structural",
    "cohort_failing_item_types",
    "cause",
    "cause_signals",
    "validity",
    "unit_importance",
    "weight",
    "impact_score",
    "evidence_n",
]

_REC_COLUMNS: list[str] = [
    "semester",
    "course_slug",
    "rank",
    "chapter",
    "target_cognitive_level",
    "segment",
    "cause_hypothesis",
    "covered_n",
    "covered_pct_segment",
    "covered_pct_cohort",
    "unit_importance",
    "weight",
    "impact_score",
    "effort_level",
    "priority_quadrant",
    "prescription_key",
    "cluster_vocab",
    "validity",
    "is_covered",
]


_INSUFFICIENT_BLOCK_LABEL = "근거 부족 단원 (코호트 전체 응답 데이터 없음)"
_INSUFFICIENT_BLOCK_COLUMNS: list[str] = ["chapter", "segment", "evidence_n", "reason"]


def _build_gap_sheet(
    wb: Workbook,
    gaps: list[UnitGap],
    insufficient: list[InsufficientEvidenceUnit],
) -> None:
    """Build the 빈틈 sheet with an adjacent 근거 부족 단원 block (H1).

    The gap rows are written first; below them (separated by a blank row) a
    labeled block lists the insufficient (zero-cohort-evidence) units so they
    are visibly "근거 부족", never silently absent.  The block is omitted only
    its header when no insufficient units exist (the label still renders).

    Args:
        wb: Target workbook.
        gaps: All UnitGap records.
        insufficient: InsufficientEvidenceUnit records to surface below the gaps.
    """
    ws = wb.create_sheet("빈틈")
    bold = Font(bold=True)

    # Header row
    for c, col in enumerate(_GAP_COLUMNS, start=1):
        ws.cell(1, c, col).font = bold

    # Sort gaps: chapter ASC, segment ASC
    sorted_gaps = sorted(gaps, key=lambda g: (g.chapter, g.segment))
    last_row = 1
    for r, gap in enumerate(sorted_gaps, start=2):
        row_dict = gap.model_dump()
        for c, col in enumerate(_GAP_COLUMNS, start=1):
            value = row_dict[col]
            # List/dict columns: convert to str for readability in xlsx
            if isinstance(value, (list, dict)):
                value = str(value)
            ws.cell(r, c, value)
        last_row = r

    # 근거 부족 단원 block — blank separator row, label, header, then units.
    label_row = last_row + 2
    ws.cell(label_row, 1, _INSUFFICIENT_BLOCK_LABEL).font = bold
    header_row = label_row + 1
    for c, col in enumerate(_INSUFFICIENT_BLOCK_COLUMNS, start=1):
        ws.cell(header_row, c, col).font = bold

    sorted_insuf = sorted(insufficient, key=lambda u: (u.chapter, u.segment))
    for r, unit in enumerate(sorted_insuf, start=header_row + 1):
        row_dict = unit.model_dump()
        for c, col in enumerate(_INSUFFICIENT_BLOCK_COLUMNS, start=1):
            ws.cell(r, c, row_dict[col])


def _build_rec_sheet(wb: Workbook, recs: list[ChangeRecommendation]) -> None:
    ws = wb.create_sheet("변경권고")
    bold = Font(bold=True)

    # Header row
    for c, col in enumerate(_REC_COLUMNS, start=1):
        ws.cell(1, c, col).font = bold

    # Sort recs: rank (None last) ASC, chapter ASC, segment ASC
    sorted_recs = sorted(
        recs,
        key=lambda r: (r.rank if r.rank is not None else 999, r.chapter, r.segment),
    )
    for row_idx, rec in enumerate(sorted_recs, start=2):
        row_dict = rec.model_dump()
        for c, col in enumerate(_REC_COLUMNS, start=1):
            ws.cell(row_idx, c, row_dict[col])


def _build_contrast_sheet(
    wb: Workbook,
    gaps: list[UnitGap],
    prescriptions: dict[tuple[str, str], str],
) -> None:
    """Build the 집단대비 (group comparison) sheet (US2 T035).

    One row per (chapter, segment) gap showing the key metrics and the
    group-specific prescription.  No student IDs are written.

    Args:
        wb: Target workbook.
        gaps: All UnitGap records (already escalated by escalate_structural).
        prescriptions: Mapping (chapter, segment) → prescription string;
            built in the pipeline before calling write_xlsx.
    """
    ws = wb.create_sheet("집단대비")
    bold = Font(bold=True)

    for c, col in enumerate(_CONTRAST_COLUMNS, start=1):
        ws.cell(1, c, col).font = bold

    sorted_gaps = sorted(gaps, key=lambda g: (g.chapter, g.segment))
    for r, gap in enumerate(sorted_gaps, start=2):
        presc = prescriptions.get((gap.chapter, gap.segment), "")
        row_values = [
            gap.chapter,
            gap.segment,
            gap.segment_mean_rate,
            gap.n_below,
            gap.is_structural,
            gap.cause,
            presc,
        ]
        for c, val in enumerate(row_values, start=1):
            ws.cell(r, c, val)


_VALIDITY_COLUMNS: list[str] = [
    "chapter",
    "verdict",
    "mean_discrimination",
    "low_disc_share",
    "bad_distractor_share",
]


def _build_validity_sheet(
    wb: Workbook,
    validity_table: list[dict],
) -> None:
    """Build the 타당도 (assessment-validity) sheet (US5 T051).

    One row per chapter showing the ValidityVerdict and the three numeric
    psychometric signals used to derive it.  Rows are sorted by chapter
    ascending.

    Note: rows where ``verdict == "문항수선"`` signal that the chapter's
    items need revision before instructional prescriptions are meaningful;
    the corresponding ``변경권고`` rows carry the repair prescription
    instead of a re-teaching prescription (SC-006).

    Args:
        wb: Target workbook.
        validity_table: List of dicts with keys matching ``_VALIDITY_COLUMNS``,
            one per chapter.  When empty, only the header row is written.
    """
    ws = wb.create_sheet("타당도")
    bold = Font(bold=True)

    for c, col in enumerate(_VALIDITY_COLUMNS, start=1):
        ws.cell(1, c, col).font = bold

    for r, row in enumerate(sorted(validity_table, key=lambda d: d["chapter"]), start=2):
        for c, col in enumerate(_VALIDITY_COLUMNS, start=1):
            ws.cell(r, c, row.get(col))


def _build_align_sheet(
    wb: Workbook,
    findings: list[AlignmentFinding],
) -> None:
    """Build the 정렬 (alignment) sheet (US4 T047).

    One row per AlignmentFinding, sorted by chapter ascending.
    ``cognitive_profile`` (a dict) is converted to a string for readability.

    Args:
        wb: Target workbook.
        findings: AlignmentFinding records from build_alignment.
    """
    ws = wb.create_sheet("정렬")
    bold = Font(bold=True)

    for c, col in enumerate(_ALIGN_COLUMNS, start=1):
        ws.cell(1, c, col).font = bold

    sorted_findings = sorted(findings, key=lambda f: f.chapter)
    for r, finding in enumerate(sorted_findings, start=2):
        row_dict = finding.model_dump()
        for c, col in enumerate(_ALIGN_COLUMNS, start=1):
            value = row_dict[col]
            if isinstance(value, dict):
                value = str(value)
            ws.cell(r, c, value)


def write_xlsx(
    gaps: list[UnitGap],
    recs: list[ChangeRecommendation],
    xlsx_path: Path,
    when: datetime.datetime,
    prescriptions: dict[tuple[str, str], str] | None = None,
    alignment_findings: list[AlignmentFinding] | None = None,
    validity_table: list[dict] | None = None,
    insufficient: list[InsufficientEvidenceUnit] | None = None,
) -> None:
    """Write ``빈틈``, ``변경권고``, ``집단대비``, ``정렬``, ``타당도`` sheets.

    Never calls ``datetime.now()`` internally.  ``finalize_xlsx`` is
    called after ``save()`` to pin ``<dcterms:modified>`` and
    ``<dcterms:created>`` so two runs with the same ``when`` produce
    byte-identical files.

    Args:
        gaps: List of UnitGap records.
        recs: List of ChangeRecommendation records.
        xlsx_path: Destination ``.xlsx`` path. Parent directory must exist.
        when: Timestamp for workbook metadata and determinism pin.
        prescriptions: Optional mapping ``(chapter, segment) → prescription``
            for the 집단대비 sheet (US2 T035).  When ``None``, the sheet is
            still written but prescription cells are empty.
        alignment_findings: Optional list of AlignmentFinding for the 정렬
            sheet (US4 T047).  When ``None``, the sheet is still written
            but contains only the header row.
        validity_table: Optional list of per-chapter validity dicts for the
            타당도 sheet (US5 T051).  Each dict must have keys:
            ``chapter``, ``verdict``, ``mean_discrimination``,
            ``low_disc_share``, ``bad_distractor_share``.  When ``None``,
            the sheet is written with only the header row.
        insufficient: Optional list of ``InsufficientEvidenceUnit`` surfaced as a
            labeled 근거 부족 단원 block inside the 빈틈 sheet (H1).  When ``None``,
            an empty block is rendered (label only).

    Raises:
        FileNotFoundError: When ``xlsx_path.parent`` does not exist.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.parent.is_dir():
        raise FileNotFoundError(f"write_xlsx: parent directory missing: {xlsx_path.parent}")

    wb = Workbook()
    # Drop the default active sheet so tab order matches spec exactly.
    default = wb.active
    if default is not None:
        wb.remove(default)

    _build_gap_sheet(wb, gaps, insufficient or [])
    _build_rec_sheet(wb, recs)
    _build_contrast_sheet(wb, gaps, prescriptions or {})
    _build_align_sheet(wb, alignment_findings or [])
    _build_validity_sheet(wb, validity_table or [])

    # Pin workbook-level metadata
    wb.properties.creator = _PRODUCER
    wb.properties.lastModifiedBy = _PRODUCER
    wb.properties.created = when
    wb.properties.modified = when

    wb.save(xlsx_path)

    # Rewrite <dcterms:modified> + <dcterms:created> after save() stamps
    # them with datetime.now() — ensures byte-identical output.
    finalize_xlsx(xlsx_path, when)


__all__ = ["write_xlsx", "_build_align_sheet", "_build_validity_sheet"]
