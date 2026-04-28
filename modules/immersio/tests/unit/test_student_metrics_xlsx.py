"""T047 — RED tests for xlsx_writer's 7th sheet `학생성적` (FR-013, FR-018).

Expanded ``write_analysis_xlsx`` accepts an optional
``student_metrics`` argument; when supplied, it appends the 학생성적
sheet with:

* 9 fixed columns (학번 / 이름 / 분반 / 응시여부 / 총점 / 100점환산 /
  분반_백분위 / 전체_백분위 / z_score)
* dynamic columns spread from ExamItem metadata (chapter / source /
  difficulty_level / expected_difficulty / item_type)
* 2 needs-map join columns (관심챕터_본인정답률 / 비호감챕터_본인정답률)

Absent students (exam_taken=False) live as rows with score columns left
as None so xlsx readers see empty cells.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from paideia_shared.schemas import (
    HistogramBin,
    ItemStatistics,
    MetadataAggregate,
    StudentExamMetrics,
)

from immersio.report.xlsx_writer import write_analysis_xlsx


def _stub_overall() -> list[dict[str, object]]:
    return [
        {"지표": "응시자 수", "값": 2},
        {"지표": "결시자 수", "값": 1},
        {"지표": "무응답 응답 수", "값": 0},
        {"지표": "만점", "값": 4.0},
        {"지표": "평균", "값": 3.0},
        {"지표": "표준편차", "값": 1.0},
        {"지표": "중앙값", "값": 3.0},
        {"지표": "최저", "값": 2.0},
        {"지표": "최고", "값": 4.0},
        {"지표": "Q1", "값": 2.5},
        {"지표": "Q3", "값": 3.5},
        {"지표": "100점환산_평균", "값": 75.0},
        {"지표": "100점환산_표준편차", "값": 25.0},
    ]


def _stub_histogram() -> list[HistogramBin]:
    return [
        HistogramBin(bin_start=0.0, bin_end=10.0, count=2, cumulative=2, cumulative_pct=100.0),
    ]


def _stub_metadata() -> list[MetadataAggregate]:
    return [
        MetadataAggregate(
            metadata_kind="분반",
            metadata_value="A",
            n=2,
            mean=3.0,
            sd=1.0,
            test_kind="N/A",
            test_p_value=None,
            levene_p_value=None,
            note=None,
        ),
    ]


def _stub_item(no: int) -> ItemStatistics:
    return ItemStatistics(
        item_no=no,
        semester="2026-1",
        course_slug="anatomy",
        chapter="1장. 서론",
        week=1,
        item_type="지식축적",
        difficulty_level=2,
        expected_difficulty="보통",
        source="교과서",
        correct_answer=1,
        n_responders=2,
        n_correct=2,
        n_omit=0,
        correct_rate=1.0,
        omit_rate=0.0,
        discrimination_index=0.0,
        point_biserial=0.0,
        top_distractor_no=None,
        top_distractor_rate=None,
        is_top_distractor_adjacent=False,
        option_distribution={1: 1.0, 2: 0.0, 3: 0.0, 4: 0.0, 5: 0.0},
        distractor_label="특이사항 없음",
    )


def _stub_metrics() -> list[StudentExamMetrics]:
    return [
        StudentExamMetrics(
            student_id="2026100001",
            name_kr="김성적",
            section="A",
            semester="2026-1",
            course_slug="anatomy",
            exam_taken=True,
            total_score=4.0,
            score_percent=100.0,
            section_percentile=75.0,
            cohort_percentile=75.0,
            z_score=1.0,
            chapter_correct_rates={"1장. 서론": 1.0, "2장. 세포와 조직": 0.5},
            source_correct_rates={"교과서": 1.0, "형성평가": 0.5},
            difficulty_correct_rates={2: 1.0, 3: 0.5},
            expected_difficulty_correct_rates={"보통": 1.0, "어려움": 0.5},
            item_type_correct_rates={"지식축적": 0.75},
            interest_chapters_correct_rate=1.0,
            aversion_chapters_correct_rate=None,
        ),
        StudentExamMetrics(
            student_id="2026100002",
            name_kr="이결시",
            section="B",
            semester="2026-1",
            course_slug="anatomy",
            exam_taken=False,
        ),
    ]


@pytest.fixture
def written_xlsx(tmp_path: Path) -> Path:
    out = tmp_path / "result.xlsx"
    write_analysis_xlsx(
        output_path=out,
        overall_rows=_stub_overall(),
        histogram_bins=_stub_histogram(),
        metadata_rows=_stub_metadata(),
        item_stats=[_stub_item(1)],
        student_metrics=_stub_metrics(),
        semester="2026-1",
        course_name_kr="인체구조와기능",
        generated_at_utc="2026-04-29T00:00:00Z",
    )
    return out


def test_xlsx_has_seven_sheets_with_학생성적_last(written_xlsx: Path) -> None:
    wb = load_workbook(written_xlsx)
    assert wb.sheetnames == [
        "전체요약",
        "1_히스토그램",
        "2_메타데이터통계",
        "3_변별력",
        "4_정답률",
        "5_오답분석",
        "학생성적",
    ]


def test_학생성적_fixed_columns_present(written_xlsx: Path) -> None:
    wb = load_workbook(written_xlsx)
    ws = wb["학생성적"]
    headers = [ws.cell(1, c).value for c in range(1, 10)]
    assert headers[:9] == [
        "학번", "이름", "분반", "응시여부", "총점",
        "100점환산", "분반_백분위", "전체_백분위", "z_score",
    ]


def test_학생성적_dynamic_columns_spread(written_xlsx: Path) -> None:
    wb = load_workbook(written_xlsx)
    ws = wb["학생성적"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # ExamItem metadata columns must spread (prefix matching)
    assert any(h and h.startswith("챕터_") for h in headers), headers
    assert any(h and h.startswith("출처_") for h in headers), headers
    assert any(h and h.startswith("난이도_") for h in headers), headers
    assert any(h and h.startswith("예상_") for h in headers), headers
    assert any(h and h.startswith("유형_") for h in headers), headers
    # needs-map join columns
    assert "관심챕터_본인정답률" in headers
    assert "비호감챕터_본인정답률" in headers


def test_학생성적_row_count_matches_master_size(written_xlsx: Path) -> None:
    wb = load_workbook(written_xlsx)
    ws = wb["학생성적"]
    # 1 header + 2 students = 3 rows
    student_ids = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    assert student_ids == ["2026100001", "2026100002"]


def test_학생성적_absent_row_has_empty_score_cells(written_xlsx: Path) -> None:
    wb = load_workbook(written_xlsx)
    ws = wb["학생성적"]
    # Row 3 = 결시 학생
    assert ws.cell(3, 1).value == "2026100002"
    assert ws.cell(3, 4).value == "결시"
    # Score columns 5-9 empty
    for c in range(5, 10):
        assert ws.cell(3, c).value is None


def test_학생성적_taker_row_has_score_filled(written_xlsx: Path) -> None:
    wb = load_workbook(written_xlsx)
    ws = wb["학생성적"]
    # Row 2 = 응시 학생
    assert ws.cell(2, 1).value == "2026100001"
    assert ws.cell(2, 4).value == "응시"
    assert ws.cell(2, 5).value == 4.0  # 총점
    assert ws.cell(2, 6).value == 100.0  # 100점환산


def test_xlsx_without_student_metrics_still_writes_six_sheets(tmp_path: Path) -> None:
    """Backwards compatibility — Phase 3 callers can still skip the 7th sheet."""
    out = tmp_path / "result.xlsx"
    write_analysis_xlsx(
        output_path=out,
        overall_rows=_stub_overall(),
        histogram_bins=_stub_histogram(),
        metadata_rows=_stub_metadata(),
        item_stats=[_stub_item(1)],
        student_metrics=None,
        semester="2026-1",
        course_name_kr="인체구조와기능",
        generated_at_utc="2026-04-29T00:00:00Z",
    )
    wb = load_workbook(out)
    assert "학생성적" not in wb.sheetnames
    assert len(wb.sheetnames) == 6


def test_xlsx_with_student_metrics_two_writes_byte_identical(tmp_path: Path) -> None:
    import hashlib

    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    for path in (a, b):
        write_analysis_xlsx(
            output_path=path,
            overall_rows=_stub_overall(),
            histogram_bins=_stub_histogram(),
            metadata_rows=_stub_metadata(),
            item_stats=[_stub_item(1)],
            student_metrics=_stub_metrics(),
            semester="2026-1",
            course_name_kr="인체구조와기능",
            generated_at_utc="2026-04-29T00:00:00Z",
        )
    sha_a = hashlib.sha256(a.read_bytes()).hexdigest()
    sha_b = hashlib.sha256(b.read_bytes()).hexdigest()
    assert sha_a == sha_b
