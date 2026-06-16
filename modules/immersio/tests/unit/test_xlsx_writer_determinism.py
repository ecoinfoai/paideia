"""T030 — RED tests for `report/xlsx_writer.py::write_analysis_xlsx` (FR-023).

Two-call byte-identical determinism + structural sanity for the 6-sheet
``시험분석결과.xlsx`` (학생성적 시트는 Phase 4 / T052 책임).

Spec 004 contracts/xlsx_sheets.md §1-§6 + research §R-01:
- Workbook.properties.creator / lastModifiedBy fixed to ``paideia/immersio/0.1.0``
- ``created`` / ``modified`` from ``created_at_utc`` (manifest single source, R-10)
- Sheet name & order: 전체요약 / 1_히스토그램 / 2_메타데이터통계 / 3_변별력 /
  4_정답률 / 5_오답분석 — exactly 6 sheets in this build (학생성적 = Phase 4)
- Two consecutive calls with identical inputs produce byte-identical files.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from immersio.report.xlsx_writer import write_analysis_xlsx
from openpyxl import load_workbook
from paideia_shared.schemas import (
    HistogramBin,
    ItemStatistics,
    MetadataAggregate,
)

EXPECTED_SHEETS = (
    "전체요약",
    "1_히스토그램",
    "2_메타데이터통계",
    "3_변별력",
    "4_정답률",
    "5_오답분석",
)


def _stub_overall() -> list[dict[str, object]]:
    return [
        {"지표": "응시자 수", "값": 10},
        {"지표": "결시자 수", "값": 1},
        {"지표": "무응답 응답 수", "값": 2},
        {"지표": "만점", "값": 100.0},
        {"지표": "평균", "값": 75.0},
        {"지표": "표준편차", "값": 10.0},
        {"지표": "중앙값", "값": 75.0},
        {"지표": "최저", "값": 60.0},
        {"지표": "최고", "값": 95.0},
        {"지표": "Q1", "값": 70.0},
        {"지표": "Q3", "값": 80.0},
        {"지표": "100점환산_평균", "값": 75.0},
        {"지표": "100점환산_표준편차", "값": 10.0},
    ]


def _stub_histogram() -> list[HistogramBin]:
    return [
        HistogramBin(bin_start=0.0, bin_end=10.0, count=0, cumulative=0, cumulative_pct=0.0),
        HistogramBin(bin_start=10.0, bin_end=20.0, count=2, cumulative=2, cumulative_pct=20.0),
        HistogramBin(bin_start=20.0, bin_end=30.0, count=8, cumulative=10, cumulative_pct=100.0),
    ]


def _stub_metadata() -> list[MetadataAggregate]:
    return [
        MetadataAggregate(
            metadata_kind="분반",
            metadata_value="A",
            n=5,
            mean=80.0,
            sd=5.0,
            test_kind="ANOVA",
            test_p_value=0.10,
            levene_p_value=None,
            note=None,
        ),
        MetadataAggregate(
            metadata_kind="분반",
            metadata_value="B",
            n=5,
            mean=70.0,
            sd=8.0,
            test_kind="ANOVA",
            test_p_value=0.10,
            levene_p_value=None,
            note=None,
        ),
    ]


def _stub_item(
    item_no: int, *, correct_rate: float = 0.7, label: str = "특이사항 없음"
) -> ItemStatistics:
    remaining = max(0.0, 1.0 - correct_rate)
    top_d = min(0.20, remaining)
    other = max(0.0, remaining - top_d) / 3.0
    return ItemStatistics(
        item_no=item_no,
        semester="2026-1",
        course_slug="anatomy",
        chapter="1장. 서론",
        week=1,
        item_type="지식축적",
        difficulty_level=2,
        expected_difficulty="보통",
        source="교과서",
        correct_answer=1,
        n_responders=10,
        n_correct=int(10 * correct_rate),
        n_omit=0,
        correct_rate=correct_rate,
        omit_rate=0.0,
        discrimination_index=0.30,
        point_biserial=0.30,
        top_distractor_no=2,
        top_distractor_rate=top_d,
        is_top_distractor_adjacent=True,
        option_distribution={1: correct_rate, 2: top_d, 3: other, 4: other, 5: other},
        distractor_label=label,
    )


def _stub_items() -> list[ItemStatistics]:
    return [_stub_item(i) for i in range(1, 4)]


@pytest.fixture
def call_writer(tmp_path: Path):
    def _call(out_path: Path) -> Path:
        write_analysis_xlsx(
            output_path=out_path,
            overall_rows=_stub_overall(),
            histogram_bins=_stub_histogram(),
            metadata_rows=_stub_metadata(),
            item_stats=_stub_items(),
            semester="2026-1",
            course_name_kr="인체구조와기능",
            generated_at_utc="2026-04-29T00:00:00Z",
        )
        return out_path

    return _call


def test_writes_six_sheets_in_canonical_order(tmp_path: Path, call_writer) -> None:
    out = tmp_path / "result.xlsx"
    call_writer(out)
    wb = load_workbook(out)
    assert tuple(wb.sheetnames) == EXPECTED_SHEETS


def test_workbook_properties_pinned_to_generated_at_utc(tmp_path: Path, call_writer) -> None:
    out = tmp_path / "result.xlsx"
    call_writer(out)
    wb = load_workbook(out)
    props = wb.properties
    assert props.creator == "paideia/immersio/0.1.0"
    assert props.lastModifiedBy == "paideia/immersio/0.1.0"
    # created / modified should reflect the pinned timestamp
    assert props.created is not None
    assert props.created.year == 2026 and props.created.month == 4 and props.created.day == 29
    assert props.modified == props.created


def test_overall_summary_sheet_has_13_rows(tmp_path: Path, call_writer) -> None:
    out = tmp_path / "result.xlsx"
    call_writer(out)
    wb = load_workbook(out)
    ws = wb["전체요약"]
    # Row 1 = header (지표 / 값); rows 2-14 carry the 13 data rows.
    assert ws.cell(1, 1).value == "지표"
    assert ws.cell(1, 2).value == "값"
    labels = [ws.cell(r, 1).value for r in range(2, 15)]
    assert labels[0] == "응시자 수"
    assert labels[-1] == "100점환산_표준편차"
    assert len(labels) == 13


def test_two_calls_produce_byte_identical_xlsx(tmp_path: Path, call_writer) -> None:
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    call_writer(a)
    call_writer(b)
    sha_a = hashlib.sha256(a.read_bytes()).hexdigest()
    sha_b = hashlib.sha256(b.read_bytes()).hexdigest()
    assert sha_a == sha_b, "xlsx bytes diverge across two identical writes"


def test_histogram_sheet_columns_match_contract(tmp_path: Path, call_writer) -> None:
    out = tmp_path / "result.xlsx"
    call_writer(out)
    wb = load_workbook(out)
    ws = wb["1_히스토그램"]
    # Find header row (which includes 구간_시작 / 구간_끝 / 도수 / 누적 / 누적_백분율)
    header_row_idx = None
    for r in range(1, ws.max_row + 1):
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if row_values[0] == "구간_시작":
            header_row_idx = r
            break
    assert header_row_idx is not None, "구간_시작 header not found in 1_히스토그램"
    headers = [ws.cell(header_row_idx, c).value for c in range(1, 6)]
    assert headers == ["구간_시작", "구간_끝", "도수", "누적", "누적_백분율"]


def test_metadata_sheet_columns_match_contract(tmp_path: Path, call_writer) -> None:
    out = tmp_path / "result.xlsx"
    call_writer(out)
    wb = load_workbook(out)
    ws = wb["2_메타데이터통계"]
    header_row_idx = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "metadata_kind":
            header_row_idx = r
            break
    assert header_row_idx is not None
    headers = [ws.cell(header_row_idx, c).value for c in range(1, 10)]
    assert headers == [
        "metadata_kind",
        "metadata_value",
        "n",
        "mean",
        "sd",
        "test_kind",
        "test_p_value",
        "levene_p_value",
        "note",
    ]


def test_distractor_sheet_lists_label_column(tmp_path: Path, call_writer) -> None:
    out = tmp_path / "result.xlsx"
    call_writer(out)
    wb = load_workbook(out)
    ws = wb["5_오답분석"]
    found = False
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value == "distractor_label":
                found = True
                break
        if found:
            break
    assert found, "distractor_label column header not found in 5_오답분석"


def test_writer_rejects_empty_items(tmp_path: Path) -> None:
    out = tmp_path / "result.xlsx"
    with pytest.raises(ValueError):
        write_analysis_xlsx(
            output_path=out,
            overall_rows=_stub_overall(),
            histogram_bins=_stub_histogram(),
            metadata_rows=_stub_metadata(),
            item_stats=[],
            semester="2026-1",
            course_name_kr="인체구조와기능",
            generated_at_utc="2026-04-29T00:00:00Z",
        )
