"""TDD tests for ``combine.xlsx_writer`` US1 sheets (T032, US1).

Verifies the workbook writer:
- 2 sheets: `상관매트릭스`, `회귀결과` present
- column headers match contracts/xlsx_workbook.md
- byte-identical re-run via Phase 1+2 ``rewrite_modified_in_zip``
- ``<dcterms:modified>`` epoch normalised
- empty inputs rejected (Fail-Fast)

T040 (US2 cluster) + T054 (US4 subgroup) extend the workbook with
additional sheets; this test set covers the US1 partial mode only.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from immersio.combine.xlsx_writer import write_us1_xlsx
from paideia_shared.schemas import (
    CorrelationCell,
    RegressionCoefficient,
    RegressionFitSummary,
)
from paideia_shared.schemas._common import STANDARD_AXIS_KEYS


def _cells() -> list[CorrelationCell]:
    return [
        CorrelationCell(
            axis_key=axis,
            exam_metric_key=metric,
            n=22,
            pearson_r=0.3,
            raw_p=0.01,
            fdr_q=0.04,
            significant_after_correction=True,
            unstable_inference_flag=False,
        )
        for axis in STANDARD_AXIS_KEYS
        for metric in ("total_score", "chapter_신경계")
    ]


def _coefs() -> list[RegressionCoefficient]:
    out: list[RegressionCoefficient] = []
    for i, axis in enumerate(STANDARD_AXIS_KEYS):
        coef = 0.5 + i * 0.1
        out.append(
            RegressionCoefficient(
                axis_key=axis,
                coef=coef,
                std_err=0.2,
                t_stat=2.5 + i * 0.5,
                raw_p=0.01,
                fdr_q=0.04,
                ci_low_95=coef - 0.5,
                ci_high_95=coef + 0.5,
                beta_standardized=0.1 + i * 0.02,
                vif=1.5,
                multicollinearity_flag=False,
            )
        )
    return out


def _fit() -> RegressionFitSummary:
    return RegressionFitSummary(
        n_complete_case=22,
        n_dropped=8,
        r2=0.45,
        r2_adj=0.32,
        f_stat=5.0,
        f_pvalue=0.001,
        regression_method="OLS",
        multiple_comparison_method="BH-FDR",
        small_sample_warning=True,
    )


# ----------------------------------------------------------------------
# Smoke
# ----------------------------------------------------------------------


def test_writes_xlsx_file(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    assert out.exists()
    assert out.stat().st_size > 0


def test_two_us1_sheets_present(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    assert "상관매트릭스" in wb.sheetnames
    assert "회귀결과" in wb.sheetnames


# ----------------------------------------------------------------------
# Sheet 1 — 상관매트릭스
# ----------------------------------------------------------------------


def test_correlation_sheet_headers(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    sheet = wb["상관매트릭스"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    expected = [
        "axis_key",
        "exam_metric_key",
        "n",
        "pearson_r",
        "raw_p",
        "fdr_q",
        "significant",
        "unstable_n_lt_20",
    ]
    assert headers == expected


def test_correlation_sheet_row_count(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    cells = _cells()
    write_us1_xlsx(
        correlation_cells=cells,
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    sheet = wb["상관매트릭스"]
    # 1 header + len(cells) data rows.
    assert sheet.max_row == 1 + len(cells)


# ----------------------------------------------------------------------
# Sheet 2 — 회귀결과
# ----------------------------------------------------------------------


def test_regression_sheet_has_fit_summary_block(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    sheet = wb["회귀결과"]
    # Fit summary block in the first 3 rows; coef table starts after.
    text_dump = "\n".join(
        " ".join(str(c.value) if c.value is not None else "" for c in row)
        for row in sheet.iter_rows(max_row=10)
    )
    assert "n_complete_case" in text_dump
    assert "R²" in text_dump or "r2" in text_dump
    assert "OLS" in text_dump


def test_regression_sheet_coef_headers(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    sheet = wb["회귀결과"]
    # Locate the row that begins with "axis_key" (coef table header).
    header_row = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and row[0] == "axis_key":
            header_row = list(row)
            break
    assert header_row is not None, "회귀결과 sheet missing 'axis_key' header row"
    expected_prefix = [
        "axis_key",
        "coef",
        "std_err",
        "t_stat",
        "raw_p",
        "fdr_q",
        "ci_low_95",
        "ci_high_95",
        "beta_standardized",
        "vif",
        "multicollinearity_flag",
    ]
    assert header_row[: len(expected_prefix)] == expected_prefix


def test_regression_sheet_8_coef_rows(tmp_path: Path) -> None:
    out = tmp_path / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    sheet = wb["회귀결과"]
    axis_rows = [
        row[0]
        for row in sheet.iter_rows(values_only=True)
        if row and row[0] in set(STANDARD_AXIS_KEYS)
    ]
    assert len(axis_rows) == 8


# ----------------------------------------------------------------------
# Determinism — vector #5 (dcterms:modified + zip date_time)
# ----------------------------------------------------------------------


def test_byte_identical_re_run(tmp_path: Path) -> None:
    """Phase 1+2 ``rewrite_modified_in_zip`` re-application keeps both runs equal."""
    out1 = tmp_path / "r1.xlsx"
    out2 = tmp_path / "r2.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out1,
    )
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out2,
    )
    assert out1.read_bytes() == out2.read_bytes()


def test_dcterms_created_pinned_to_epoch(tmp_path: Path) -> None:
    """``<dcterms:created>`` must also be pinned (not just ``modified``).

    Without explicit ``wb.properties.created`` set, openpyxl writes
    ``datetime.now()`` into core.xml so two runs with non-zero wall-clock
    delta diverge. This test asserts the pinned epoch lands.
    """
    import io
    import zipfile

    out = tmp_path / "epoch.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    with zipfile.ZipFile(out, "r") as zf:
        core = zf.read("docProps/core.xml").decode("utf-8")
    assert "<dcterms:created" in core
    assert "2000-01-01T00:00:00Z" in core
    assert core.count("2000-01-01T00:00:00Z") >= 2  # both created + modified


def test_byte_identical_across_wallclock_delta(tmp_path: Path) -> None:
    """Re-runs separated by >1s wall-clock delta must still match.

    Reproduces the e2e regression where xlsx differed only in
    ``<dcterms:created>`` because openpyxl's default uses ``datetime.now()``.
    """
    import time

    out1 = tmp_path / "wc1.xlsx"
    out2 = tmp_path / "wc2.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out1,
    )
    time.sleep(1.1)
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out2,
    )
    assert out1.read_bytes() == out2.read_bytes()


# ----------------------------------------------------------------------
# Fail-Fast
# ----------------------------------------------------------------------


def test_empty_correlation_rejected(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="empty"):
        write_us1_xlsx(
            correlation_cells=[],
            regression_coefs=_coefs(),
            regression_fit=_fit(),
            out_path=tmp_path / "x.xlsx",
        )


def test_empty_regression_rejected(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="empty"):
        write_us1_xlsx(
            correlation_cells=_cells(),
            regression_coefs=[],
            regression_fit=_fit(),
            out_path=tmp_path / "x.xlsx",
        )


def test_creates_parent_directory(tmp_path: Path) -> None:
    nested = tmp_path / "deep" / "nest" / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=nested,
    )
    assert nested.exists()
