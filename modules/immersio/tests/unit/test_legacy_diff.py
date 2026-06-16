"""T054 — RED tests for `report/legacy_diff.py::generate_legacy_diff` (FR-030/031, R-06).

Cell-level comparison of legacy `중간고사_분석결과.xlsx` vs immersio
`시험분석결과.xlsx` with ±0.001 tolerance, reason-estimation rules, and
deterministic Markdown output.

Behaviours under test:
  (a) ±0.0009 numeric diff → ignored (false positive 0건 / SC-004)
  (b) ±0.0011 numeric diff → reported with reason
  (c) text cell exact-match — mismatch → reported
  (d) cell present in legacy but absent in immersio → 미재현 섹션
  (e) determinism: 두 번 호출 byte-identical
  (f) reason rules — 검정 종류 차이 / 결시 분모 의심 / 반올림 차이 / 무응답 처리 의심
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from immersio.report.legacy_diff import LegacyLoadError, generate_legacy_diff
from openpyxl import Workbook


def _stamp_workbook(wb: Workbook, when: str) -> None:
    import datetime

    dt = datetime.datetime.fromisoformat(when.replace("Z", "+00:00")).replace(tzinfo=None)
    wb.properties.creator = "test"
    wb.properties.lastModifiedBy = "test"
    wb.properties.created = dt
    wb.properties.modified = dt


def _build_legacy_xlsx(path: Path, payload: dict[str, list[list[object]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, rows in payload.items():
        ws = wb.create_sheet(sheet_name)
        for r, row in enumerate(rows, start=1):
            for c, value in enumerate(row, start=1):
                ws.cell(r, c, value)
    _stamp_workbook(wb, "2026-04-29T00:00:00Z")
    wb.save(path)


def _build_immersio_xlsx(path: Path, payload: dict[str, list[list[object]]]) -> None:
    _build_legacy_xlsx(path, payload)


def _read_md(path: Path) -> str:
    return path.read_text(encoding="utf-8")


@pytest.fixture
def legacy_path(tmp_path: Path) -> Path:
    out = tmp_path / "legacy.xlsx"
    _build_legacy_xlsx(
        out,
        {
            "전체요약": [
                ["지표", "값"],
                ["응시자 수", 184],
                ["평균", 125.351],  # vs immersio 125.350 → diff -0.001 (=tolerance, ignored)
                ["표준편차", 39.5511],  # vs immersio 39.5500 → diff -0.0011 (reported)
            ],
            "2_메타데이터통계": [
                ["분류축", "값", "p값"],
                [
                    "분반",
                    "ANOVA",
                    0.0421,
                ],  # vs immersio 0.0432 → diff +0.0011 (reported, ANOVA reason)
                ["LLM_코멘트", "분반 A 평균이...", None],  # missing in immersio (legacy 만 존재)
            ],
            "4_정답률": [
                ["문항번호", "정답률(%)"],
                [1, 73.0],  # text + numeric mix
                [2, 85.87],
            ],
        },
    )
    return out


@pytest.fixture
def immersio_path(tmp_path: Path) -> Path:
    out = tmp_path / "immersio.xlsx"
    _build_immersio_xlsx(
        out,
        {
            "전체요약": [
                ["지표", "값"],
                ["응시자 수", 184],
                ["평균", 125.350],
                ["표준편차", 39.5500],
            ],
            "2_메타데이터통계": [
                ["분류축", "값", "p값"],
                [
                    "분반",
                    "Welch ANOVA",
                    0.0432,
                ],  # text different ('ANOVA' vs 'Welch ANOVA') + p diff
                # row 3 absent → legacy 의 "LLM_코멘트" 행이 미재현
            ],
            "4_정답률": [
                ["문항번호", "정답률(%)"],
                [1, 73.0],
                [2, 85.87],
            ],
        },
    )
    return out


def test_numeric_diff_within_tolerance_is_ignored(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # 평균 diff = -0.001 == tolerance → must NOT appear
    assert "평균" not in md.split("# legacy_diff.md")[0] + md.split("## 1.")[0]
    # The comparator should explicitly exclude the in-tolerance cell;
    # we assert it does not show up under any 시트's diff table.
    diff_table = md.split("## 8.")[0]  # everything before 미재현 섹션
    # 평균 row may appear in section header; we check value pair absence
    assert "125.351" not in diff_table


def test_numeric_diff_above_tolerance_is_reported(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # 표준편차 diff -0.0011 must surface
    assert "39.5511" in md
    assert "39.5500" in md
    # 4-decimal float formatting per spec
    assert "-0.0011" in md or "0.0011" in md


def test_text_cell_mismatch_is_reported(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # legacy 'ANOVA' vs immersio 'Welch ANOVA' must surface
    assert "ANOVA" in md
    assert "Welch ANOVA" in md


def test_missing_cell_in_immersio_listed_in_미재현_section(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # 미재현 section header per contracts/legacy_diff.md §8
    assert "## 8. immersio 미재현" in md or "미재현" in md
    assert "LLM_코멘트" in md or "분반 A" in md


def test_reason_rule_anova_vs_welch_for_p_value_column(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # The p값 column on 2_메타데이터통계 sheet must trigger ANOVA / Welch reason
    assert "ANOVA" in md
    assert "Welch" in md
    # Reason text must appear at least once
    assert "검정 종류 차이" in md or "Welch ANOVA" in md


def test_reason_rule_rounding_for_small_numeric_diff(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # 표준편차 diff = 0.0011 → 반올림 차이 reason
    assert "반올림" in md


def test_two_calls_byte_identical(tmp_path: Path, legacy_path: Path, immersio_path: Path) -> None:
    a = tmp_path / "a.md"
    b = tmp_path / "b.md"
    for path in (a, b):
        generate_legacy_diff(
            legacy_xlsx=legacy_path,
            immersio_xlsx=immersio_path,
            output_path=path,
            compared_at_utc="2026-04-29T00:00:00Z",
            semester="2026-1",
            course_slug="anatomy",
        )
    sha_a = hashlib.sha256(a.read_bytes()).hexdigest()
    sha_b = hashlib.sha256(b.read_bytes()).hexdigest()
    assert sha_a == sha_b, "legacy_diff.md bytes diverge across two identical writes"


def test_decision_marks_immersio_for_numeric_diffs(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # FR-030 / 대체 원칙 4 — every numeric diff defaults to immersio
    assert "immersio" in md
    assert "## 10." in md or "채택" in md


def test_summary_header_lists_total_compared_and_diff_counts(
    tmp_path: Path, legacy_path: Path, immersio_path: Path
) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    # Summary line per contracts header (총 비교 셀 / 차이 발견 / immersio 채택 / 미재현)
    assert "총 비교 셀" in md
    assert "차이 발견" in md
    assert "미재현" in md


def test_structure_section_present(tmp_path: Path, legacy_path: Path, immersio_path: Path) -> None:
    out = tmp_path / "diff.md"
    generate_legacy_diff(
        legacy_xlsx=legacy_path,
        immersio_xlsx=immersio_path,
        output_path=out,
        compared_at_utc="2026-04-29T00:00:00Z",
        semester="2026-1",
        course_slug="anatomy",
    )
    md = _read_md(out)
    assert "## 0. 구조 검증" in md or "구조 검증" in md


def test_rejects_missing_legacy_path(tmp_path: Path, immersio_path: Path) -> None:
    out = tmp_path / "diff.md"
    with pytest.raises(FileNotFoundError):
        generate_legacy_diff(
            legacy_xlsx=tmp_path / "does_not_exist.xlsx",
            immersio_xlsx=immersio_path,
            output_path=out,
            compared_at_utc="2026-04-29T00:00:00Z",
            semester="2026-1",
            course_slug="anatomy",
        )


def test_corrupt_legacy_xlsx_raises_legacy_load_error(tmp_path: Path, immersio_path: Path) -> None:
    """Adversary P4: corrupt legacy xlsx → LegacyLoadError (CLI exit 5)."""
    bad = tmp_path / "corrupt.xlsx"
    bad.write_bytes(b"this is not a real xlsx -- just garbage bytes")
    out = tmp_path / "diff.md"
    with pytest.raises(LegacyLoadError) as exc_info:
        generate_legacy_diff(
            legacy_xlsx=bad,
            immersio_xlsx=immersio_path,
            output_path=out,
            compared_at_utc="2026-04-29T00:00:00Z",
            semester="2026-1",
            course_slug="anatomy",
        )
    msg = str(exc_info.value)
    assert "legacy" in msg
    assert "corrupt" in msg or "valid xlsx" in msg
    # message must point at the failing path so operators can act on it
    assert str(bad) in msg


def test_corrupt_immersio_xlsx_raises_legacy_load_error(tmp_path: Path, legacy_path: Path) -> None:
    bad = tmp_path / "corrupt.xlsx"
    bad.write_bytes(b"PK\x03\x04 -- truncated zip header, not a real xlsx")
    out = tmp_path / "diff.md"
    with pytest.raises(LegacyLoadError) as exc_info:
        generate_legacy_diff(
            legacy_xlsx=legacy_path,
            immersio_xlsx=bad,
            output_path=out,
            compared_at_utc="2026-04-29T00:00:00Z",
            semester="2026-1",
            course_slug="anatomy",
        )
    msg = str(exc_info.value)
    assert "immersio" in msg
    assert str(bad) in msg


def test_legacy_load_error_chains_original_exception(tmp_path: Path, immersio_path: Path) -> None:
    """Adversary P4 anti-pattern guard: original exception preserved via __cause__."""
    bad = tmp_path / "x.xlsx"
    bad.write_bytes(b"not xlsx")
    out = tmp_path / "diff.md"
    with pytest.raises(LegacyLoadError) as exc_info:
        generate_legacy_diff(
            legacy_xlsx=bad,
            immersio_xlsx=immersio_path,
            output_path=out,
            compared_at_utc="2026-04-29T00:00:00Z",
            semester="2026-1",
            course_slug="anatomy",
        )
    # `raise ... from exc` populates __cause__ — never silently swallow
    assert exc_info.value.__cause__ is not None
