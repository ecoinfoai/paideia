"""Unit tests for the roster-only attendance branch (T014, FR-027).

Spec 004 research §R-08a — parse_attendance_xlsx 가
- (a) 기존 ('학번', '이름', W01..W16, '비고') 형식 → 기존 동작
- (b) 신규 ('sn', '분반', '출석부연번', '학번', '이름', ...) 형식 → roster-only
- (c) 빈 시트 → ValueError
- (d) 어느 헤더와도 매칭 안 됨 → ValueError
"""

from __future__ import annotations

from pathlib import Path

# Import ``immersio.ingest`` first to break the io ↔ ingest circular import.
# Both ``io.attendance`` (this module's target) and ``ingest.pipeline`` cross-
# reference each other; loading ``immersio.ingest`` up-front populates
# ``sys.modules`` so the subsequent ``from immersio.io import ...`` can resolve
# ``parse_attendance_xlsx`` without re-entering a partially initialized
# ``io/__init__``. This mirrors how the existing integration tests succeed
# (they import ingest before io transitively).
import immersio.ingest  # noqa: F401  # required-for: io ↔ ingest import order
import pandas as pd  # noqa: F401  # imported for fixture / type clarity
import pytest
from immersio.io import parse_attendance_xlsx
from openpyxl import Workbook


def _build_w01_w16_attendance(target: Path) -> None:
    """Build a small W01..W16 fixture identical to existing builder format."""
    wb = Workbook()
    sh = wb.active
    sh.title = "출석"
    header = ["학번", "이름"] + [f"W{week:02d}" for week in range(1, 17)] + ["비고"]
    sh.append(header)
    sh.append(["2026000001", "학생A"] + ["O"] * 16 + [""])
    sh.append(["2026000002", "학생B"] + ["O"] * 14 + ["X", "O", ""])
    wb.save(target)


def _build_roster_only(target: Path) -> None:
    """Build a roster-only namesheet (no W01..W16, no attendance codes)."""
    wb = Workbook()
    sh = wb.active
    sh.title = "출석부"
    header = ["sn", "분반", "출석부연번", "학번", "이름", "비고"]
    sh.append(header)
    sh.append([1, "A", 1, "2026000001", "학생A", ""])
    sh.append([2, "A", 2, "2026000002", "학생B", ""])
    sh.append([3, "B", 1, "2026000003", "학생C", ""])
    wb.save(target)


def _build_empty(target: Path) -> None:
    wb = Workbook()
    sh = wb.active
    sh.title = "출석"
    wb.save(target)


def _build_unknown_header(target: Path) -> None:
    wb = Workbook()
    sh = wb.active
    sh.title = "출석"
    sh.append(["foo", "bar", "baz"])
    sh.append(["a", "b", "c"])
    wb.save(target)


def test_attendance_w01_w16_format_works(tmp_path: Path) -> None:
    target = tmp_path / "출석부.xlsx"
    _build_w01_w16_attendance(target)
    df = parse_attendance_xlsx(target)
    assert list(df["student_id"]) == ["2026000001", "2026000002"]
    assert df.loc[df["student_id"] == "2026000002", "attendance_absent_count"].iloc[0] == 1
    # Existing format → attendance counts populated, not None.
    assert df["attendance_present_count"].notna().all()


def test_attendance_roster_only_format_returns_section_and_no_counts(tmp_path: Path) -> None:
    target = tmp_path / "출석부.xlsx"
    _build_roster_only(target)
    df = parse_attendance_xlsx(target)
    assert set(df["student_id"]) == {"2026000001", "2026000002", "2026000003"}
    # roster-only → attendance_count 컬럼은 모두 None
    for col in (
        "attendance_present_count",
        "attendance_absent_count",
        "attendance_late_count",
        "attendance_excused_count",
    ):
        assert df[col].isna().all(), f"{col} must be None in roster-only mode"
    # section 컬럼은 추출됨
    assert "section" in df.columns
    assert set(df["section"]) == {"A", "B"}
    # name_kr 또한 정상 추출
    assert df.loc[df["student_id"] == "2026000003", "name_kr"].iloc[0] == "학생C"


def test_attendance_empty_workbook_raises(tmp_path: Path) -> None:
    target = tmp_path / "출석부.xlsx"
    _build_empty(target)
    with pytest.raises(ValueError, match=r"empty|header"):
        parse_attendance_xlsx(target)


def test_attendance_unknown_header_raises(tmp_path: Path) -> None:
    target = tmp_path / "출석부.xlsx"
    _build_unknown_header(target)
    with pytest.raises(ValueError, match="header"):
        parse_attendance_xlsx(target)
