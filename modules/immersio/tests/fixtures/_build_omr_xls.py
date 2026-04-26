"""Build synthetic OMR XLS fixtures for the bronze_minimal tree.

The department OMR system emits four ``.xls`` files (one per section A/B/C/D),
each holding four sheets named 결과·결시·OX·문항분석. To keep the test
suite self-contained without binary blobs in git, this script regenerates
the fixtures from declarative Python data.

Outputs are written as ``.xlsx`` files because the standard openpyxl
writer is the simplest deterministic builder available; the ingest parser
auto-routes between ``.xls`` and ``.xlsx`` via pandas engines.

Run:
    uv run python modules/immersio/tests/fixtures/_build_omr_xls.py
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Final

from openpyxl import Workbook

ROSTER_BY_SECTION: Final[dict[str, list[tuple[str, str]]]] = {
    "A": [("2026000001", "학생A"), ("2026099002", "응시외G")],
    "B": [("2026000002", "학생B"), ("2026000003", "학생C")],
    "C": [("2026000004", "학생D")],
    "D": [("2026000005", "학생E")],
}

ITEMS: Final[list[dict[str, object]]] = [
    {
        "item_no": 1,
        "answer_key": "3",
        "chapter": "1장 세포",
        "source": "textbook",
        "expected_difficulty": "easy",
        "bloom": "knowledge",
        "points": 2.0,
        "text": "세포막의 기본 구성 분자는?",
    },
    {
        "item_no": 2,
        "answer_key": "1",
        "chapter": "2장 조직",
        "source": "textbook",
        "expected_difficulty": "medium",
        "bloom": "comprehension",
        "points": 2.0,
        "text": "상피조직의 분류 기준 중 옳은 것은?",
    },
    {
        "item_no": 3,
        "answer_key": "4",
        "chapter": "5장 신경계",
        "source": "formative",
        "expected_difficulty": "hard",
        "bloom": "application",
        "points": 2.5,
        "text": "축삭의 활동전위 전달 속도에 영향을 주는 요인은?",
    },
    {
        "item_no": 4,
        "answer_key": "2",
        "chapter": "6장 근육계",
        "source": "quiz",
        "expected_difficulty": "medium",
        "bloom": "comprehension",
        "points": 2.0,
        "text": "골격근 수축의 슬라이딩 필라멘트 이론에서 가장 중요한 단백질은?",
    },
    {
        "item_no": 5,
        "answer_key": "5",
        "chapter": "7장 순환계",
        "source": "textbook",
        "expected_difficulty": "easy",
        "bloom": "knowledge",
        "points": 1.5,
        "text": "심장의 4 chamber 중 좌심실의 역할은?",
    },
]

# Per-student answer matrix. None means blank cell (no answer).
# "0" string means the student wrote 0 (different from blank).
RESPONSES_BY_SECTION: Final[dict[str, dict[str, list[str | None]]]] = {
    "A": {
        "2026000001": ["3", "1", "4", "2", "5"],
        "2026099002": ["3", "1", "4", "2", None],
    },
    "B": {
        "2026000002": ["3", "2", "4", "2", "5"],
        "2026000003": ["3", "1", "4", "0", "5"],  # "0" intentional (T040a)
    },
    "C": {
        "2026000004": ["3", "1", "1", "2", "5"],
    },
    "D": {
        # Student E is intentionally absent (no row in 결과 sheet, listed in 결시).
    },
}

ABSENT_BY_SECTION: Final[dict[str, list[str]]] = {
    "A": [],
    "B": [],
    "C": [],
    "D": ["2026000005"],
}


def _write_section_workbook(path: Path, section: str) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    # 결과 sheet: 학번, 이름, item_1, item_2, ..., item_N, 점수
    results_sheet = workbook.create_sheet("결과")
    header = ["학번", "이름"] + [f"item_{item['item_no']}" for item in ITEMS] + ["점수"]
    results_sheet.append(header)
    responses = RESPONSES_BY_SECTION[section]
    roster = ROSTER_BY_SECTION[section]
    for student_id, name in roster:
        if student_id not in responses:
            continue
        answers = responses[student_id]
        score = sum(
            float(item["points"])
            for item, answer in zip(ITEMS, answers, strict=True)
            if answer is not None and answer == item["answer_key"]
        )
        row: list[object] = [student_id, name]
        for answer in answers:
            row.append("" if answer is None else answer)
        row.append(score)
        results_sheet.append(row)

    # 결시 sheet: 학번, 이름
    absent_sheet = workbook.create_sheet("결시")
    absent_sheet.append(["학번", "이름"])
    for absent_id in ABSENT_BY_SECTION[section]:
        for sid, sname in roster:
            if sid == absent_id:
                absent_sheet.append([sid, sname])

    # OX sheet: 학번, item_1_OX, item_2_OX, ..., item_N_OX
    ox_sheet = workbook.create_sheet("OX")
    ox_sheet.append(
        ["학번"] + [f"item_{item['item_no']}_OX" for item in ITEMS]
    )
    for student_id, _name in roster:
        if student_id not in responses:
            continue
        answers = responses[student_id]
        ox_row: list[object] = [student_id]
        for item, answer in zip(ITEMS, answers, strict=True):
            if answer is None:
                ox_row.append("")
            elif answer == item["answer_key"]:
                ox_row.append("O")
            else:
                ox_row.append("X")
        ox_sheet.append(ox_row)

    # 문항분석 sheet: item_no, chapter, source, expected_difficulty, bloom,
    #                answer_key, points, text
    analysis_sheet = workbook.create_sheet("문항분석")
    analysis_sheet.append(
        [
            "item_no",
            "chapter",
            "source",
            "expected_difficulty",
            "bloom",
            "answer_key",
            "points",
            "text",
        ]
    )
    for item in ITEMS:
        analysis_sheet.append(
            [
                item["item_no"],
                item["chapter"],
                item["source"],
                item["expected_difficulty"],
                item["bloom"],
                item["answer_key"],
                item["points"],
                item["text"],
            ]
        )

    workbook.save(path)


def build_all(target_dir: Path) -> Iterable[Path]:
    target_dir.mkdir(parents=True, exist_ok=True)
    for section in ("A", "B", "C", "D"):
        target = target_dir / f"인체구조와기능_{section}반_결과.xlsx"
        _write_section_workbook(target, section)
        yield target


if __name__ == "__main__":
    base = Path(__file__).parent / "bronze_minimal" / "시험성적"
    for path in build_all(base):
        print(f"wrote {path}")
