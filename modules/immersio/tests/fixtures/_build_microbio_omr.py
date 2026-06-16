"""Build microbiology OMR XLS fixtures (US3 portability)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

ROSTER = {
    "A": [("2027000001", "학생가")],
    "B": [("2027000002", "학생나")],
    "C": [("2027000003", "학생다")],
    "D": [],
}

ITEMS = [
    {
        "item_no": 1,
        "answer_key": "2",
        "chapter": "1장 세균학",
        "source": "textbook",
        "expected_difficulty": "easy",
        "bloom": "knowledge",
        "points": 2.0,
        "text": "그람 양성균의 세포벽 주성분?",
    },
    {
        "item_no": 2,
        "answer_key": "5",
        "chapter": "2장 바이러스학",
        "source": "textbook",
        "expected_difficulty": "medium",
        "bloom": "comprehension",
        "points": 2.5,
        "text": "RNA 바이러스의 일반적 특징은?",
    },
    {
        "item_no": 3,
        "answer_key": "1",
        "chapter": "3장 진균학",
        "source": "quiz",
        "expected_difficulty": "hard",
        "bloom": "application",
        "points": 2.0,
        "text": "효모의 출아 생식이 가장 잘 일어나는 환경은?",
    },
]

RESPONSES = {
    "A": {"2027000001": ["2", "5", "1"]},
    "B": {"2027000002": ["2", "1", "1"]},
    "C": {"2027000003": ["3", "5", "2"]},
}


def _section_workbook(target: Path, section: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    results = wb.create_sheet("결과")
    results.append(["학번", "이름"] + [f"item_{item['item_no']}" for item in ITEMS] + ["점수"])
    for sid, name in ROSTER[section]:
        if sid not in RESPONSES.get(section, {}):
            continue
        answers = RESPONSES[section][sid]
        score = sum(
            float(item["points"])
            for item, ans in zip(ITEMS, answers, strict=True)
            if ans == item["answer_key"]
        )
        results.append([sid, name, *answers, score])

    wb.create_sheet("결시").append(["학번", "이름"])

    ox = wb.create_sheet("OX")
    ox.append(["학번"] + [f"item_{item['item_no']}_OX" for item in ITEMS])
    for sid, _name in ROSTER[section]:
        if sid not in RESPONSES.get(section, {}):
            continue
        answers = RESPONSES[section][sid]
        ox.append(
            [sid]
            + [
                "O" if ans == item["answer_key"] else "X"
                for item, ans in zip(ITEMS, answers, strict=True)
            ]
        )

    analysis = wb.create_sheet("문항분석")
    analysis.append(
        [
            "item_no",
            "chapter",
            "source",
            "expected_difficulty",
            "bloom",
            "answer_key",
            "points",
            "text",
        ]
    )
    for item in ITEMS:
        analysis.append(
            [
                item["item_no"],
                item["chapter"],
                item["source"],
                item["expected_difficulty"],
                item["bloom"],
                item["answer_key"],
                item["points"],
                item["text"],
            ]
        )

    wb.save(target)


if __name__ == "__main__":
    base = Path(__file__).parent / "bronze_minimal_microbio" / "시험성적"
    base.mkdir(parents=True, exist_ok=True)
    for section in ("A", "B", "C", "D"):
        if not ROSTER[section]:
            # Department format expects all four sections; emit empty workbook anyway.
            wb = Workbook()
            wb.remove(wb.active)
            for sheet_name in ("결과", "결시", "OX"):
                wb.create_sheet(sheet_name).append(
                    ["학번", "이름"]
                    if sheet_name != "OX"
                    else ["학번"] + [f"item_{item['item_no']}_OX" for item in ITEMS]
                )
            analysis = wb.create_sheet("문항분석")
            analysis.append(
                [
                    "item_no",
                    "chapter",
                    "source",
                    "expected_difficulty",
                    "bloom",
                    "answer_key",
                    "points",
                    "text",
                ]
            )
            for item in ITEMS:
                analysis.append(
                    [
                        item["item_no"],
                        item["chapter"],
                        item["source"],
                        item["expected_difficulty"],
                        item["bloom"],
                        item["answer_key"],
                        item["points"],
                        item["text"],
                    ]
                )
            target = base / f"미생물학_{section}반_결과.xlsx"
            wb.save(target)
            print(f"wrote (empty) {target}")
            continue
        target = base / f"미생물학_{section}반_결과.xlsx"
        _section_workbook(target, section)
        print(f"wrote {target}")
