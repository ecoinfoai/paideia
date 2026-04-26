"""contracts/cli.md exit code 4 — data integrity (duplicate student_id)."""

from __future__ import annotations

import shutil
from pathlib import Path

from immersio.cli.main import app
from openpyxl import load_workbook

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures"
GOOD_BRONZE = FIXTURES / "bronze_minimal"
GOOD_MAPPING = FIXTURES / "mappings" / "anatomy.diagnostic.yaml"


def test_cli_returns_4_on_duplicate_student_id(tmp_path: Path) -> None:
    bronze = tmp_path / "bronze"
    shutil.copytree(GOOD_BRONZE, bronze)

    target = bronze / "시험성적" / "인체구조와기능_A반_결과.xlsx"
    workbook = load_workbook(target)
    sheet = workbook["결과"]
    duplicate_id = sheet.cell(row=2, column=1).value
    sheet.cell(row=3, column=1).value = duplicate_id
    workbook.save(target)

    out = tmp_path / "silver"
    code = app(
        argv=[
            "ingest",
            "--bronze-dir",
            str(bronze),
            "--mapping",
            str(GOOD_MAPPING),
            "--output-dir",
            str(out),
            "--no-git-commit",
        ]
    )
    assert code == 4
    assert not (out / "2026-1-anatomy").exists()
