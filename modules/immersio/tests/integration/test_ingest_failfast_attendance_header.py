"""US2: attendance template header drift triggers Fail-Fast (T058a)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from immersio.ingest import IngestValidationError, run_ingest


def test_attendance_header_drift_blocks_silver(
    corrupt_bronze: Path, corrupt_mapping: Path, tmp_path: Path
) -> None:
    target = corrupt_bronze / "출석" / "출석부.xlsx"
    workbook = load_workbook(target)
    sheet = workbook.active
    # Replace W01 with Week01.
    sheet.cell(row=1, column=3).value = "Week01"
    workbook.save(target)

    out = tmp_path / "silver"
    with pytest.raises(IngestValidationError) as exc:
        run_ingest(
            bronze_dir=corrupt_bronze,
            mapping_path=corrupt_mapping,
            output_dir=out,
            no_git_commit=True,
        )
    rendered = str(exc.value)
    assert "header mismatch" in rendered or "Week01" in rendered or "W01" in rendered
    assert not (out / "2026-1-anatomy").exists()
