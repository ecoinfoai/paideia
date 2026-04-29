"""T066 — End-to-end Phase 1+2 pipeline (FR-032/033/034).

Drives ``run_immersio_phase1`` over a minimal synthetic silver cohort
and verifies all 9 산출 파일 land:

  silver/
    학생지표.parquet
    manifest.json
  gold/
    시험분석결과.xlsx (7 sheets)
    시험품질보고서.md
    시험품질보고서.pdf
    figs/fig1_전체성적_히스토그램.png
    figs/fig2_메타데이터별_정답률.png
    legacy_diff.md
    manifest.json

Plus T067 byte-identical determinism + T068 LLM-call detection in the
same module so a single fixture build serves all three integration
guarantees.
"""

from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import pandas as pd
import pytest

from immersio.analyze.pipeline import PipelineArgs, run_immersio_phase1
from immersio import fonts as _fonts


@pytest.fixture(autouse=True)
def _patch_fonts(monkeypatch: pytest.MonkeyPatch) -> None:
    """Bypass NanumGothic install requirement for e2e tests.

    The pipeline calls ``resolve_korean_font_paths`` for both reportlab
    and matplotlib registration; we substitute DejaVu Sans (bundled with
    matplotlib) so tests run on any CI host.
    """
    import matplotlib
    matplotlib.use("Agg")
    from matplotlib import font_manager

    deja_vu = Path(font_manager.findfont("DejaVu Sans", fallback_to_default=True))
    monkeypatch.setattr(_fonts, "resolve_korean_font_paths", lambda: (deja_vu, deja_vu))


def _seed_silver(silver_dir: Path) -> None:
    """Build a minimal but pipeline-complete silver cohort: 10 학생 × 4 문항."""
    silver_dir.mkdir(parents=True, exist_ok=True)

    items = pd.DataFrame(
        [
            {
                "semester": "2026-1",
                "course_slug": "anatomy",
                "item_no": i,
                "chapter": "1장. 서론" if i <= 2 else "2장. 세포와 조직",
                "week": 1,
                "item_type": "지식축적",
                "difficulty_level": 2,
                "expected_difficulty": "보통",
                "source": "교과서",
                "correct_answer": 1,
                "answer_key": "1",
                "points": 1.0,
                "bloom": "knowledge",
                "text": f"문항 {i}",
            }
            for i in range(1, 5)
        ]
    )
    items.to_parquet(silver_dir / "exam_item.parquet")

    masters = []
    for i in range(1, 11):
        sid = f"202610{i:04d}"
        exam_taken = i <= 7  # 7 응시, 3 결시
        masters.append(
            {
                "student_id": sid,
                "semester": "2026-1",
                "course_slug": "anatomy",
                "on_roster": True,
                "section": "A" if i <= 5 else "B",
                "name_kr": f"학생{i}",
                "diagnostic_responded": True,
                "exam_taken": exam_taken,
                "exam_absent": not exam_taken,
                "attendance_recorded": True,
                "exam_total_score": float(4 - (i - 1) % 4) if exam_taken else None,
                "exam_max_score": 4.0 if exam_taken else None,
                "attendance_present_count": None,
                "attendance_absent_count": None,
                "attendance_late_count": None,
                "attendance_excused_count": None,
                "axis_scores": {"placeholder": 0.0},
            }
        )
    pd.DataFrame(masters).to_parquet(silver_dir / "student_master.parquet")

    rows = []
    correctness_pattern = [
        [True, True, True, True],   # student 1 — 4/4
        [True, True, True, False],  # 2 — 3/4
        [True, True, False, True],  # 3 — 3/4
        [True, True, False, False], # 4 — 2/4
        [True, False, True, False], # 5 — 2/4
        [True, False, False, False],# 6 — 1/4
        [False, False, False, False],# 7 — 0/4
    ]
    for idx, flags in enumerate(correctness_pattern, start=1):
        sid = f"202610{idx:04d}"
        for item_no, ok in enumerate(flags, start=1):
            rows.append(
                {
                    "student_id": sid,
                    "semester": "2026-1",
                    "course_slug": "anatomy",
                    "item_no": item_no,
                    "response": "1" if ok else "2",
                    "is_correct": ok,
                    "is_omit": False,
                }
            )
    pd.DataFrame(rows).to_parquet(silver_dir / "exam_result.parquet")

    diagnostic = pd.DataFrame(
        [
            {
                "student_id": "20261000001",
                "semester": "2026-1",
                "course_slug": "anatomy",
                "axis": "interest_topics",
                "axis_kind": "multiselect_onehot",
                "option_key": "혈액과 면역",
                "value_int": None,
                "value_bool": True,
                "value_text": None,
                "source_column": "Q11",
            }
        ]
    )
    diagnostic.to_parquet(silver_dir / "diagnostic_response.parquet")


@pytest.fixture
def synthetic_silver(tmp_path: Path) -> Path:
    silver_root = tmp_path / "silver"
    silver_dir = silver_root / "immersio" / "2026-1-anatomy"
    _seed_silver(silver_dir)
    return silver_root


@pytest.fixture
def gold_root(tmp_path: Path) -> Path:
    return tmp_path / "gold"


def _make_args(
    *,
    silver_root: Path,
    gold_root: Path,
    legacy_xlsx: Path | None = None,
    no_needs_map: bool = True,
) -> PipelineArgs:
    return PipelineArgs(
        semester="2026-1",
        course_slug="anatomy",
        bronze_dir=silver_root.parent / "bronze",
        silver_root=silver_root,
        gold_root=gold_root,
        legacy_xlsx=legacy_xlsx,
        created_at_utc_override="2026-04-29T00:00:00Z",
        seed=42,
        no_needs_map=no_needs_map,
    )


def test_pipeline_writes_all_gold_artefacts(synthetic_silver: Path, gold_root: Path) -> None:
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    gold_dir = gold_root / "immersio" / "2026-1-anatomy"
    expected = [
        "시험분석결과.xlsx",
        "시험품질보고서.md",
        "시험품질보고서.pdf",
        "legacy_diff.md",
        "manifest.json",
        "figs/fig1_전체성적_히스토그램.png",
        "figs/fig2_메타데이터별_정답률.png",
    ]
    for rel in expected:
        path = gold_dir / rel
        assert path.is_file(), f"missing gold artefact: {rel}"
        assert path.stat().st_size > 0


def test_pipeline_writes_silver_mirrors(synthetic_silver: Path, gold_root: Path) -> None:
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    silver_dir = synthetic_silver / "immersio" / "2026-1-anatomy"
    assert (silver_dir / "학생지표.parquet").is_file()
    assert (silver_dir / "manifest.json").is_file()


def test_pipeline_manifest_has_all_required_fields(
    synthetic_silver: Path, gold_root: Path
) -> None:
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    gold_dir = gold_root / "immersio" / "2026-1-anatomy"
    payload = json.loads((gold_dir / "manifest.json").read_text(encoding="utf-8"))
    required = {
        "schema_version",
        "semester",
        "course_slug",
        "generated_at_utc",
        "exam_item_yaml_sha256",
        "omr_xls_sha256_list",
        "attendance_sha256",
        "run_seed",
        "ruleset_version",
        "total_items",
        "total_responders",
        "total_absent",
        "total_omit_responses",
        "silver_outputs",
        "gold_outputs",
        "legacy_diff_total_cells",
        "legacy_diff_diff_cells",
        "legacy_diff_immersio_chose_count",
        "notes",
    }
    missing = required - set(payload.keys())
    assert not missing, f"manifest missing fields: {missing}"
    # Every sha256 must be a 64-char hex string
    for sha_field in ("exam_item_yaml_sha256", "attendance_sha256"):
        assert len(payload[sha_field]) == 64


def test_pipeline_silver_parquet_round_trip(synthetic_silver: Path, gold_root: Path) -> None:
    """T066 (b) — round-trip of silver 학생지표.parquet."""
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    silver_dir = synthetic_silver / "immersio" / "2026-1-anatomy"
    df = pd.read_parquet(silver_dir / "학생지표.parquet")
    sids = df["student_id"].tolist()
    assert sids == sorted(sids), "학생지표.parquet must be sorted by student_id"
    # 10 students total (7 응시 + 3 결시) — pipeline never drops absent rows
    assert len(df) == 10


def test_pipeline_xlsx_has_seven_sheets(synthetic_silver: Path, gold_root: Path) -> None:
    """T066 (c) — xlsx contains all 7 sheets including 학생성적."""
    from openpyxl import load_workbook

    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    gold_dir = gold_root / "immersio" / "2026-1-anatomy"
    wb = load_workbook(gold_dir / "시험분석결과.xlsx")
    assert wb.sheetnames == [
        "전체요약",
        "1_히스토그램",
        "2_메타데이터통계",
        "3_변별력",
        "4_정답률",
        "5_오답분석",
        "학생성적",
    ]


def test_pipeline_legacy_diff_skipped_when_legacy_absent(
    synthetic_silver: Path, gold_root: Path
) -> None:
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root, legacy_xlsx=None)
    )
    assert rc == 0
    gold_dir = gold_root / "immersio" / "2026-1-anatomy"
    diff_md = (gold_dir / "legacy_diff.md").read_text(encoding="utf-8")
    assert "comparison skipped" in diff_md or "legacy xlsx" in diff_md


# --- T067: byte-equal determinism ----------------------------------------


def _hash_outputs(gold_dir: Path, silver_dir: Path) -> dict[str, str]:
    """sha256 every artefact except manifest.json (manifest carries paths)."""
    out: dict[str, str] = {}
    for path in [
        gold_dir / "시험분석결과.xlsx",
        gold_dir / "시험품질보고서.md",
        gold_dir / "시험품질보고서.pdf",
        gold_dir / "figs/fig1_전체성적_히스토그램.png",
        gold_dir / "figs/fig2_메타데이터별_정답률.png",
        silver_dir / "학생지표.parquet",
    ]:
        out[path.name] = hashlib.sha256(path.read_bytes()).hexdigest()
    return out


def test_two_runs_byte_identical(synthetic_silver: Path, tmp_path: Path) -> None:
    gold_a = tmp_path / "gold_a"
    gold_b = tmp_path / "gold_b"
    silver_b = tmp_path / "silver_b"
    silver_b.mkdir()
    # Seed silver_b with the exact same files
    src_silver = synthetic_silver / "immersio" / "2026-1-anatomy"
    dst_silver = silver_b / "immersio" / "2026-1-anatomy"
    dst_silver.mkdir(parents=True)
    for f in src_silver.iterdir():
        dst_silver.joinpath(f.name).write_bytes(f.read_bytes())

    assert run_immersio_phase1(_make_args(silver_root=synthetic_silver, gold_root=gold_a)) == 0
    assert run_immersio_phase1(_make_args(silver_root=silver_b, gold_root=gold_b)) == 0

    sha_a = _hash_outputs(
        gold_a / "immersio" / "2026-1-anatomy",
        synthetic_silver / "immersio" / "2026-1-anatomy",
    )
    sha_b = _hash_outputs(
        gold_b / "immersio" / "2026-1-anatomy",
        silver_b / "immersio" / "2026-1-anatomy",
    )
    assert sha_a == sha_b, (
        "byte-identical determinism failed for one of the artefacts:\n"
        f"  a: {sha_a}\n  b: {sha_b}"
    )


# --- T068: LLM call detection --------------------------------------------


def test_no_llm_modules_imported_after_pipeline(
    synthetic_silver: Path, gold_root: Path
) -> None:
    """SC-006 — pipeline must NOT import anthropic/openai/instructor."""
    forbidden = ("anthropic", "openai", "instructor")
    pre = {m for m in sys.modules if m.startswith(forbidden)}
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    post = {m for m in sys.modules if m.startswith(forbidden)}
    new_llm = post - pre
    assert not new_llm, f"forbidden LLM module imported during pipeline: {new_llm}"


def test_no_socket_calls_during_pipeline(
    synthetic_silver: Path, gold_root: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """SC-006 — pipeline must NOT open network sockets."""
    import socket

    blocked: list[str] = []
    real_socket = socket.socket

    def _spy(*args, **kwargs):
        # Allow Unix-domain sockets (some test infra uses them); block AF_INET.
        if args and args[0] in (socket.AF_INET, socket.AF_INET6):
            blocked.append(f"socket({args})")
        return real_socket(*args, **kwargs)

    monkeypatch.setattr(socket, "socket", _spy)
    rc = run_immersio_phase1(
        _make_args(silver_root=synthetic_silver, gold_root=gold_root)
    )
    assert rc == 0
    assert not blocked, f"pipeline opened network sockets: {blocked}"
