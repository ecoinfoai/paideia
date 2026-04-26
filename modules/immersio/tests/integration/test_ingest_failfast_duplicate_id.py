"""US2: duplicate student_id in OMR results triggers a violation.

Per contracts/cli.md, post-normalization duplicates are exit code 4
(data integrity), distinct from generic format violations (exit 1).
The aggregator type is ``DataIntegrityError`` (subclass of
``IngestValidationError``).
"""

from __future__ import annotations

from pathlib import Path

import pytest
from immersio.cli.main import app
from immersio.ingest import DataIntegrityError, run_ingest
from openpyxl import load_workbook


def test_duplicate_student_id_in_omr(
    corrupt_bronze: Path, corrupt_mapping: Path, tmp_path: Path
) -> None:
    target = corrupt_bronze / "시험성적" / "인체구조와기능_A반_결과.xlsx"
    workbook = load_workbook(target)
    sheet = workbook["결과"]
    # Force a duplicate student_id by overwriting row 3 with row 2's id.
    duplicate_id = sheet.cell(row=2, column=1).value
    sheet.cell(row=3, column=1).value = duplicate_id
    workbook.save(target)

    out = tmp_path / "silver"
    with pytest.raises(DataIntegrityError) as exc:
        run_ingest(
            bronze_dir=corrupt_bronze,
            mapping_path=corrupt_mapping,
            output_dir=out,
            no_git_commit=True,
        )
    rendered = str(exc.value)
    assert "duplicate student_id" in rendered or str(duplicate_id) in rendered
    assert not (out / "2026-1-anatomy").exists()


def test_duplicate_student_id_cli_returns_exit_code_4(
    corrupt_bronze: Path, corrupt_mapping: Path, tmp_path: Path
) -> None:
    """CLI wraps DataIntegrityError to exit code 4 per contracts/cli.md."""
    target = corrupt_bronze / "시험성적" / "인체구조와기능_A반_결과.xlsx"
    workbook = load_workbook(target)
    sheet = workbook["결과"]
    sheet.cell(row=3, column=1).value = sheet.cell(row=2, column=1).value
    workbook.save(target)

    out = tmp_path / "silver"
    code = app(
        argv=[
            "ingest",
            "--bronze-dir",
            str(corrupt_bronze),
            "--mapping",
            str(corrupt_mapping),
            "--output-dir",
            str(out),
            "--no-git-commit",
        ]
    )
    assert code == 4
    assert not (out / "2026-1-anatomy").exists()

