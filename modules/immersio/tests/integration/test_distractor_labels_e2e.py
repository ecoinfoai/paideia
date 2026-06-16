"""T056 — End-to-end label showcase (Phase 6 US4, FR-019/FR-021, SC-007).

Drives the synthetic 44-item fixture from
``tests.fixtures.build_synthetic_44`` through the report writers
(``md_writer`` + ``xlsx_writer``) and asserts:

  (a) 5_오답분석 시트의 라벨 컬럼에 6 종 라벨 모두 등장
  (b) 변별력 < 0 문항이 시험품질보고서.md 의 해석 단락에 명시 인용
  (c) 5_오답분석 시트의 변별력 < 0 행이 굵게 강조 표기 (Font(bold=True))

This is a *unit-of-integration* test: it skips the silver→bronze→OMR
pipeline (Phase 8 territory) and feeds the report writers directly with
deterministic ItemStatistics instances. The contract being verified is
the wiring between US1's labelling rule output (T029/T039) and US4's
operator-facing surfaces (T058/T059).
"""

from __future__ import annotations

import sys
from pathlib import Path
from pathlib import Path as _Path

import pytest
from immersio.report.md_writer import render_quality_report_md
from immersio.report.xlsx_writer import write_analysis_xlsx
from openpyxl import load_workbook
from paideia_shared.schemas import HistogramBin, MetadataAggregate

sys.path.insert(0, str(_Path(__file__).resolve().parent.parent / "fixtures"))
from build_synthetic_44 import build_label_showcase_items  # noqa: E402

EXPECTED_LABELS = (
    "역변별 의심 — 출제 재검토",
    "모두 풀 수 있는 기본 문항",
    "어려운 변별 우수 문항(유지 권장)",
    "시간 부족 또는 포기형",
    "근접 distractor에 의한 변별 성공형",
    "변별 기여 적음 — 차년도 교체 검토",
)


def _stub_overall(n_responders: int = 100) -> list[dict[str, object]]:
    return [
        {"지표": "응시자 수", "값": n_responders},
        {"지표": "결시자 수", "값": 0},
        {"지표": "무응답 응답 수", "값": 0},
        {"지표": "만점", "값": 100.0},
        {"지표": "평균", "값": 65.0},
        {"지표": "표준편차", "값": 15.0},
        {"지표": "중앙값", "값": 65.0},
        {"지표": "최저", "값": 30.0},
        {"지표": "최고", "값": 95.0},
        {"지표": "Q1", "값": 50.0},
        {"지표": "Q3", "값": 80.0},
        {"지표": "100점환산_평균", "값": 65.0},
        {"지표": "100점환산_표준편차", "값": 15.0},
    ]


def _stub_histogram() -> list[HistogramBin]:
    return [
        HistogramBin(bin_start=0.0, bin_end=10.0, count=0, cumulative=0, cumulative_pct=0.0),
        HistogramBin(
            bin_start=90.0, bin_end=100.0, count=100, cumulative=100, cumulative_pct=100.0
        ),
    ]


def _stub_metadata() -> list[MetadataAggregate]:
    return [
        MetadataAggregate(
            metadata_kind="분반",
            metadata_value="A",
            n=100,
            mean=65.0,
            sd=15.0,
            test_kind="N/A",
            test_p_value=None,
            levene_p_value=None,
            note=None,
        ),
    ]


@pytest.fixture
def items():
    return build_label_showcase_items()


def test_six_labels_all_present_in_5_오답분석_sheet(items, tmp_path: Path) -> None:
    out = tmp_path / "result.xlsx"
    write_analysis_xlsx(
        output_path=out,
        overall_rows=_stub_overall(),
        histogram_bins=_stub_histogram(),
        metadata_rows=_stub_metadata(),
        item_stats=items,
        semester="2026-1",
        course_name_kr="인체구조와기능",
        generated_at_utc="2026-04-29T00:00:00Z",
    )
    wb = load_workbook(out)
    ws = wb["5_오답분석"]

    label_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == "distractor_label":
            label_col = c
            break
    assert label_col is not None, "distractor_label column not found"

    labels_in_sheet = {
        ws.cell(r, label_col).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, label_col).value is not None
    }
    for expected in EXPECTED_LABELS:
        assert expected in labels_in_sheet, (
            f"label {expected!r} missing from 5_오답분석 sheet; found: {labels_in_sheet}"
        )


def test_negative_discrimination_row_is_bold(items, tmp_path: Path) -> None:
    out = tmp_path / "result.xlsx"
    write_analysis_xlsx(
        output_path=out,
        overall_rows=_stub_overall(),
        histogram_bins=_stub_histogram(),
        metadata_rows=_stub_metadata(),
        item_stats=items,
        semester="2026-1",
        course_name_kr="인체구조와기능",
        generated_at_utc="2026-04-29T00:00:00Z",
    )
    wb = load_workbook(out)
    ws = wb["5_오답분석"]

    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == 1:
            target_row = r
            break
    assert target_row is not None, "item_no=1 row not found"
    for c in range(1, 9):
        font = ws.cell(target_row, c).font
        assert font.bold is True, f"5_오답분석 row {target_row} col {c} not bold (item_no=1, D<0)"


def test_md_writer_quotes_negative_discrimination_items(items) -> None:
    md = render_quality_report_md(
        overall_rows=_stub_overall(),
        histogram_bins=_stub_histogram(),
        metadata_rows=_stub_metadata(),
        item_stats=items,
        semester="2026-1",
        course_name_kr="인체구조와기능",
        generated_at_utc="2026-04-29T00:00:00Z",
    )
    section3 = md.split("## (3)")[1].split("## (4)")[0]
    assert "1" in section3, "negative-D item_no=1 not cited in §(3)"
    # Per FR-021 + dispatch: explicit (해당 문항: ...) format
    assert "해당 문항" in section3, "FR-021 template requires '(해당 문항: ...)' phrasing in §(3)"


def test_md_writer_label_buckets_list_each_label(items) -> None:
    md = render_quality_report_md(
        overall_rows=_stub_overall(),
        histogram_bins=_stub_histogram(),
        metadata_rows=_stub_metadata(),
        item_stats=items,
        semester="2026-1",
        course_name_kr="인체구조와기능",
        generated_at_utc="2026-04-29T00:00:00Z",
    )
    section5 = md.split("## (5)")[1].split("## (6)")[0]
    for label in EXPECTED_LABELS:
        assert label in section5, f"label {label!r} not surfaced in §(5)"


def test_xlsx_two_writes_byte_identical_with_full_label_cohort(items, tmp_path: Path) -> None:
    import hashlib

    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    for path in (a, b):
        write_analysis_xlsx(
            output_path=path,
            overall_rows=_stub_overall(),
            histogram_bins=_stub_histogram(),
            metadata_rows=_stub_metadata(),
            item_stats=items,
            semester="2026-1",
            course_name_kr="인체구조와기능",
            generated_at_utc="2026-04-29T00:00:00Z",
        )
    sha_a = hashlib.sha256(a.read_bytes()).hexdigest()
    sha_b = hashlib.sha256(b.read_bytes()).hexdigest()
    assert sha_a == sha_b
