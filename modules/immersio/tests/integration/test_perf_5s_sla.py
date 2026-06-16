"""SC-001: a 184-student ingest must complete within 5 seconds.

The synthetic fixture is materialized lazily into tmp_path so the binary
blobs stay out of git.
"""

from __future__ import annotations

import time
from pathlib import Path

import pytest
import yaml
from immersio.ingest import run_ingest
from openpyxl import Workbook

ITEMS = [
    {
        "item_no": item_no,
        "answer_key": str(((item_no - 1) % 5) + 1),
        "chapter": f"{(item_no - 1) // 10 + 1}장",
        "source": ("textbook" if item_no % 3 == 0 else "formative"),
        "expected_difficulty": "medium",
        "bloom": "knowledge",
        "points": 2.0,
        "text": f"문항 {item_no}",
    }
    for item_no in range(1, 51)  # 50 items
]


def _make_diag_csv(out: Path, n: int) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    headers = "학번,Q01,Q02,Q05,Q11,Q62"
    rows = [headers]
    for i in range(n):
        sid = f"2026{i:06d}"
        rows.append(f"{sid},매우 그렇다,그렇다,약간 그렇다,신경계;근육계,자유서술 텍스트")
    out.write_text("\n".join(rows), encoding="utf-8")


def _make_omr(dir_path: Path, students_by_section: dict[str, list[str]]) -> None:
    dir_path.mkdir(parents=True, exist_ok=True)
    for section, students in students_by_section.items():
        wb = Workbook()
        wb.remove(wb.active)
        results = wb.create_sheet("결과")
        results.append(["학번", "이름"] + [f"item_{item['item_no']}" for item in ITEMS] + ["점수"])
        for sid in students:
            answers = [item["answer_key"] for item in ITEMS]
            score = sum(float(item["points"]) for item in ITEMS)
            results.append([sid, "이름"] + answers + [score])
        wb.create_sheet("결시").append(["학번", "이름"])
        ox = wb.create_sheet("OX")
        ox.append(["학번"] + [f"item_{item['item_no']}_OX" for item in ITEMS])
        for sid in students:
            ox.append([sid] + ["O"] * len(ITEMS))
        analysis = wb.create_sheet("문항분석")
        analysis.append(
            [
                "item_no",
                "chapter",
                "source",
                "expected_difficulty",
                "bloom",
                "answer_key",
                "points",
                "text",
            ]
        )
        for item in ITEMS:
            analysis.append(
                [
                    item["item_no"],
                    item["chapter"],
                    item["source"],
                    item["expected_difficulty"],
                    item["bloom"],
                    item["answer_key"],
                    item["points"],
                    item["text"],
                ]
            )
        wb.save(dir_path / f"인체구조와기능_{section}반_결과.xlsx")


def _make_attendance(out: Path, ids: list[str]) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "출석"
    sheet.append(["학번", "이름"] + [f"W{week:02d}" for week in range(1, 17)] + ["비고"])
    for sid in ids:
        sheet.append([sid, "이름"] + ["O"] * 16 + [""])
    wb.save(out)


def _make_yaml(out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(yaml.safe_dump({"items": ITEMS}, allow_unicode=True), encoding="utf-8")


def _make_mapping(out: Path) -> None:
    out.write_text(
        """
metadata:
  semester: "2026-1"
  course_slug: anatomy
  course_name_kr: "인체구조와기능"
  mapping_version: 1
columns:
  - source: "학번"
    kind: identity
  - source: "Q01"
    axis: motivation
    kind: likert
    aggregate: mean
  - source: "Q02"
    axis: motivation
    kind: likert
    aggregate: mean
  - source: "Q05"
    axis: anxiety
    kind: likert
  - source: "Q11"
    axis: interest
    kind: multiselect
  - source: "Q62"
    axis: anxiety
    kind: freetext
axes:
  required:
    - motivation
    - anxiety
    - interest
  optional:
    - anxiety
""".lstrip(),
        encoding="utf-8",
    )


@pytest.fixture
def synthetic_184_bronze(tmp_path: Path) -> tuple[Path, Path]:
    """Materialize a synthetic 184-student Bronze tree."""
    bronze = tmp_path / "bronze"
    n = 184
    ids_by_section: dict[str, list[str]] = {"A": [], "B": [], "C": [], "D": []}
    for i in range(n):
        sid = f"2026{i:06d}"
        ids_by_section[("A", "B", "C", "D")[i % 4]].append(sid)

    all_ids = [sid for ids in ids_by_section.values() for sid in ids]
    _make_diag_csv(bronze / "진단평가" / "diag.csv", n)
    _make_omr(bronze / "시험성적", ids_by_section)
    _make_attendance(bronze / "출석" / "출석부.xlsx", all_ids)
    _make_yaml(bronze / "시험문제" / "exam.yaml")

    mapping = tmp_path / "anatomy.yaml"
    _make_mapping(mapping)
    return bronze, mapping


def test_184_student_ingest_under_5s(
    synthetic_184_bronze: tuple[Path, Path], tmp_path: Path
) -> None:
    bronze, mapping = synthetic_184_bronze
    out = tmp_path / "silver"
    start = time.perf_counter()
    run_ingest(
        bronze_dir=bronze,
        mapping_path=mapping,
        output_dir=out,
        no_git_commit=True,
    )
    elapsed = time.perf_counter() - start
    assert elapsed < 5.0, f"184-student ingest exceeded SC-001 SLA: {elapsed:.2f}s"
