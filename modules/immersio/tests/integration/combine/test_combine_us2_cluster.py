"""Integration test — US2 cluster pipeline e2e (T044).

Drives ``run_us1_pipeline(include_cluster=True)`` on:
- silver_phase3_minimal (k=3 well-separated clusters)
- silver_phase3_no_clusters (k=1 fallback)

Asserts that:
- US2 wiring 산출 9 파일 (silver 2 + gold md/pdf/xlsx + figs 3) land
- xlsx 에 `군집비교` sheet 추가
- manifest.posthoc_method_used 가 cluster_compare 결과와 정합
- §4 군집별 비교 가 markdown 에 active section 으로 등장
"""

from __future__ import annotations

import importlib.util
import json
from pathlib import Path
from types import ModuleType

import openpyxl
import pyarrow.parquet as pq
import pytest


def _load_builder() -> ModuleType:
    here = Path(__file__).resolve()
    builder_path = here.parents[2] / "fixtures" / "build_silver_phase3.py"
    spec = importlib.util.spec_from_file_location(
        "build_silver_phase3", builder_path
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"could not load builder from {builder_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ----------------------------------------------------------------------
# k=3 happy path
# ----------------------------------------------------------------------


def test_us2_k3_pipeline_lands_three_sheets(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us2_k3")
    builder.build_silver_phase3_minimal(tmp)

    rc = run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
    )
    assert rc == 0

    xlsx_path = tmp / "gold" / "immersio" / "2026-1-anatomy" / "결합분석.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    assert wb.sheetnames == ["상관매트릭스", "회귀결과", "군집비교"]


def test_us2_k3_emits_fig5(tmp_path_factory: pytest.TempPathFactory) -> None:
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us2_k3_fig5")
    builder.build_silver_phase3_minimal(tmp)
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
    )
    fig5 = (
        tmp
        / "gold"
        / "immersio"
        / "2026-1-anatomy"
        / "figs"
        / "fig5_cluster_boxplot.png"
    )
    assert fig5.exists()


def test_us2_k3_manifest_posthoc_set(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """k=3 well-separated → posthoc_method_used != 'N/A'."""
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us2_k3_manifest")
    builder.build_silver_phase3_minimal(tmp)
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
    )
    manifest_path = (
        tmp
        / "silver"
        / "immersio"
        / "2026-1-anatomy"
        / "manifest_phase3.json"
    )
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["posthoc_method_used"] in {"Tukey_HSD", "Games_Howell"}


def test_us2_k3_md_contains_section_4_active(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """§4 군집별 비교 가 placeholder 가 아닌 active section 으로 등장."""
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us2_k3_md")
    builder.build_silver_phase3_minimal(tmp)
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
    )
    md_path = (
        tmp / "gold" / "immersio" / "2026-1-anatomy" / "결합분석보고서.md"
    )
    text = md_path.read_text(encoding="utf-8")
    assert "## 4. 군집별 비교" in text
    assert "(US2 미수행)" not in text
    assert "fig5_cluster_boxplot.png" in text


# ----------------------------------------------------------------------
# k=1 fallback
# ----------------------------------------------------------------------


def test_us2_k1_fallback_yields_na_posthoc(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """k=1 fallback → cluster_compare emits N/A test + posthoc."""
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us2_k1")
    builder.build_silver_phase3_no_clusters(tmp)
    rc = run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
    )
    assert rc == 0

    manifest_path = (
        tmp
        / "silver"
        / "immersio"
        / "2026-1-anatomy"
        / "manifest_phase3.json"
    )
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["posthoc_method_used"] == "N/A"


def test_us2_k1_fallback_no_pairwise_rows(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """k=1 → xlsx 군집비교 sheet 의 사후 비교 block 이 비어 있음."""
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us2_k1_xlsx")
    builder.build_silver_phase3_no_clusters(tmp)
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
    )
    xlsx_path = tmp / "gold" / "immersio" / "2026-1-anatomy" / "결합분석.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheet = wb["군집비교"]
    pair_rows = [
        row
        for row in sheet.iter_rows(values_only=True)
        if row and isinstance(row[0], str) and "-" in row[0]
        and row[0] not in {"raw_p", "ci_low_95", "ci_high_95"}
    ]
    # k=1 → 0 pairwise. (Header rows like '군집별 통계' / '검정 결과' / '사후 비교'
    # excluded by string check.)
    assert pair_rows == []


# ----------------------------------------------------------------------
# Backward-compat: include_cluster=False emits 2 sheets (Phase 4 path)
# ----------------------------------------------------------------------


def test_us1_partial_default_emits_two_sheets(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    tmp = tmp_path_factory.mktemp("us1_default")
    builder.build_silver_phase3_minimal(tmp)
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        # include_cluster default False
    )
    xlsx_path = tmp / "gold" / "immersio" / "2026-1-anatomy" / "결합분석.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    assert wb.sheetnames == ["상관매트릭스", "회귀결과"]
    fig5 = (
        tmp
        / "gold"
        / "immersio"
        / "2026-1-anatomy"
        / "figs"
        / "fig5_cluster_boxplot.png"
    )
    assert not fig5.exists(), "fig5 must not land when include_cluster=False"


# ----------------------------------------------------------------------
# Silver byte-identical re-run (vector composition retained in US2 mode)
# ----------------------------------------------------------------------


def test_us2_silver_byte_identical_re_run(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    from immersio.combine.pipeline import run_us1_pipeline

    builder = _load_builder()
    a = tmp_path_factory.mktemp("us2_byte_a")
    b = tmp_path_factory.mktemp("us2_byte_b")
    builder.build_silver_phase3_minimal(a)
    builder.build_silver_phase3_minimal(b)
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=a / "silver",
        gold_dir=a / "gold",
        include_cluster=True,
    )
    run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=b / "silver",
        gold_dir=b / "gold",
        include_cluster=True,
    )
    pa = (
        a / "silver" / "immersio" / "2026-1-anatomy" / "진단×시험결합.parquet"
    ).read_bytes()
    pb = (
        b / "silver" / "immersio" / "2026-1-anatomy" / "진단×시험결합.parquet"
    ).read_bytes()
    assert pa == pb
