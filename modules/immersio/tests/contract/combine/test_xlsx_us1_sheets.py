"""Contract test — xlsx US1 2 sheets schema (T035, US1).

본 contract 는 ``combine.xlsx_writer.write_us1_xlsx`` 출력이
``contracts/xlsx_workbook.md`` 의 시트 이름 + 헤더 스키마와 1:1 정합
함을 담보. T032 unit test 가 함수 boundary 검증 (헤더 row + row count
+ byte-identical re-run); T035 는 spec contract 차원 게이트.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from immersio.combine.xlsx_writer import write_us1_xlsx
from paideia_shared.schemas import (
    CorrelationCell,
    RegressionCoefficient,
    RegressionFitSummary,
)
from paideia_shared.schemas._common import STANDARD_AXIS_KEYS


def _cells() -> list[CorrelationCell]:
    return [
        CorrelationCell(
            axis_key=axis,
            exam_metric_key=metric,
            n=22,
            pearson_r=0.3,
            raw_p=0.01,
            fdr_q=0.04,
            significant_after_correction=True,
            unstable_inference_flag=False,
        )
        for axis in STANDARD_AXIS_KEYS
        for metric in ("total_score", "chapter_신경계")
    ]


def _coefs() -> list[RegressionCoefficient]:
    out: list[RegressionCoefficient] = []
    for i, axis in enumerate(STANDARD_AXIS_KEYS):
        coef = 0.5 + i * 0.1
        out.append(
            RegressionCoefficient(
                axis_key=axis,
                coef=coef,
                std_err=0.2,
                t_stat=2.5,
                raw_p=0.01,
                fdr_q=0.04,
                ci_low_95=coef - 0.5,
                ci_high_95=coef + 0.5,
                beta_standardized=0.1 + i * 0.02,
                vif=1.5,
                multicollinearity_flag=False,
            )
        )
    return out


def _fit() -> RegressionFitSummary:
    return RegressionFitSummary(
        n_complete_case=22,
        n_dropped=8,
        r2=0.45,
        r2_adj=0.32,
        f_stat=5.0,
        f_pvalue=0.001,
        regression_method="OLS",
        multiple_comparison_method="BH-FDR",
        small_sample_warning=True,
    )


@pytest.fixture(scope="module")
def workbook(tmp_path_factory: pytest.TempPathFactory) -> openpyxl.Workbook:
    out = tmp_path_factory.mktemp("xlsx_contract") / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    return openpyxl.load_workbook(out, read_only=True)


# ----------------------------------------------------------------------
# Sheet inventory + ordering
# ----------------------------------------------------------------------


def test_us1_sheet_inventory_matches_contract(
    workbook: openpyxl.Workbook,
) -> None:
    """contracts/xlsx_workbook.md US1 partial: 정확히 2 시트.

    US2 (`군집비교`) + US4 (`부분군비교`) 는 후속 phase 가 추가.
    """
    expected = ["상관매트릭스", "회귀결과"]
    assert list(workbook.sheetnames) == expected


# ----------------------------------------------------------------------
# Sheet 1 — 상관매트릭스
# ----------------------------------------------------------------------


def test_correlation_sheet_header_schema(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["상관매트릭스"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    expected = [
        "axis_key",
        "exam_metric_key",
        "n",
        "pearson_r",
        "raw_p",
        "fdr_q",
        "significant",
        "unstable_n_lt_20",
    ]
    assert headers == expected


def test_correlation_first_data_row_axis_is_standard_key(
    workbook: openpyxl.Workbook,
) -> None:
    sheet = workbook["상관매트릭스"]
    rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    first_axis = rows[0][0]
    assert first_axis in STANDARD_AXIS_KEYS


# ----------------------------------------------------------------------
# Sheet 2 — 회귀결과
# ----------------------------------------------------------------------


def test_regression_sheet_starts_with_fit_block_label(
    workbook: openpyxl.Workbook,
) -> None:
    sheet = workbook["회귀결과"]
    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    assert first_row[0] == "적합 지표"


def test_regression_sheet_fit_field_row_present(
    workbook: openpyxl.Workbook,
) -> None:
    """row 2 = field labels of the fit summary."""
    sheet = workbook["회귀결과"]
    row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert "n_complete_case" in row
    assert "R²" in row
    assert "regression_method" in row


def test_regression_sheet_coef_header_row_present(
    workbook: openpyxl.Workbook,
) -> None:
    sheet = workbook["회귀결과"]
    expected_coef_headers = [
        "axis_key",
        "coef",
        "std_err",
        "t_stat",
        "raw_p",
        "fdr_q",
        "ci_low_95",
        "ci_high_95",
        "beta_standardized",
        "vif",
        "multicollinearity_flag",
    ]
    found = False
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "axis_key":
            assert list(row[: len(expected_coef_headers)]) == expected_coef_headers
            found = True
            break
    assert found, "회귀결과 sheet has no 'axis_key' header row"


def test_regression_sheet_emits_8_axis_rows(
    workbook: openpyxl.Workbook,
) -> None:
    sheet = workbook["회귀결과"]
    axes_seen: set[str] = set()
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] in set(STANDARD_AXIS_KEYS):
            axes_seen.add(row[0])
    assert axes_seen == set(STANDARD_AXIS_KEYS)


# ----------------------------------------------------------------------
# Determinism — vector #5 dcterms epoch
# ----------------------------------------------------------------------


def test_dcterms_modified_pinned(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """xlsx 의 docProps/core.xml `<dcterms:modified>` 값이 epoch 2000-01-01 으로 pin."""
    import zipfile

    out = tmp_path_factory.mktemp("xlsx_dcterms") / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    with zipfile.ZipFile(out, "r") as zf:
        core_xml = zf.read("docProps/core.xml").decode("utf-8")
    assert "2000-01-01T00:00:00Z" in core_xml, (
        f"<dcterms:modified> not pinned to epoch (got core.xml excerpt: "
        f"{core_xml[:500]})"
    )
