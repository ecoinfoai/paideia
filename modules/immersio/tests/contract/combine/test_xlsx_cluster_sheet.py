"""Contract test — xlsx `군집비교` sheet 3-block schema (T043, US2).

Verifies that ``write_us1_xlsx`` with ``cluster_rows/header/pairwise``
adds the `군집비교` sheet whose 3-block structure matches
``contracts/xlsx_workbook.md`` §Sheet 3.
"""

from __future__ import annotations

import openpyxl
import pytest
from immersio.combine.xlsx_writer import write_us1_xlsx
from paideia_shared.schemas import (
    CorrelationCell,
    RegressionCoefficient,
    RegressionFitSummary,
)
from paideia_shared.schemas._common import STANDARD_AXIS_KEYS
from paideia_shared.schemas.cluster_score_comparison import (
    ClusterPairwise,
    ClusterRow,
    ClusterScoreComparison,
)


def _cells() -> list[CorrelationCell]:
    return [
        CorrelationCell(
            axis_key=axis,
            exam_metric_key="total_score",
            n=22,
            pearson_r=0.3,
            raw_p=0.01,
            fdr_q=0.04,
            significant_after_correction=True,
            unstable_inference_flag=False,
        )
        for axis in STANDARD_AXIS_KEYS
    ]


def _coefs() -> list[RegressionCoefficient]:
    return [
        RegressionCoefficient(
            axis_key=axis,
            coef=0.5,
            std_err=0.2,
            t_stat=2.5,
            raw_p=0.01,
            fdr_q=0.04,
            ci_low_95=0.0,
            ci_high_95=1.0,
            beta_standardized=0.1,
            vif=1.5,
            multicollinearity_flag=False,
        )
        for axis in STANDARD_AXIS_KEYS
    ]


def _fit() -> RegressionFitSummary:
    return RegressionFitSummary(
        n_complete_case=22,
        n_dropped=8,
        r2=0.45,
        r2_adj=0.32,
        f_stat=5.0,
        f_pvalue=0.001,
        regression_method="OLS",
        multiple_comparison_method="BH-FDR",
        small_sample_warning=True,
    )


def _cluster_payload() -> tuple:
    rows = [
        ClusterRow(
            cluster_id=0,
            cluster_label="A",
            n=10,
            mean=60.0,
            std=5.0,
            ci_low_95=56.0,
            ci_high_95=64.0,
        ),
        ClusterRow(
            cluster_id=1,
            cluster_label="B",
            n=12,
            mean=75.0,
            std=4.0,
            ci_low_95=72.0,
            ci_high_95=78.0,
        ),
        ClusterRow(
            cluster_id=2,
            cluster_label="C",
            n=15,
            mean=85.0,
            std=6.0,
            ci_low_95=82.0,
            ci_high_95=88.0,
        ),
    ]
    header = ClusterScoreComparison(
        k_used=3,
        test_used="ANOVA",
        levene_p=0.42,
        test_stat=85.0,
        raw_p=0.0001,
        eta_squared=0.65,
        omega_squared=None,
        posthoc_test="Tukey_HSD",
    )
    pairwise = [
        ClusterPairwise(
            cluster_pair=(0, 1),
            mean_diff=-15.0,
            raw_p=0.001,
            fdr_q=0.001,
            significant_after_correction=True,
        ),
        ClusterPairwise(
            cluster_pair=(0, 2),
            mean_diff=-25.0,
            raw_p=0.0001,
            fdr_q=0.0001,
            significant_after_correction=True,
        ),
        ClusterPairwise(
            cluster_pair=(1, 2),
            mean_diff=-10.0,
            raw_p=0.005,
            fdr_q=0.005,
            significant_after_correction=True,
        ),
    ]
    return rows, header, pairwise


@pytest.fixture(scope="module")
def workbook(tmp_path_factory: pytest.TempPathFactory) -> openpyxl.Workbook:
    rows, header, pairwise = _cluster_payload()
    out = tmp_path_factory.mktemp("xlsx_cluster_contract") / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
        cluster_rows=rows,
        cluster_header=header,
        cluster_pairwise=pairwise,
    )
    return openpyxl.load_workbook(out, read_only=True)


def test_three_sheets_in_us2_mode(workbook: openpyxl.Workbook) -> None:
    """US2 mode: `상관매트릭스` + `회귀결과` + `군집비교` (정확히 3)."""
    assert workbook.sheetnames == ["상관매트릭스", "회귀결과", "군집비교"]


def test_block1_label_and_headers(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["군집비교"]
    rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
    assert rows[0][0] == "군집별 통계"
    assert list(rows[1]) == [
        "cluster_id",
        "cluster_label",
        "n",
        "mean",
        "std",
        "ci_low_95",
        "ci_high_95",
        "excluded_reason",
    ]


def test_block1_three_cluster_rows(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["군집비교"]
    cluster_ids: list[int] = []
    for row in sheet.iter_rows(values_only=True):
        if row and isinstance(row[0], int) and row[0] in {0, 1, 2}:
            cluster_ids.append(row[0])
    assert cluster_ids == [0, 1, 2]


def test_block2_omnibus_header_present(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["군집비교"]
    seen = False
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "검정 결과":
            seen = True
            break
    assert seen, "군집비교 sheet missing '검정 결과' block label"


def test_block2_field_row(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["군집비교"]
    found = False
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "k_used":
            assert list(row) == [
                "k_used",
                "test_used",
                "levene_p",
                "test_stat",
                "raw_p",
                "eta_squared",
                "omega_squared",
                "posthoc_test",
            ]
            found = True
            break
    assert found


def test_block3_pairwise_block(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["군집비교"]
    pair_rows: list[str] = []
    in_pairwise_block = False
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "사후 비교":
            in_pairwise_block = True
            continue
        if in_pairwise_block and row and isinstance(row[0], str) and "-" in row[0]:
            pair_rows.append(row[0])
    assert pair_rows == ["0-1", "0-2", "1-2"]


def test_us1_mode_omits_cluster_sheet(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """Backward-compat: US1 mode (no cluster_* args) still emits 2 sheets."""
    out = tmp_path_factory.mktemp("xlsx_us1_only") / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
    )
    wb = openpyxl.load_workbook(out, read_only=True)
    assert wb.sheetnames == ["상관매트릭스", "회귀결과"]
