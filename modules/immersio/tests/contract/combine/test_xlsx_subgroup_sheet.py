"""Contract test — xlsx `부분군비교` sheet 4-meta sub-block (T055a, US4).

Verifies write_us1_xlsx with subgroup_* args adds the 4 sub-blocks
matching ``contracts/xlsx_workbook.md`` §Sheet 4.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from immersio.combine.xlsx_writer import write_us1_xlsx
from paideia_shared.schemas import (
    CorrelationCell,
    RegressionCoefficient,
    RegressionFitSummary,
)
from paideia_shared.schemas._common import STANDARD_AXIS_KEYS
from paideia_shared.schemas.subgroup_score_comparison import (
    SubgroupRow,
    SubgroupScoreComparison,
)


def _cells() -> list[CorrelationCell]:
    return [
        CorrelationCell(
            axis_key=axis,
            exam_metric_key="total_score",
            n=22,
            pearson_r=0.3,
            raw_p=0.01,
            fdr_q=0.04,
            significant_after_correction=True,
            unstable_inference_flag=False,
        )
        for axis in STANDARD_AXIS_KEYS
    ]


def _coefs() -> list[RegressionCoefficient]:
    return [
        RegressionCoefficient(
            axis_key=axis,
            coef=0.5,
            std_err=0.2,
            t_stat=2.5,
            raw_p=0.01,
            fdr_q=0.04,
            ci_low_95=0.0,
            ci_high_95=1.0,
            beta_standardized=0.1,
            vif=1.5,
            multicollinearity_flag=False,
        )
        for axis in STANDARD_AXIS_KEYS
    ]


def _fit() -> RegressionFitSummary:
    return RegressionFitSummary(
        n_complete_case=22,
        n_dropped=8,
        r2=0.45,
        r2_adj=0.32,
        f_stat=5.0,
        f_pvalue=0.001,
        regression_method="OLS",
        multiple_comparison_method="BH-FDR",
        small_sample_warning=True,
    )


def _subgroup_payload() -> tuple:
    rows = [
        SubgroupRow(meta_kind="section", meta_value="A", n=15, mean=70.0, std=5.0),
        SubgroupRow(meta_kind="section", meta_value="B", n=20, mean=75.0, std=4.0),
        SubgroupRow(
            meta_kind="prior_biology",
            meta_value="(메타 미정의)",
            n=0,
            excluded_reason="R-10 매핑 미정의",
        ),
        SubgroupRow(
            meta_kind="occupation",
            meta_value="(메타 미정의)",
            n=0,
            excluded_reason="R-10 매핑 미정의",
        ),
        SubgroupRow(
            meta_kind="education",
            meta_value="(메타 미정의)",
            n=0,
            excluded_reason="R-10 매핑 미정의",
        ),
    ]
    headers = [
        SubgroupScoreComparison(
            meta_kind="section",
            test_used="t_test_welch",
            levene_p=None,
            test_stat=None,
            raw_p=0.05,
            fdr_q=0.05,
            effect_size_kind="cohen_d",
            effect_size_value=-0.3,
            n_categories_compared=2,
        ),
        SubgroupScoreComparison(
            meta_kind="prior_biology",
            test_used="N/A",
            levene_p=None,
            test_stat=None,
            raw_p=None,
            fdr_q=None,
            effect_size_kind="cohen_d",
            effect_size_value=None,
            n_categories_compared=0,
        ),
        SubgroupScoreComparison(
            meta_kind="occupation",
            test_used="N/A",
            levene_p=None,
            test_stat=None,
            raw_p=None,
            fdr_q=None,
            effect_size_kind="cohen_d",
            effect_size_value=None,
            n_categories_compared=0,
        ),
        SubgroupScoreComparison(
            meta_kind="education",
            test_used="N/A",
            levene_p=None,
            test_stat=None,
            raw_p=None,
            fdr_q=None,
            effect_size_kind="cohen_d",
            effect_size_value=None,
            n_categories_compared=0,
        ),
    ]
    return rows, headers


@pytest.fixture(scope="module")
def workbook(tmp_path_factory: pytest.TempPathFactory) -> openpyxl.Workbook:
    rows, headers = _subgroup_payload()
    out = tmp_path_factory.mktemp("xlsx_subgroup_contract") / "결합분석.xlsx"
    write_us1_xlsx(
        correlation_cells=_cells(),
        regression_coefs=_coefs(),
        regression_fit=_fit(),
        out_path=out,
        subgroup_rows=rows,
        subgroup_headers=headers,
    )
    return openpyxl.load_workbook(out, read_only=True)


def test_subgroup_sheet_present_us4_mode(
    workbook: openpyxl.Workbook,
) -> None:
    """US4 mode 시 정확히 [상관매트릭스, 회귀결과, 부분군비교] 3 sheets."""
    assert workbook.sheetnames == ["상관매트릭스", "회귀결과", "부분군비교"]


def test_four_meta_sub_blocks_present(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["부분군비교"]
    meta_labels: list[str] = []
    for row in sheet.iter_rows(values_only=True):
        if row and isinstance(row[0], str) and row[0].startswith("메타: "):
            meta_labels.append(row[0])
    assert meta_labels == [
        "메타: section",
        "메타: prior_biology",
        "메타: occupation",
        "메타: education",
    ]


def test_section_block_emits_two_categories(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["부분군비교"]
    in_section = False
    cat_rows: list[str] = []
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "메타: section":
            in_section = True
            continue
        if in_section and row and row[0] == "메타: prior_biology":
            break
        if in_section and row and row[0] in {"A", "B"}:
            cat_rows.append(row[0])
    assert sorted(cat_rows) == ["A", "B"]


def test_undefined_meta_emits_placeholder(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["부분군비교"]
    placeholder_count = 0
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "(메타 미정의)":
            placeholder_count += 1
    assert placeholder_count == 3, (
        f"expected 3 (메타 미정의) rows for prior_biology/occupation/education, "
        f"got {placeholder_count}"
    )


def test_test_used_header_present(workbook: openpyxl.Workbook) -> None:
    sheet = workbook["부분군비교"]
    seen_test_used = False
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "test_used":
            seen_test_used = True
            break
    assert seen_test_used


# ----------------------------------------------------------------------
# T056a — first e2e 9-artifact inventory (full pipeline)
# ----------------------------------------------------------------------


def test_full_pipeline_emits_all_artefacts(
    tmp_path_factory: pytest.TempPathFactory,
) -> None:
    """T056a — silver_phase3_minimal e2e with cluster + subgroup → 11 산출
    (silver 2 + gold md/pdf/xlsx + 4 figs + ... )."""
    import importlib.util
    from types import ModuleType

    here = Path(__file__).resolve()
    builder_path = here.parents[2] / "fixtures" / "build_silver_phase3.py"
    spec = importlib.util.spec_from_file_location("build_silver_phase3", builder_path)
    if spec is None or spec.loader is None:
        pytest.skip("builder unavailable")
    module: ModuleType = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    from immersio.combine.pipeline import run_us1_pipeline

    tmp = tmp_path_factory.mktemp("us4_full_e2e")
    module.build_silver_phase3_minimal(tmp)
    rc = run_us1_pipeline(
        semester="2026-1",
        course_slug="anatomy",
        silver_dir=tmp / "silver",
        gold_dir=tmp / "gold",
        include_cluster=True,
        include_subgroup=True,
    )
    assert rc == 0
    target = tmp / "gold" / "immersio" / "2026-1-anatomy"
    silver_target = tmp / "silver" / "immersio" / "2026-1-anatomy"
    assert (silver_target / "진단×시험결합.parquet").exists()
    assert (silver_target / "manifest_phase3.json").exists()
    assert (target / "결합분석보고서.md").exists()
    assert (target / "결합분석보고서.pdf").exists()
    assert (target / "결합분석.xlsx").exists()
    figs = target / "figs"
    assert (figs / "fig3_corr_heatmap.png").exists()
    assert (figs / "fig4_beta_bar.png").exists()
    assert (figs / "fig5_cluster_boxplot.png").exists()
    assert (figs / "fig6_subgroup_panels.png").exists()

    # xlsx must carry 4 sheets in US2+US4 wiring.
    wb = openpyxl.load_workbook(target / "결합분석.xlsx", read_only=True)
    assert wb.sheetnames == [
        "상관매트릭스",
        "회귀결과",
        "군집비교",
        "부분군비교",
    ]
