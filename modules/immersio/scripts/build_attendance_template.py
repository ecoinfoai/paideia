"""Build modules/immersio/templates/attendance.xlsx.

The output is the canonical attendance template referenced by spec FR-021
and contracts/attendance_template.md. Users copy the file into their
``data/bronze/출석/출석부.xlsx`` location and only fill in data rows.

Run:
    uv run python modules/immersio/scripts/build_attendance_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER: tuple[str, ...] = (
    ("학번", "이름") + tuple(f"W{week:02d}" for week in range(1, 17)) + ("비고",)
)


def build(target: Path, *, max_rows: int = 200) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "출석"
    sheet.append(list(HEADER))

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_index in range(1, len(HEADER) + 1):
        cell = sheet.cell(row=1, column=col_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 12
    for week_index in range(3, 19):
        sheet.column_dimensions[get_column_letter(week_index)].width = 6
    sheet.column_dimensions[get_column_letter(len(HEADER))].width = 24

    weekly_dv = DataValidation(
        type="list",
        formula1='"O,X,L,E"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="잘못된 출석 코드",
        error="허용 어휘: O(출석)·X(결석)·L(지각)·E(공결) 또는 빈 셀.",
    )
    weekly_dv.add(f"C2:R{max_rows}")
    sheet.add_data_validation(weekly_dv)

    student_id_dv = DataValidation(
        type="textLength",
        operator="equal",
        formula1=10,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="학번 길이",
        error="학번은 10자리 텍스트여야 합니다.",
    )
    student_id_dv.add(f"A2:A{max_rows}")
    sheet.add_data_validation(student_id_dv)

    sheet["A1"].comment = Comment(
        "헤더(학번·이름·W01..W16·비고)는 변경하지 마세요. "
        "출석 코드: O 출석 / X 결석 / L 지각 / E 공결.",
        author="immersio",
    )

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)


if __name__ == "__main__":
    out = Path(__file__).resolve().parents[1] / "templates" / "attendance.xlsx"
    build(out)
    print(f"wrote {out}")
