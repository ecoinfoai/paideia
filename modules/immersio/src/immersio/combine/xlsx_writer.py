"""4-sheet xlsx workbook writer — US1 sheets `상관매트릭스` + `회귀결과` (T032).

FR-001 / FR-002 + research §R6 #4 + §R13 vector #5 (`<dcterms:modified>`
epoch pin + zip entry date_time fix). Phase 1+2 의 public
``rewrite_modified_in_zip`` 을 직접 후처리로 호출하여 byte-identical
재실행 보장.

Public:
- :func:`write_us1_xlsx(correlation_cells, regression_coefs,
  regression_fit, out_path)` — 2-sheet 워크북 (US2/US4 sheets 는 후속
  T041/T055 가 추가).

Sheets:
- `상관매트릭스`: header (8 columns) + 1 row per correlation cell
- `회귀결과`: 3-row fit summary block + blank row + coef table (header +
  8 axis rows + warnings sub-block)

Determinism:
- 행 정렬: caller 가 결정 (correlation 모듈은 STANDARD_AXIS_KEYS ×
  metric alphabetic; regression 은 STANDARD_AXIS_KEYS)
- ``<dcterms:modified>`` 후처리: Phase 1+2 ``rewrite_modified_in_zip``
  with epoch ``2000-01-01T00:00:00Z`` (research §R13 vector #5 정합)
"""

from __future__ import annotations

import datetime
from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from immersio.report.xlsx_writer import rewrite_modified_in_zip

from paideia_shared.schemas import (
    CorrelationCell,
    RegressionCoefficient,
    RegressionFitSummary,
)

# Determinism vector #5 — fixed epoch (matches Phase 1+2 conventions).
_EPOCH_MODIFIED = datetime.datetime(
    2000, 1, 1, 0, 0, 0, tzinfo=datetime.timezone.utc
)

_CORRELATION_HEADERS = (
    "axis_key",
    "exam_metric_key",
    "n",
    "pearson_r",
    "raw_p",
    "fdr_q",
    "significant",
    "unstable_n_lt_20",
)

_REGRESSION_HEADERS = (
    "axis_key",
    "coef",
    "std_err",
    "t_stat",
    "raw_p",
    "fdr_q",
    "ci_low_95",
    "ci_high_95",
    "beta_standardized",
    "vif",
    "multicollinearity_flag",
)


def _build_correlation_sheet(
    wb: Workbook, cells: Sequence[CorrelationCell]
) -> None:
    sheet = wb.create_sheet("상관매트릭스")
    sheet.append(list(_CORRELATION_HEADERS))
    for c in cells:
        sheet.append(
            [
                c.axis_key,
                c.exam_metric_key,
                c.n,
                c.pearson_r,
                c.raw_p,
                c.fdr_q,
                c.significant_after_correction,
                c.unstable_inference_flag,
            ]
        )
    # Modest column widths so headers are visible without auto-fit.
    for i, _ in enumerate(_CORRELATION_HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = 16


def _build_regression_sheet(
    wb: Workbook,
    coefs: Sequence[RegressionCoefficient],
    fit: RegressionFitSummary,
) -> None:
    sheet = wb.create_sheet("회귀결과")

    # Block 1 — fit summary (3 rows: header label, fields, values).
    sheet.append(["적합 지표"])
    sheet.append(
        [
            "n_complete_case",
            "n_dropped",
            "R²",
            "R²_adj",
            "F_stat",
            "F_pvalue",
            "regression_method",
            "multiple_comparison_method",
            "small_sample_warning",
        ]
    )
    sheet.append(
        [
            fit.n_complete_case,
            fit.n_dropped,
            fit.r2,
            fit.r2_adj,
            fit.f_stat,
            fit.f_pvalue,
            fit.regression_method,
            fit.multiple_comparison_method,
            fit.small_sample_warning,
        ]
    )
    sheet.append([])  # blank separator row

    # Block 2 — coefficient table.
    sheet.append(list(_REGRESSION_HEADERS))
    for c in coefs:
        sheet.append(
            [
                c.axis_key,
                c.coef,
                c.std_err,
                c.t_stat,
                c.raw_p,
                c.fdr_q,
                c.ci_low_95,
                c.ci_high_95,
                c.beta_standardized,
                c.vif,
                c.multicollinearity_flag,
            ]
        )

    for i, _ in enumerate(_REGRESSION_HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = 16


_CLUSTER_ROW_HEADERS = (
    "cluster_id",
    "cluster_label",
    "n",
    "mean",
    "std",
    "ci_low_95",
    "ci_high_95",
    "excluded_reason",
)

_CLUSTER_PAIRWISE_HEADERS = (
    "cluster_pair",
    "mean_diff",
    "raw_p",
    "fdr_q",
    "significant_after_correction",
)


def _build_cluster_sheet(
    wb: "Workbook",
    rows: object,
    header: object,
    pairwise: object,
) -> None:
    """T041 — sheet `군집비교` 3-block (rows + ANOVA header + posthoc)."""
    sheet = wb.create_sheet("군집비교")

    # Block 1 — per-cluster rows.
    sheet.append(["군집별 통계"])
    sheet.append(list(_CLUSTER_ROW_HEADERS))
    for r in rows:
        sheet.append(
            [
                r.cluster_id,
                r.cluster_label,
                r.n,
                r.mean,
                r.std,
                r.ci_low_95,
                r.ci_high_95,
                r.excluded_reason,
            ]
        )
    sheet.append([])

    # Block 2 — omnibus header.
    sheet.append(["검정 결과"])
    sheet.append(
        [
            "k_used",
            "test_used",
            "levene_p",
            "test_stat",
            "raw_p",
            "eta_squared",
            "omega_squared",
            "posthoc_test",
        ]
    )
    sheet.append(
        [
            header.k_used,
            header.test_used,
            header.levene_p,
            header.test_stat,
            header.raw_p,
            header.eta_squared,
            header.omega_squared,
            header.posthoc_test,
        ]
    )
    sheet.append([])

    # Block 3 — posthoc pairwise.
    sheet.append(["사후 비교"])
    sheet.append(list(_CLUSTER_PAIRWISE_HEADERS))
    for p in pairwise:
        sheet.append(
            [
                f"{p.cluster_pair[0]}-{p.cluster_pair[1]}",
                p.mean_diff,
                p.raw_p,
                p.fdr_q,
                p.significant_after_correction,
            ]
        )

    for i in range(1, len(_CLUSTER_ROW_HEADERS) + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 18


def _build_subgroup_sheet(
    wb: Workbook,
    rows: object,
    headers: object,
) -> None:
    """T055 — sheet `부분군비교` 4-meta sub-blocks."""
    sheet = wb.create_sheet("부분군비교")
    headers_by_meta = {h.meta_kind: h for h in headers}
    rows_by_meta: dict[str, list[object]] = {}
    for r in rows:
        rows_by_meta.setdefault(r.meta_kind, []).append(r)

    meta_order = ("section", "prior_biology", "occupation", "education")
    for meta_kind in meta_order:
        sheet.append([f"메타: {meta_kind}"])
        sheet.append(["meta_value", "n", "mean", "std", "excluded_reason"])
        for r in rows_by_meta.get(meta_kind, []):
            sheet.append([r.meta_value, r.n, r.mean, r.std, r.excluded_reason])
        h = headers_by_meta.get(meta_kind)
        if h is not None:
            sheet.append([])
            sheet.append(
                [
                    "test_used",
                    "levene_p",
                    "test_stat",
                    "raw_p",
                    "fdr_q",
                    "effect_size_kind",
                    "effect_size_value",
                    "n_categories_compared",
                ]
            )
            sheet.append(
                [
                    h.test_used,
                    h.levene_p,
                    h.test_stat,
                    h.raw_p,
                    h.fdr_q,
                    h.effect_size_kind,
                    h.effect_size_value,
                    h.n_categories_compared,
                ]
            )
        sheet.append([])

    for i in range(1, 9):
        sheet.column_dimensions[get_column_letter(i)].width = 18


def write_us1_xlsx(
    *,
    correlation_cells: Sequence[CorrelationCell],
    regression_coefs: Sequence[RegressionCoefficient],
    regression_fit: RegressionFitSummary,
    out_path: Path,
    cluster_rows: object | None = None,
    cluster_header: object | None = None,
    cluster_pairwise: object | None = None,
    subgroup_rows: object | None = None,
    subgroup_headers: object | None = None,
) -> None:
    """Write the combined-analysis workbook to ``out_path``.

    US1 partial mode (default): 2 sheets — `상관매트릭스` + `회귀결과`.
    US2 wiring (T041): supply ``cluster_rows`` + ``cluster_header`` (and
    optionally ``cluster_pairwise``) to add `군집비교` sheet 3-block.

    Args:
        correlation_cells: ``compute_correlation_matrix`` output.
        regression_coefs: ``compute_ols_regression`` 1st elem.
        regression_fit: ``compute_ols_regression`` 2nd elem.
        out_path: ``.xlsx`` destination. Parent dir auto-created.
        cluster_rows: ``compute_cluster_score_comparison`` 1st elem (US2).
        cluster_header: ``compute_cluster_score_comparison`` 2nd elem (US2).
        cluster_pairwise: ``compute_cluster_score_comparison`` 3rd elem
            (US2). Empty list permitted (k=1 fallback).

    Raises:
        ValueError: If correlation_cells / regression_coefs is empty.
    """
    if not correlation_cells:
        raise ValueError("write_us1_xlsx: correlation_cells is empty")
    if not regression_coefs:
        raise ValueError("write_us1_xlsx: regression_coefs is empty")

    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet that openpyxl creates so our sheets are
    # the only ones present (deterministic sheet order).
    if wb.active is not None and wb.active.title == "Sheet":
        wb.remove(wb.active)

    _build_correlation_sheet(wb, correlation_cells)
    _build_regression_sheet(wb, regression_coefs, regression_fit)

    if cluster_rows is not None and cluster_header is not None:
        _build_cluster_sheet(
            wb, cluster_rows, cluster_header, cluster_pairwise or []
        )

    if subgroup_rows is not None and subgroup_headers is not None:
        _build_subgroup_sheet(wb, subgroup_rows, subgroup_headers)

    # Determinism vector #5 — pin core.xml dcterms:created/modified to a
    # fixed epoch. openpyxl strips tzinfo internally, so we hand it a
    # naive UTC datetime sourced from _EPOCH_MODIFIED.
    _epoch_naive = _EPOCH_MODIFIED.replace(tzinfo=None)
    wb.properties.created = _epoch_naive
    wb.properties.modified = _epoch_naive

    wb.save(out_path)

    # Determinism vector #5 — repack zip to also pin <dcterms:modified>
    # (openpyxl rewrites it on save) and zip entry mtimes.
    rewrite_modified_in_zip(out_path, _EPOCH_MODIFIED)


__all__ = ["write_us1_xlsx"]
