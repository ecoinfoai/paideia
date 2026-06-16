"""Quality-analysis xlsx writer (T045, FR-001/002, R-01).

Spec 004 contracts/xlsx_sheets.md §1-§6 — six sheets in this build:

  1. 전체요약          (13 indicator/value rows)
  2. 1_히스토그램       (구간_시작 / 구간_끝 / 도수 / 누적 / 누적_백분율)
  3. 2_메타데이터통계   (metadata_kind / value / n / mean / sd / test_kind /
                        test_p_value / levene_p_value / note)
  4. 3_변별력          (per-item discrimination + 판정 + 메타)
  5. 4_정답률          (per-item correct rate + 메타 + 정답·최다오답)
  6. 5_오답분석        (per-item label + 무응답률 + 인접 distractor 등)

The 7th sheet (학생성적) lives behind the ``student_metrics`` kwarg
(Phase 4 / T052) — opt-in so Phase 3 callers preserve the original
6-sheet output.

Determinism (FR-023 / R-10):
* ``Workbook.properties.creator`` and ``lastModifiedBy`` pinned to a
  fixed string so legacy/build-host metadata never leaks in.
* ``created`` / ``modified`` parsed from the operator's
  ``generated_at_utc`` argument — that string is a single source of
  truth shared by xlsx Producer, pdf CreationDate, and png Software.
* openpyxl's ``save()`` overwrites ``modified`` with ``datetime.now()``.
  Earlier revisions tried a ``contextmanager`` monkey-patch on
  ``openpyxl.writer.excel.datetime.datetime`` but that bled into other
  test fixtures' openpyxl interactions during the same pytest session
  (xlsx bytes diverged when test_legacy_diff fixtures preceded
  test_two_runs_byte_identical). The current implementation rewrites
  the ``<dcterms:modified>`` element of ``docProps/core.xml`` *after*
  ``save()`` lands by repacking the zip — purely local file work, zero
  shared-namespace mutation. The repack uses ``ZipInfo`` with a fixed
  ``date_time=(1980,1,1,0,0,0)`` so the resulting bytes match
  openpyxl's own pin and the two-call byte-equal property holds even
  across heavy fixture interleaving.
"""

from __future__ import annotations

import datetime
import io
import re
import zipfile
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from paideia_shared.schemas import (
    HistogramBin,
    ItemStatistics,
    MetadataAggregate,
    StudentExamMetrics,
)

_PRODUCER = "paideia/immersio/0.1.0"
_FIXED_ZIP_DATE = (1980, 1, 1, 0, 0, 0)
_CORE_XML_PATH = "docProps/core.xml"
_MODIFIED_RE = re.compile(
    r"(<dcterms:modified[^>]*>)([^<]+)(</dcterms:modified>)",
    re.DOTALL,
)


def rewrite_modified_in_zip(xlsx_path: Path, when: datetime.datetime) -> None:
    """Repack ``xlsx_path`` with a pinned ``<dcterms:modified>`` value.

    openpyxl 의 save() 가 modified 를 datetime.now() 로 덮어씀. 본 함수는
    save() 후 호출되어:
      1. 원본 xlsx 의 모든 entry 를 메모리 buffer 에 unpack (date_time 보존
         불필요 — 어차피 새 archive 에서 fixed mtime 사용).
      2. ``docProps/core.xml`` 안의 ``<dcterms:modified>`` 텍스트만
         ``when.strftime("%Y-%m-%dT%H:%M:%SZ")`` 로 교체.
      3. 새 archive 를 fixed ZipInfo (date_time=1980-01-01) 로 다시 작성.
      4. 결과 bytes 로 ``xlsx_path`` 덮어쓰기 (atomic via .write_bytes).

    ``when`` 은 ISO 8601 UTC ('YYYY-MM-DDTHH:MM:SSZ') 형식으로 직렬화.
    """
    iso = when.strftime("%Y-%m-%dT%H:%M:%SZ")
    with zipfile.ZipFile(xlsx_path, "r") as src:
        members: list[tuple[str, bytes, int]] = []
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == _CORE_XML_PATH:
                text = data.decode("utf-8")
                text = _MODIFIED_RE.sub(rf"\g<1>{iso}\g<3>", text, count=1)
                data = text.encode("utf-8")
            members.append((info.filename, data, info.compress_type))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for name, data, compress_type in members:
            zi = zipfile.ZipInfo(filename=name, date_time=_FIXED_ZIP_DATE)
            # Preserve the openpyxl-chosen compression so the archive
            # behaves identically to the original.
            zi.compress_type = compress_type or zipfile.ZIP_DEFLATED
            dst.writestr(zi, data)
    xlsx_path.write_bytes(buf.getvalue())


def _parse_iso8601_utc(s: str) -> datetime.datetime:
    """Parse the manifest ISO8601 string into a tz-naive UTC datetime.

    openpyxl's ``properties.created`` setter strips tzinfo internally, so
    we hand it a naive datetime sourced from the explicit UTC ISO string.
    """
    if not isinstance(s, str) or not s:
        raise ValueError(f"generated_at_utc must be a non-empty string, got {s!r}")
    s_norm = s.replace("Z", "+00:00")
    try:
        dt = datetime.datetime.fromisoformat(s_norm)
    except ValueError as exc:
        raise ValueError(f"generated_at_utc is not a valid ISO8601 string: {s!r}") from exc
    if dt.tzinfo is not None:
        dt = dt.astimezone(datetime.UTC).replace(tzinfo=None)
    return dt


def _stamp_workbook(
    wb: Workbook, semester: str, course_name_kr: str, when: datetime.datetime
) -> None:
    wb.properties.creator = _PRODUCER
    wb.properties.lastModifiedBy = _PRODUCER
    wb.properties.created = when
    wb.properties.modified = when
    wb.properties.title = f"시험분석결과 — {semester} {course_name_kr}"


def _build_overall_sheet(wb: Workbook, rows: Sequence[Mapping[str, object]]) -> None:
    ws = wb.create_sheet("전체요약")
    bold = Font(bold=True)
    ws.cell(1, 1, "지표").font = bold
    ws.cell(1, 2, "값").font = bold
    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1, row["지표"])
        ws.cell(i, 2, row["값"])


def _build_histogram_sheet(wb: Workbook, bins: Sequence[HistogramBin]) -> None:
    ws = wb.create_sheet("1_히스토그램")
    bold = Font(bold=True)
    headers = ("구간_시작", "구간_끝", "도수", "누적", "누적_백분율")
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h).font = bold
    for i, b in enumerate(bins, start=2):
        ws.cell(i, 1, b.bin_start)
        ws.cell(i, 2, b.bin_end)
        ws.cell(i, 3, b.count)
        ws.cell(i, 4, b.cumulative)
        ws.cell(i, 5, b.cumulative_pct)


def _build_metadata_sheet(wb: Workbook, rows: Sequence[MetadataAggregate]) -> None:
    ws = wb.create_sheet("2_메타데이터통계")
    bold = Font(bold=True)
    headers = (
        "metadata_kind",
        "metadata_value",
        "n",
        "mean",
        "sd",
        "test_kind",
        "test_p_value",
        "levene_p_value",
        "note",
    )
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h).font = bold
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r.metadata_kind)
        ws.cell(i, 2, r.metadata_value)
        ws.cell(i, 3, r.n)
        ws.cell(i, 4, r.mean)
        ws.cell(i, 5, r.sd)
        ws.cell(i, 6, r.test_kind)
        ws.cell(i, 7, r.test_p_value)
        ws.cell(i, 8, r.levene_p_value)
        ws.cell(i, 9, r.note)


def _verdict_for_discrimination(d: float) -> str:
    if d < 0:
        return "역변별"
    if d >= 0.40:
        return "우수"
    if d >= 0.30:
        return "양호"
    if d >= 0.20:
        return "보통"
    if d >= 0.10:
        return "개선필요"
    return "약함"


def _build_discrimination_sheet(wb: Workbook, items: Sequence[ItemStatistics]) -> None:
    ws = wb.create_sheet("3_변별력")
    bold = Font(bold=True)
    headers = (
        "문항번호",
        "정답률",
        "변별력지수",
        "점-이연_상관",
        "변별력_판정",
        "챕터",
        "문제유형",
        "난이도",
    )
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h).font = bold
    for i, it in enumerate(sorted(items, key=lambda x: x.item_no), start=2):
        ws.cell(i, 1, it.item_no)
        ws.cell(i, 2, it.correct_rate)
        ws.cell(i, 3, it.discrimination_index)
        ws.cell(i, 4, it.point_biserial)
        ws.cell(i, 5, _verdict_for_discrimination(it.discrimination_index))
        ws.cell(i, 6, it.chapter)
        ws.cell(i, 7, it.item_type)
        ws.cell(i, 8, it.difficulty_level)


def _build_correct_rate_sheet(wb: Workbook, items: Sequence[ItemStatistics]) -> None:
    ws = wb.create_sheet("4_정답률")
    bold = Font(bold=True)
    headers = (
        "문항번호",
        "정답률(%)",
        "챕터",
        "문제유형",
        "출처",
        "난이도",
        "예상_난이도",
        "정답번호",
        "최다오답번호",
        "최다오답률(%)",
        "무응답(%)",
    )
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h).font = bold
    for i, it in enumerate(sorted(items, key=lambda x: x.item_no), start=2):
        ws.cell(i, 1, it.item_no)
        ws.cell(i, 2, round(it.correct_rate * 100, 2))
        ws.cell(i, 3, it.chapter)
        ws.cell(i, 4, it.item_type)
        ws.cell(i, 5, it.source)
        ws.cell(i, 6, it.difficulty_level)
        ws.cell(i, 7, it.expected_difficulty)
        ws.cell(i, 8, it.correct_answer)
        ws.cell(i, 9, it.top_distractor_no)
        ws.cell(
            i,
            10,
            round(it.top_distractor_rate * 100, 2) if it.top_distractor_rate is not None else None,
        )
        ws.cell(i, 11, round(it.omit_rate * 100, 2))


def _build_distractor_sheet(wb: Workbook, items: Sequence[ItemStatistics]) -> None:
    ws = wb.create_sheet("5_오답분석")
    bold = Font(bold=True)
    headers = (
        "문항번호",
        "정답률(%)",
        "변별력지수",
        "무응답률(%)",
        "최다오답번호",
        "최다오답률(%)",
        "인접_distractor",
        "distractor_label",
    )
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h).font = bold
    for i, it in enumerate(sorted(items, key=lambda x: x.item_no), start=2):
        ws.cell(i, 1, it.item_no)
        ws.cell(i, 2, round(it.correct_rate * 100, 2))
        cell_d = ws.cell(i, 3, it.discrimination_index)
        ws.cell(i, 4, round(it.omit_rate * 100, 2))
        ws.cell(i, 5, it.top_distractor_no)
        ws.cell(
            i,
            6,
            round(it.top_distractor_rate * 100, 2) if it.top_distractor_rate is not None else None,
        )
        ws.cell(i, 7, "yes" if it.is_top_distractor_adjacent else "no")
        cell_label = ws.cell(i, 8, it.distractor_label)
        # Per US4 (T059), 변별력 < 0 행은 굵게 강조.
        if it.discrimination_index < 0:
            for c in range(1, 9):
                ws.cell(i, c).font = bold
        # Otherwise leave the cell defaults; only the header row + this row
        # carry styling so the xlsx remains visually scannable.
        _ = cell_d, cell_label  # references retained for clarity


def _build_student_score_sheet(wb: Workbook, metrics: Sequence[StudentExamMetrics]) -> None:
    """Build the 7th sheet `학생성적` per contracts/xlsx_sheets.md §7."""
    ws = wb.create_sheet("학생성적")
    bold = Font(bold=True)

    fixed_headers = [
        "학번",
        "이름",
        "분반",
        "응시여부",
        "총점",
        "100점환산",
        "분반_백분위",
        "전체_백분위",
        "z_score",
    ]

    # Collect dynamic header lists across all takers so absent students
    # also see the columns (empty cells).
    chapter_keys: list[str] = []
    source_keys: list[str] = []
    difficulty_keys: list[int] = []
    expected_keys: list[str] = []
    item_type_keys: list[str] = []
    for m in metrics:
        for k in m.chapter_correct_rates:
            if k not in chapter_keys:
                chapter_keys.append(k)
        for k in m.source_correct_rates:
            if k not in source_keys:
                source_keys.append(k)
        for k in m.difficulty_correct_rates:
            if k not in difficulty_keys:
                difficulty_keys.append(k)
        for k in m.expected_difficulty_correct_rates:
            if k not in expected_keys:
                expected_keys.append(k)
        for k in m.item_type_correct_rates:
            if k not in item_type_keys:
                item_type_keys.append(k)

    # Sort dynamic keys so output is deterministic across runs.
    chapter_keys.sort()
    source_keys.sort()
    difficulty_keys.sort()
    expected_keys.sort()
    item_type_keys.sort()

    dynamic_headers: list[str] = []
    dynamic_headers.extend(f"챕터_{k}" for k in chapter_keys)
    dynamic_headers.extend(f"출처_{k}" for k in source_keys)
    dynamic_headers.extend(f"난이도_{k}" for k in difficulty_keys)
    dynamic_headers.extend(f"예상_{k}" for k in expected_keys)
    dynamic_headers.extend(f"유형_{k}" for k in item_type_keys)
    dynamic_headers.append("관심챕터_본인정답률")
    dynamic_headers.append("비호감챕터_본인정답률")

    all_headers = fixed_headers + dynamic_headers
    for c, h in enumerate(all_headers, start=1):
        ws.cell(1, c, h).font = bold

    # Sort metrics by student_id for deterministic row order.
    sorted_metrics = sorted(metrics, key=lambda m: m.student_id)
    for i, m in enumerate(sorted_metrics, start=2):
        ws.cell(i, 1, m.student_id)
        ws.cell(i, 2, m.name_kr)
        ws.cell(i, 3, m.section)
        ws.cell(i, 4, "응시" if m.exam_taken else "결시")
        if m.exam_taken:
            ws.cell(i, 5, m.total_score)
            ws.cell(i, 6, m.score_percent)
            ws.cell(i, 7, m.section_percentile)
            ws.cell(i, 8, m.cohort_percentile)
            ws.cell(i, 9, m.z_score)
        # else: score columns remain None (empty cells) per FR-013.

        col_cursor = 10
        for k in chapter_keys:
            ws.cell(i, col_cursor, m.chapter_correct_rates.get(k))
            col_cursor += 1
        for k in source_keys:
            ws.cell(i, col_cursor, m.source_correct_rates.get(k))
            col_cursor += 1
        for k in difficulty_keys:
            ws.cell(i, col_cursor, m.difficulty_correct_rates.get(k))
            col_cursor += 1
        for k in expected_keys:
            ws.cell(i, col_cursor, m.expected_difficulty_correct_rates.get(k))
            col_cursor += 1
        for k in item_type_keys:
            ws.cell(i, col_cursor, m.item_type_correct_rates.get(k))
            col_cursor += 1
        ws.cell(i, col_cursor, m.interest_chapters_correct_rate)
        col_cursor += 1
        ws.cell(i, col_cursor, m.aversion_chapters_correct_rate)


def write_analysis_xlsx(
    *,
    output_path: Path,
    overall_rows: Sequence[Mapping[str, object]],
    histogram_bins: Sequence[HistogramBin],
    metadata_rows: Sequence[MetadataAggregate],
    item_stats: Iterable[ItemStatistics],
    semester: str,
    course_name_kr: str,
    generated_at_utc: str,
    student_metrics: Iterable[StudentExamMetrics] | None = None,
) -> None:
    """Write ``시험분석결과.xlsx`` to ``output_path``.

    Six sheets always written. When ``student_metrics`` is supplied the
    7th sheet ``학생성적`` is appended (Phase 4 / T052 join path). Phase
    3 callers may pass ``None`` to skip that sheet — backwards
    compatibility for the MVP without student-level join.

    Args:
        output_path: Destination ``.xlsx`` path. Parent directory must
            exist; the function refuses to ``mkdir`` (fail-fast).
        overall_rows: 13-row payload from
            ``analysis.overall_summary.compute_overall_summary``.
        histogram_bins: List of ``HistogramBin`` from
            ``analysis.histogram.compute_score_histogram``.
        metadata_rows: List of ``MetadataAggregate`` from
            ``analysis.metadata_stats.compute_metadata_aggregates``.
        item_stats: Iterable of ``ItemStatistics`` from
            ``analysis.item_stats.compute_item_statistics``.
        student_metrics: Optional iterable of ``StudentExamMetrics`` —
            when supplied, the 7th sheet ``학생성적`` is added with
            ExamItem metadata columns spread per FR-013/018.
        semester: e.g. ``"2026-1"`` (used in workbook title).
        course_name_kr: e.g. ``"인체구조와기능"`` (used in workbook title).
        generated_at_utc: ISO8601 UTC timestamp pinning Workbook
            ``created``/``modified`` for byte-identical determinism.

    Raises:
        ValueError: When required inputs are empty / malformed.
        FileNotFoundError: When ``output_path.parent`` does not exist.
    """
    if not overall_rows:
        raise ValueError("write_analysis_xlsx: overall_rows is empty")
    items = list(item_stats)
    if not items:
        raise ValueError("write_analysis_xlsx: item_stats is empty")
    metrics_list: list[StudentExamMetrics] | None
    if student_metrics is None:
        metrics_list = None
    else:
        metrics_list = list(student_metrics)
        if not metrics_list:
            raise ValueError(
                "write_analysis_xlsx: student_metrics is an empty iterable; "
                "pass None to skip the 학생성적 sheet"
            )
    output_path = Path(output_path)
    if not output_path.parent.is_dir():
        raise FileNotFoundError(
            f"write_analysis_xlsx: parent directory missing: {output_path.parent}"
        )

    when = _parse_iso8601_utc(generated_at_utc)

    wb = Workbook()
    # Workbook() seeds a default 'Sheet' which we drop so the tab order
    # mirrors contracts/xlsx_sheets.md exactly.
    default = wb.active
    wb.remove(default)

    _build_overall_sheet(wb, overall_rows)
    _build_histogram_sheet(wb, histogram_bins)
    _build_metadata_sheet(wb, metadata_rows)
    _build_discrimination_sheet(wb, items)
    _build_correct_rate_sheet(wb, items)
    _build_distractor_sheet(wb, items)
    if metrics_list is not None:
        _build_student_score_sheet(wb, metrics_list)

    _stamp_workbook(wb, semester, course_name_kr, when)
    wb.save(output_path)
    # Rewrite docProps/core.xml's <dcterms:modified> to the pinned
    # timestamp so two consecutive saves on the same input produce
    # byte-identical files (FR-023). See module docstring for why we
    # repack instead of monkey-patching openpyxl.writer.excel.
    rewrite_modified_in_zip(output_path, when)


__all__ = ["write_analysis_xlsx", "rewrite_modified_in_zip"]
