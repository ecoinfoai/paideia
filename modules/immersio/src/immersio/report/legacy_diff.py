"""Legacy ↔ immersio xlsx cell-level diff (T055, FR-003/030/031, R-06).

Compares ``data/silver/legacy/중간고사_분석결과.xlsx`` against
``data/gold/immersio/{semester}-{course}/시험분석결과.xlsx`` cell-by-cell
and emits a deterministic Markdown report that the operator can take to
the academic department review meeting.

Tolerance (Clarification Q1 / FR-030):
* numeric cells: ``|legacy − immersio| > 0.001`` → reported with diff +
  reason; ``≤ 0.001`` ignored (false positive 차단 / SC-004)
* text cells: exact-match mismatch → reported

Reason rules (FR-031): per-(sheet, header) if-then mapping covering
검정 종류 차이 (ANOVA → Welch) / 결시 분모 의심 / 무응답 처리 의심 / 반올림 /
default "운영자 검토 필요" fallback.

Determinism (FR-023): sheet index ASC → cell-coordinate (row, col) ASC
ordering, fixed ``f"{value:.4f}"`` numeric formatting, no system clock
read.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


class LegacyLoadError(RuntimeError):
    """Raised when a legacy / immersio xlsx fails to load.

    Wraps low-level openpyxl errors (``InvalidFileException``,
    ``BadZipFile``, ``ValueError``) so the orchestrator can map every
    load failure to a single CLI exit code (FR-033 → exit 5) without
    leaking implementation-detail traces to operators.
    """

_TOLERANCE = 0.001

_SHEETS_OF_INTEREST: tuple[str, ...] = (
    "전체요약",
    "1_히스토그램",
    "2_메타데이터통계",
    "3_변별력",
    "4_정답률",
    "5_오답분석",
    "학생성적",
)


def _column_letter(c: int) -> str:
    """Openpyxl-style column letter for 1-based index ``c``."""
    out = ""
    while c > 0:
        c, rem = divmod(c - 1, 26)
        out = chr(65 + rem) + out
    return out


def _cell_address(row: int, col: int) -> str:
    return f"{_column_letter(col)}{row}"


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _fmt(value: Any) -> str:
    """Stable string for the diff table — ``f"{x:.4f}"`` for numerics."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if _is_number(value):
        return f"{float(value):.4f}"
    return str(value)


def _detect_header_row(sheet: Any) -> int:
    """Locate the first row whose first cell is a non-blank string label.

    Legacy xlsx prefixes each sheet with a Korean title row + memo rows
    (see fixture ``legacy_xlsx_anchors.json``), so the actual header is
    rarely row 1. Scan the first 6 rows and return the first row that
    looks like a header (string in col 1, > 1 non-blank cells). When
    no candidate is found, fall back to row 1.
    """
    max_scan = min(sheet.max_row, 6)
    for r in range(1, max_scan + 1):
        first = sheet.cell(r, 1).value
        if not isinstance(first, str):
            continue
        non_blank = sum(
            1 for c in range(1, sheet.max_column + 1) if sheet.cell(r, c).value not in (None, "")
        )
        if non_blank >= 2:
            return r
    return 1


def _build_header_lookup(sheet: Any, header_row: int) -> dict[int, str]:
    """``{column_index: header_label}`` for the chosen header row."""
    return {
        c: str(sheet.cell(header_row, c).value)
        for c in range(1, sheet.max_column + 1)
        if sheet.cell(header_row, c).value not in (None, "")
    }


def _estimate_reason(
    *,
    legacy: Any,
    immersio: Any,
    sheet: str,
    header: str,
    diff: float | None,
) -> str:
    """FR-031 reason rules, evaluated in priority order."""
    if diff is None:
        return "텍스트 셀 불일치"
    abs_diff = abs(diff)
    if abs_diff < _TOLERANCE:
        return ""

    # 검정 종류 차이 (ANOVA → Welch ANOVA 자동 폴백)
    if sheet == "2_메타데이터통계" and ("p값" in header or "p_value" in header.lower()):
        return (
            "검정 종류 차이 (legacy: ANOVA / immersio: Welch ANOVA 자동 폴백 — Levene p < 0.05)"
        )

    # 결시·무응답 정의 차이
    if sheet == "전체요약" and header in ("평균", "표준편차", "값"):
        if abs_diff > 1.0:
            return (
                "결시·무응답 처리 정의 차이 의심 "
                "(legacy 측이 결시를 분모에 포함했을 가능성)"
            )

    if sheet == "3_변별력" and "변별력" in header:
        return "27% 분위 동점자 처리 차이 의심"

    if sheet == "4_정답률" and "정답률" in header:
        return (
            "무응답 처리 차이 의심 "
            "(legacy 측이 무응답을 분모에서 제외 가능성)"
        )

    if abs_diff < 0.01:
        return "반올림 차이 (자릿수 정의)"

    return "수치 정의 차이 미상 — 운영자 검토 필요"


def _build_cell_grid(sheet: Any) -> dict[tuple[int, int], Any]:
    """``{(row, col): value}`` snapshot of the sheet."""
    out: dict[tuple[int, int], Any] = {}
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(r, c).value
            if v is None or v == "":
                continue
            out[(r, c)] = v
    return out


def _diff_pair(
    sheet_name: str,
    legacy_sheet: Any,
    immersio_sheet: Any,
) -> tuple[list[dict], list[dict], int]:
    """Compare two sheets cell-by-cell.

    Returns ``(diffs, missing_in_immersio, total_compared)``.
    """
    legacy_grid = _build_cell_grid(legacy_sheet)
    immersio_grid = _build_cell_grid(immersio_sheet)

    legacy_header_row = _detect_header_row(legacy_sheet)
    legacy_headers = _build_header_lookup(legacy_sheet, legacy_header_row)

    diffs: list[dict] = []
    missing: list[dict] = []
    total_compared = 0

    all_coords = sorted(set(legacy_grid) | set(immersio_grid))
    for row, col in all_coords:
        legacy_value = legacy_grid.get((row, col))
        immersio_value = immersio_grid.get((row, col))
        if legacy_value is None and immersio_value is None:
            continue
        total_compared += 1

        # missing in immersio
        if immersio_value is None:
            header = legacy_headers.get(col, f"col{col}")
            missing.append(
                {
                    "sheet": sheet_name,
                    "cell": _cell_address(row, col),
                    "legacy": legacy_value,
                    "header": header,
                    "reason": "legacy 만 존재 — immersio 의도적 미재현",
                }
            )
            continue

        # numeric diff
        if _is_number(legacy_value) and _is_number(immersio_value):
            diff = float(immersio_value) - float(legacy_value)
            # Compare on the rounded magnitude (4 decimal places — same
            # precision as the rendered f"{value:.4f}" report) so
            # floating-point representation noise around the ±0.001
            # boundary does not surface as a false positive (SC-004).
            if round(abs(diff), 4) <= _TOLERANCE:
                continue
            header = legacy_headers.get(col, f"col{col}")
            reason = _estimate_reason(
                legacy=legacy_value,
                immersio=immersio_value,
                sheet=sheet_name,
                header=header,
                diff=diff,
            )
            diffs.append(
                {
                    "sheet": sheet_name,
                    "cell": _cell_address(row, col),
                    "header": header,
                    "kind": "numeric",
                    "legacy": legacy_value,
                    "immersio": immersio_value,
                    "diff": diff,
                    "reason": reason,
                }
            )
            continue

        # text exact-match
        if str(legacy_value) != str(immersio_value):
            header = legacy_headers.get(col, f"col{col}")
            reason = _estimate_reason(
                legacy=legacy_value,
                immersio=immersio_value,
                sheet=sheet_name,
                header=header,
                diff=None,
            )
            diffs.append(
                {
                    "sheet": sheet_name,
                    "cell": _cell_address(row, col),
                    "header": header,
                    "kind": "text",
                    "legacy": legacy_value,
                    "immersio": immersio_value,
                    "diff": None,
                    "reason": reason,
                }
            )

    return diffs, missing, total_compared


def _render_md(
    *,
    legacy_path: Path,
    immersio_path: Path,
    compared_at_utc: str,
    semester: str,
    course_slug: str,
    structure_lines: list[str],
    per_sheet_diffs: dict[str, list[dict]],
    missing_entries: list[dict],
    total_compared: int,
) -> str:
    total_diffs = sum(len(v) for v in per_sheet_diffs.values())
    lines: list[str] = [
        f"# legacy_diff.md — {semester} {course_slug}",
        "",
        f"**immersio xlsx**: {immersio_path}",
        f"**legacy xlsx**: {legacy_path}",
        f"**비교 시각**: {compared_at_utc}",
        f"**tolerance**: 절대 오차 ±{_TOLERANCE:.3f} (수치 셀)",
        (
            f"**총 비교 셀**: {total_compared} / 차이 발견: {total_diffs} / "
            f"immersio 채택: {total_diffs} / 미재현: {len(missing_entries)}"
        ),
        "",
        "---",
        "",
        "## 0. 구조 검증",
        "",
        "| 항목 | 결과 |",
        "|---|---|",
        *structure_lines,
        "",
        "---",
        "",
    ]

    for idx, sheet_name in enumerate(_SHEETS_OF_INTEREST, start=1):
        rows = per_sheet_diffs.get(sheet_name, [])
        lines.append(f"## {idx}. 시트 `{sheet_name}` — 차이 {len(rows)}건")
        lines.append("")
        if not rows:
            lines.append("(차이 없음)")
            lines.append("")
            lines.append("---")
            lines.append("")
            continue
        lines.append("| 셀 | 컬럼 | legacy | immersio | diff | 사유 추정 | 채택 |")
        lines.append("|---|---|---|---|---|---|---|")
        for r in rows:
            diff_str = _fmt(r["diff"]) if r["kind"] == "numeric" else "—"
            lines.append(
                f"| {r['cell']} | {r['header']} | {_fmt(r['legacy'])} "
                f"| {_fmt(r['immersio'])} | {diff_str} "
                f"| {r['reason']} | immersio |"
            )
        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append("## 8. immersio 미재현 셀 (legacy 만 존재)")
    lines.append("")
    if missing_entries:
        lines.append("| 시트 | 셀 | 컬럼 | legacy 값 | 미재현 사유 |")
        lines.append("|---|---|---|---|---|")
        for r in missing_entries:
            lines.append(
                f"| {r['sheet']} | {r['cell']} | {r['header']} "
                f"| {_fmt(r['legacy'])} | {r['reason']} |"
            )
    else:
        lines.append("(미재현 셀 없음)")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 9. 사유 추정 룰 (FR-031)")
    lines.append("")
    lines.append("- 검정 종류 차이: `2_메타데이터통계` 의 p값 컬럼 — legacy ANOVA / immersio Welch ANOVA 자동 폴백.")
    lines.append("- 결시·무응답 처리: `전체요약` 의 평균/표준편차 차이 > 1.0 → legacy 가 결시를 분모에 포함했을 가능성.")
    lines.append("- 27% 분위 동점자: `3_변별력` 의 변별력지수 — boundary 동점자 포함/제외 처리 차이.")
    lines.append("- 무응답 처리: `4_정답률` — legacy 가 무응답을 분모에서 제외 가능성.")
    lines.append("- 반올림: 그 외 |diff| < 0.01 — 자릿수 정의 차이.")
    lines.append("- 그 외: 운영자 검토 필요 (수치 정의 차이 미상).")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 10. 채택 결정 (모두 immersio)")
    lines.append("")
    lines.append(
        "FR-030 / 대체 원칙 4 에 따라 차이가 있는 모든 셀에서 immersio 수치를 "
        "채택. legacy 의 LLM 산 통계는 신뢰 근거가 없으므로 immersio 결정론적 "
        "수치를 정답으로 본다."
    )
    lines.append("")
    return "\n".join(lines)


def _load_xlsx_or_raise(path: Path, *, role: str) -> Any:
    """Load ``path`` via openpyxl, wrapping low-level errors as ``LegacyLoadError``.

    Args:
        path: Target xlsx path (must exist — caller should ``is_file()``
            check first).
        role: ``"legacy"`` or ``"immersio"`` — surfaced in the error
            message so the operator knows which input is corrupt.

    Returns:
        The loaded ``openpyxl.Workbook``.

    Raises:
        LegacyLoadError: When the file cannot be parsed as xlsx. Logs at
            error level before raising so the orchestrator's stderr
            stream carries the trace + actionable suggestion.
    """
    try:
        return load_workbook(path, data_only=True)
    except (InvalidFileException, BadZipFile, ValueError) as exc:
        logger.error(
            "legacy_diff: failed to load %s xlsx %s — %s: %s",
            role,
            path,
            type(exc).__name__,
            exc,
        )
        raise LegacyLoadError(
            f"{role} xlsx is corrupt or not a valid xlsx file: {path} "
            f"({type(exc).__name__}: {exc}). Verify the file with "
            f"`unzip -l {path}` (every xlsx is a zip archive); if "
            f"unreadable, restore from backup or re-export. CLI exit 5."
        ) from exc


def _structure_check(
    legacy_wb: Any, immersio_wb: Any
) -> tuple[list[str], list[str]]:
    """Return (markdown rows, list of sheet names common to both)."""
    legacy_sheets = list(legacy_wb.sheetnames)
    immersio_sheets = list(immersio_wb.sheetnames)
    rows = [
        f"| 시트 개수 (legacy/immersio) | {len(legacy_sheets)} / {len(immersio_sheets)} |",
    ]
    if legacy_sheets == immersio_sheets:
        rows.append("| 시트명·순서 | 일치 ✓ |")
    else:
        rows.append("| 시트명·순서 | 불일치 — 시트 비교는 교집합만 진행 |")
    rows.append("| 차트 anchor 좌표 | 본 spec 검증 범위 외 (R-13 fixture 참조) |")
    common = [s for s in legacy_sheets if s in immersio_sheets]
    return rows, common


def generate_legacy_diff(
    *,
    legacy_xlsx: Path,
    immersio_xlsx: Path,
    output_path: Path,
    compared_at_utc: str,
    semester: str,
    course_slug: str,
) -> None:
    """Compare ``legacy_xlsx`` against ``immersio_xlsx`` → write ``legacy_diff.md``.

    Args:
        legacy_xlsx: Path to the legacy reference xlsx (read-only).
        immersio_xlsx: Path to the immersio-generated xlsx (read-only).
        output_path: Target ``.md`` path. Parent directory must exist.
        compared_at_utc: ISO8601 UTC timestamp recorded in the report
            header (manifest single source — does NOT read the wall clock).
        semester: e.g. ``"2026-1"``.
        course_slug: e.g. ``"anatomy"``.

    Raises:
        FileNotFoundError: When either input xlsx is missing or the
            output parent directory does not exist.
        LegacyLoadError: When either input xlsx exists but is corrupt /
            not a valid xlsx (CLI exit 5 per FR-033). Wraps
            ``openpyxl`` 's ``InvalidFileException``, ``zipfile``'s
            ``BadZipFile``, and ``ValueError`` so callers don't have to
            reach into openpyxl internals.
    """
    legacy_xlsx = Path(legacy_xlsx)
    immersio_xlsx = Path(immersio_xlsx)
    output_path = Path(output_path)
    if not legacy_xlsx.is_file():
        raise FileNotFoundError(
            f"generate_legacy_diff: legacy xlsx not found: {legacy_xlsx}"
        )
    if not immersio_xlsx.is_file():
        raise FileNotFoundError(
            f"generate_legacy_diff: immersio xlsx not found: {immersio_xlsx}"
        )
    if not output_path.parent.is_dir():
        raise FileNotFoundError(
            f"generate_legacy_diff: parent directory missing: {output_path.parent}"
        )

    legacy_wb = _load_xlsx_or_raise(legacy_xlsx, role="legacy")
    immersio_wb = _load_xlsx_or_raise(immersio_xlsx, role="immersio")

    structure_rows, common_sheets = _structure_check(legacy_wb, immersio_wb)
    logger.info(
        "legacy_diff: comparing %d common sheets (legacy=%d, immersio=%d)",
        len(common_sheets),
        len(legacy_wb.sheetnames),
        len(immersio_wb.sheetnames),
    )

    per_sheet_diffs: dict[str, list[dict]] = {}
    missing_entries: list[dict] = []
    total_compared = 0

    for sheet_name in common_sheets:
        diffs, missing, compared = _diff_pair(
            sheet_name, legacy_wb[sheet_name], immersio_wb[sheet_name]
        )
        per_sheet_diffs[sheet_name] = diffs
        missing_entries.extend(missing)
        total_compared += compared

    md = _render_md(
        legacy_path=legacy_xlsx,
        immersio_path=immersio_xlsx,
        compared_at_utc=compared_at_utc,
        semester=semester,
        course_slug=course_slug,
        structure_lines=structure_rows,
        per_sheet_diffs=per_sheet_diffs,
        missing_entries=missing_entries,
        total_compared=total_compared,
    )
    output_path.write_text(md, encoding="utf-8")
    logger.info(
        "legacy_diff: wrote %s (compared=%d, diffs=%d, missing=%d)",
        output_path,
        total_compared,
        sum(len(v) for v in per_sheet_diffs.values()),
        len(missing_entries),
    )


__all__ = ["LegacyLoadError", "generate_legacy_diff"]
